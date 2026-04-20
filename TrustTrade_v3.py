import asyncio
import re
from dataclasses import dataclass
from telethon import TelegramClient, events
from telethon.tl.types import InputPeerChannel
from telethon.errors import SessionPasswordNeededError, FloodWaitError, PeerIdInvalidError, ChatIdInvalidError
from typing import Optional, Set, List, Tuple
import logging
from logging.handlers import RotatingFileHandler
from binance.client import Client as BinanceClient
from binance.enums import SIDE_BUY, SIDE_SELL, ORDER_TYPE_LIMIT, TIME_IN_FORCE_GTC, ORDER_TYPE_TAKE_PROFIT
from binance.exceptions import BinanceAPIException
import binance
from packaging import version
from functools import partial
import os
import sys
import time
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor
import math
import json
from flask import Flask, jsonify, render_template_string
import threading
from waitress import serve
from datetime import datetime, timezone
import atexit
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric import ed25519 
from cryptography.hazmat.backends import default_backend
from cryptography.exceptions import InvalidSignature
import numpy as np # Used for technical indicator calculation
import pandas as pd
from openpyxl import load_workbook
import aiohttp  # For Telegram alerts

# Fix for Windows: aiohttp requires SelectorEventLoop on Windows
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

load_dotenv()

# --- License Validation Start ---

LICENSE_FILE_NAME = "license.lic"
# Use the actual public key read from the file
LICENSE_PUBLIC_KEY = b"""-----BEGIN PUBLIC KEY-----
MCowBQYDK2VwAyEAhntNkn6cqJqG2mVICC+BLoX+Gss3gYH/kMaWTgxS6P8=
-----END PUBLIC KEY-----"""

class LicenseError(Exception):
    """Custom exception for license-related errors"""
    pass

def get_base_path():
    """Get the base path for the application, accounting for frozen executables."""
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        # Running in a PyInstaller bundle
        return sys._MEIPASS
    # Running as a script
    return os.path.dirname(os.path.abspath(__file__))

def validate_license() -> dict:
    """Validate the license file's signature and expiration date."""
    possible_paths = [
        os.path.join(get_base_path(), LICENSE_FILE_NAME),
        os.path.join(os.path.dirname(sys.executable), LICENSE_FILE_NAME), # For installed scripts/executables
        os.path.join(os.getcwd(), LICENSE_FILE_NAME), # Current working directory
        # Consider adding paths relative to the main script if needed
        os.path.join(os.path.dirname(os.path.abspath(__file__)), LICENSE_FILE_NAME) 
    ]
    # Remove duplicates
    possible_paths = list(dict.fromkeys(possible_paths))

    license_path = None
    for path in possible_paths:
        if os.path.exists(path):
            license_path = path
            logger.info(f"Found license file at: {license_path}")
            break

    if not license_path:
        raise LicenseError(
            f"License file '{LICENSE_FILE_NAME}' not found. Please ensure it exists in one of these locations:\n" +
            "\n".join(possible_paths)
        )

    try:
        with open(license_path, 'rb') as f:
            license_data = f.read()

        # Ed25519 signatures are typically 64 bytes
        if len(license_data) < 64:
            raise LicenseError("License file is corrupted or invalid (too short)")

        signature = license_data[:64]
        content = license_data[64:]

        public_key = serialization.load_pem_public_key(
            LICENSE_PUBLIC_KEY,
            backend=default_backend()
        )
        
        # Verify the signature
        public_key.verify(signature, content)

        # Decode and parse the license content
        license_info = json.loads(content.decode('utf-8'))

        # Validate required fields
        required_fields = ['issued_to', 'expiry', 'features', 'metadata'] # Removed 'trading_pairs'
        for field in required_fields:
            if field not in license_info:
                raise LicenseError(f"License is missing required field: '{field}'")
        
        if 'max_open_positions' not in license_info.get('metadata', {}) or \
           'max_leverage' not in license_info.get('metadata', {}):
             raise LicenseError("License metadata is missing 'max_open_positions' or 'max_leverage'")


        # Parse and validate expiry date
        expiry_date_str = license_info['expiry']
        try:
            # Handle potential timezone info (ISO format)
            if expiry_date_str.endswith('Z'):
                 expiry_date = datetime.fromisoformat(expiry_date_str.replace('Z', '+00:00')).astimezone(timezone.utc)
            else:
                 expiry_date = datetime.fromisoformat(expiry_date_str).astimezone(timezone.utc)

        except ValueError:
             raise LicenseError(f"Invalid expiry date format: {expiry_date_str}")


        current_date = datetime.now(timezone.utc)

        if expiry_date < current_date:
            raise LicenseError(f"License expired on {expiry_date.strftime('%Y-%m-%d %H:%M UTC')}")

        # Validate required features for the trading bot
        required_features = {"futures_trading"} # Example: Check if futures trading is allowed
        if not required_features.issubset(set(license_info.get('features', []))):
            missing = required_features - set(license_info.get('features', []))
            raise LicenseError(f"License is missing required features: {', '.join(missing)}")

        logger.info(f"License validated successfully for: {license_info['issued_to']}")
        logger.info(f"Expires on: {expiry_date.strftime('%Y-%m-%d %H:%M UTC')}")
        logger.info(f"Allowed features: {', '.join(license_info['features'])}")
        # Removed logging for allowed pairs
        logger.info(f"Max Positions: {license_info['metadata']['max_open_positions']}, Max Leverage: {license_info['metadata']['max_leverage']}")

        return license_info # Return the validated license info

    except InvalidSignature:
        raise LicenseError("License signature verification failed. The license file may be tampered with or invalid.")
    except json.JSONDecodeError:
        raise LicenseError("License content is not valid JSON.")
    except LicenseError: # Re-raise specific license errors
        raise
    except Exception as e:
        # Catch other potential errors during validation
        logger.error(f"An unexpected error occurred during license validation: {e}", exc_info=True)
        raise LicenseError(f"License validation failed due to an unexpected error: {e}")

# --- License Validation End ---


# Validate required environment variables
required_env_vars = [
    'API_ID', 'API_HASH', 'PHONE_NUMBER', 'CHANNEL_ID',
    'TRADING_API_KEY', 'TRADING_SECRET_KEY', 'MONITORING_API_KEY', 'MONITORING_SECRET_KEY', 'TRADING_MODE'
]
missing_vars = [var for var in required_env_vars if not os.getenv(var)]
if missing_vars:
    raise ValueError(f"Missing required environment variables: {', '.join(missing_vars)}")

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.DEBUG)
console_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logger.addHandler(console_handler)

log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'log.txt')
file_handler = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=5, encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logger.addHandler(file_handler)

# Utility function for EMA calculation
def calculate_ema(prices, period):
    if not prices or len(prices) < period:
        return None
    # Use pandas-like multiplier for standard EMA
    k = 2 / (period + 1)
    ema = [0.0] * len(prices)
    # Initialize with SMA for the first period
    ema[period-1] = sum(prices[:period]) / period 
    for i in range(period, len(prices)):
        ema[i] = prices[i] * k + ema[i-1] * (1 - k)
    return ema[-1]

# Utility function for RSI calculation
def calculate_rsi(prices, period=14):
    if not prices or len(prices) < period + 1:
        return None
    
    # Calculate price changes
    changes = [prices[i] - prices[i-1] for i in range(1, len(prices))]
    
    # Calculate gains and losses
    gains = [c if c > 0 else 0 for c in changes]
    losses = [-c if c < 0 else 0 for c in changes]
    
    # Calculate initial average gain and loss (SMA for first 14 periods)
    avg_gain = sum(gains[:period]) / period
    avg_loss = sum(losses[:period]) / period
    
    # Calculate subsequent EMA-based average gain and loss
    for i in range(period, len(gains)):
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period
        
    # Calculate RS and RSI
    if avg_loss == 0:
        rs = 999.0 if avg_gain > 0 else 0.0
    else:
        rs = avg_gain / avg_loss
    
    rsi = 100 - (100 / (1 + rs))
    return rsi

# Utility function for MACD histogram calculation
def calculate_macd_histogram(closes, fast_period=12, slow_period=26, signal_period=9):
    if not closes or len(closes) < slow_period + signal_period:
        return None
        
    def _ema_list(prices, period):
        k = 2 / (period + 1)
        ema = [0.0] * len(prices)
        if len(prices) < period: return None
        ema[period-1] = sum(prices[:period]) / period
        for i in range(period, len(prices)):
            ema[i] = prices[i] * k + ema[i-1] * (1 - k)
        return ema[period-1:] # Return from the first valid point

    ema_fast_list = _ema_list(closes, fast_period)
    ema_slow_list = _ema_list(closes, slow_period)
    
    if not ema_fast_list or not ema_slow_list: return None
    
    # Ensure lists are aligned based on the slower period
    offset = slow_period - fast_period
    
    # Calculate MACD Line
    macd_line = [ema_fast_list[i + offset] - ema_slow_list[i] for i in range(len(ema_slow_list))]
    
    # Calculate Signal Line (EMA of MACD Line)
    signal_line_list = _ema_list(macd_line, signal_period)
    
    if not signal_line_list: return None
    
    # Align MACD and Signal Line for histogram calculation
    macd_aligned = macd_line[signal_period-1:]
    
    # Calculate MACD Histogram
    histogram = [macd_aligned[i] - signal_line_list[i] for i in range(len(macd_aligned))]
    
    # Check if the histogram is rising or falling (use the last two values)
    if len(histogram) >= 2:
        if histogram[-1] > histogram[-2]:
            return 'RISING', histogram[-1]
        elif histogram[-1] < histogram[-2]:
            return 'FALLING', histogram[-1]
        else:
            return 'FLAT', histogram[-1]
    return 'FLAT', histogram[-1] if histogram else 0.0

# --- STRATEGY 2 & 4: ATR Calculation for Dynamic Sizing and Adaptive SL ---
def calculate_atr(highs: list, lows: list, closes: list, period: int = 14) -> float:
    """
    Calculate Average True Range for volatility measurement.
    Used for dynamic position sizing and adaptive stop-loss.
    
    Args:
        highs: List of high prices
        lows: List of low prices
        closes: List of close prices
        period: ATR period (default 14)
    
    Returns:
        ATR value as float
    """
    if len(highs) < period + 1 or len(lows) < period + 1 or len(closes) < period + 1:
        return 0.0
    
    true_ranges = []
    for i in range(1, len(closes)):
        high_low = highs[i] - lows[i]
        high_close = abs(highs[i] - closes[i-1])
        low_close = abs(lows[i] - closes[i-1])
        true_ranges.append(max(high_low, high_close, low_close))
    
    if len(true_ranges) < period:
        return 0.0
    
    return sum(true_ranges[-period:]) / period

# --- PHASE 3: Telegram Alert Manager ---

class AlertManager:
    """Manages Telegram alerts with severity levels and throttling"""
    
    def __init__(self, bot_token: str = None, chat_id: str = None, alert_level: str = "ALL"):
        self.bot_token = bot_token
        self.chat_id = chat_id
        self.alert_level = alert_level.upper()
        self.enabled = bool(bot_token and chat_id)
        self.last_alert_time = {}  # For throttling
        self.throttle_seconds = 300  # 5 minutes throttle for warnings
        
        if self.enabled:
            logger.info(f"Telegram alerts enabled. Level: {self.alert_level}")
        else:
            logger.info("Telegram alerts disabled (no bot token or chat ID)")
    
    async def send_telegram(self, message: str) -> bool:
        """Send message via Telegram Bot API"""
        if not self.enabled:
            return False
        
        try:
            url = f"https://api.telegram.org/bot{self.bot_token}/sendMessage"
            data = {
                "chat_id": self.chat_id,
                "text": message,
                "parse_mode": "HTML",
                "disable_web_page_preview": True
            }
            
            # Create SSL context that doesn't verify certificates
            # This is needed for corporate proxies/antivirus that intercept HTTPS
            import ssl
            ssl_context = ssl.create_default_context()
            ssl_context.check_hostname = False
            ssl_context.verify_mode = ssl.CERT_NONE
            
            connector = aiohttp.TCPConnector(ssl=ssl_context)
            
            async with aiohttp.ClientSession(connector=connector) as session:
                async with session.post(url, json=data, timeout=aiohttp.ClientTimeout(total=10)) as response:
                    if response.status == 200:
                        return True
                    else:
                        error_text = await response.text()
                        logger.error(f"Telegram API error: {response.status} - {error_text}")
                        return False
                        
        except asyncio.TimeoutError:
            logger.error("Timeout sending Telegram alert")
            return False
        except Exception as e:
            logger.error(f"Error sending Telegram alert: {e}")
            return False
    
    async def send_alert(self, level: str, title: str, message: str, symbol: str = None) -> bool:
        """
        Send alert with severity level and throttling
        
        Args:
            level: "CRITICAL", "WARNING", or "INFO"
            title: Alert title
            message: Alert message
            symbol: Optional symbol for throttling key
        """
        if not self.enabled:
            return False
        
        # Check alert level filter
        if level == "INFO" and self.alert_level == "CRITICAL":
            return False
        if level == "WARNING" and self.alert_level == "CRITICAL":
            return False
        if level == "INFO" and self.alert_level == "WARNING":
            return False
        
        # Throttling (except for CRITICAL)
        if level != "CRITICAL" and symbol:
            throttle_key = f"{level}_{symbol}_{title}"
            now = time.time()
            if throttle_key in self.last_alert_time:
                if now - self.last_alert_time[throttle_key] < self.throttle_seconds:
                    logger.debug(f"Throttled alert: {throttle_key}")
                    return False
            self.last_alert_time[throttle_key] = now
        
        # Format message
        emoji = {"CRITICAL": "🚨", "WARNING": "⚠️", "INFO": "ℹ️"}.get(level, "📢")
        formatted_message = f"{emoji} <b>{title}</b>\n\n{message}\n\n<i>Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>"
        
        # Send
        success = await self.send_telegram(formatted_message)
        if success:
            logger.debug(f"Sent {level} alert: {title}")
        return success

# --- END PHASE 3: Alert Manager ---

@dataclass
class TradingSignal:
    coin: str
    direction: str
    entry_prices: list[float]
    targets: list[float]
    leverage: str
    is_take_profit: bool = False
    profit: float = 0.0
    timestamp: float = 0.0
    message: str = ""

class TelegramSignalParser:
    def __init__(self):
        # Patterns for the old signal format (kept for backward compatibility)
        self.old_patterns = {
            'coin': r'Coin pair: ([A-Z0-9]+)',
            'direction': r'Order: (buy|sell)',
        }
        
        # Patterns for the new signal format: 🔥#BEAT/USDT (Short📉, x20)🔥
        self.new_patterns = {
            'signal_header': r'#([A-Z0-9]+)/USDT\s*\(\s*(Long|Short)[^,]*,\s*x(\d+)\s*\)',
            'entry_price': r'Entry\s*-\s*([0-9.]+)',
            'tp_levels': r'([0-9.]+)\s*\(\d+%\s*of\s*profit\)',
            'tp_price': r'Price\s*-\s*([0-9.]+)',  # For TP signals
            'tp_profit': r'Profit\s*-\s*(\d+)%',   # Profit percentage
        }

    def _normalize_text(self, text: str) -> str:
        # Added new emojis from the signal format
        replacements = {
            '📌': '', '⭕️': '', '📈': '', '📉': '', '✴️': '', '⚠️': '',
            "'": "'", '-': '-', ',': ',', '"': '"', ' ': ' ',
            '🟢': '', '🔴': '', '⭐': '', '🚀': '', '💠': '',
            '🇱🇰': '', '🔥': '', '🔔': '',
            '✅': '', '⏰': '', '⚠': ''
        }
        for old, new in replacements.items():
            text = text.replace(old, new)
        # Keep the regex to remove any other non-ASCII characters
        return re.sub(r'[^\x00-\x7F]+', ' ', text)

    def parse_message(self, text: str, timestamp: float) -> Optional[TradingSignal]:
        try:
            # Normalize text to remove emojis and other special characters
            normalized_text = self._normalize_text(text)
            logger.info(f"Parsing normalized message: {normalized_text[:200]}...")
            
            # Try to parse new signal format first: #SYMBOL/USDT (Direction, xLeverage)
            signal_header_match = re.search(self.new_patterns['signal_header'], normalized_text, re.IGNORECASE)
            
            if signal_header_match:
                # New format detected
                symbol_name = signal_header_match.group(1).upper()
                direction_str = signal_header_match.group(2).upper()
                leverage_value = signal_header_match.group(3)
                
                coin = f"{symbol_name}USDT"
                direction = 'LONG' if direction_str == 'LONG' else 'SHORT'
                leverage = f"{leverage_value}X"
                
                # Extract entry price
                entry_match = re.search(self.new_patterns['entry_price'], normalized_text)
                entry_price = float(entry_match.group(1)) if entry_match else None
                
                if entry_price:
                    logger.info(f"New signal format detected: Coin={coin}, Direction={direction}, Leverage={leverage}, Entry={entry_price}")
                    
                    # Return signal with entry price for limit order
                    return TradingSignal(
                        coin=coin,
                        direction=direction,
                        entry_prices=[entry_price],  # Single entry price for limit order
                        targets=[],  # Will be calculated based on ROI from env
                        leverage=leverage,
                        is_take_profit=False,
                        profit=0.0,
                        timestamp=timestamp,
                        message=text
                    )
                else:
                    # Check if this is a TP signal (has Price and Profit but no Entry)
                    tp_price_match = re.search(self.new_patterns['tp_price'], normalized_text)
                    tp_profit_match = re.search(self.new_patterns['tp_profit'], normalized_text)
                    
                    if tp_price_match and tp_profit_match:
                        tp_price = float(tp_price_match.group(1))
                        profit_percent = float(tp_profit_match.group(1))
                        
                        logger.info(f"TP signal detected: Coin={coin}, Direction={direction}, Leverage={leverage}, Price={tp_price}, Profit={profit_percent}%")
                        
                        # Return TP signal
                        return TradingSignal(
                            coin=coin,
                            direction=direction,
                            entry_prices=[],  # No entry for TP signals
                            targets=[tp_price],  # TP price as target
                            leverage=leverage,
                            is_take_profit=True,  # Mark as TP signal
                            profit=profit_percent,
                            timestamp=timestamp,
                            message=text
                        )
                    else:
                        logger.warning(f"New format detected but no entry price or TP data found for {coin}")
                        return None
            
            # Try old format: Coin pair: X, Order: buy/sell
            coin_match = re.search(self.old_patterns['coin'], normalized_text, re.IGNORECASE)
            direction_match = re.search(self.old_patterns['direction'], normalized_text, re.IGNORECASE)

            if coin_match and direction_match:
                coin_name = coin_match.group(1).upper()
                # Remove .P suffix if present and add USDT only if not already present
                coin_name = coin_name.replace('.P', '').replace('.PERP', '')
                if not coin_name.endswith('USDT'):
                    coin = f"{coin_name}USDT"
                else:
                    coin = coin_name
                
                direction_raw = direction_match.group(1).upper()
                direction = 'LONG' if direction_raw == 'BUY' else 'SHORT'

                logger.info(f"Old signal format detected: Coin={coin}, Direction={direction}")

                # This is a market order signal (old format)
                return TradingSignal(
                    coin=coin,
                    direction=direction,
                    entry_prices=[],  # Empty for market order
                    targets=[],       # Empty, will be calculated based on ROI
                    leverage=f"{os.getenv('DEFAULT_LEVERAGE', 20)}X",
                    is_take_profit=False,
                    profit=0.0,
                    timestamp=timestamp,
                    message=text
                )
            else:
                logger.warning("Message does not match any known signal format.")
                return None

        except Exception as e:
            logger.error(f"Parsing error: {e}", exc_info=True)
            return None

class TelegramSignalBot:
    def __init__(self, api_id: int, api_hash: str, phone: str, channel_id: str, 
                 trading_api_key: str, trading_secret_key: str, monitoring_api_key: str, monitoring_secret_key: str, 
                 trading_mode: str):
        # Validate python-binance version for Algo Trading Service support (Dec 2025 migration)
        required_version = "1.0.33"
        try:
            current_version = binance.__version__
            if version.parse(current_version) < version.parse(required_version):
                error_msg = (
                    f"python-binance version {required_version} or higher required for Algo Trading support.\n"
                    f"Current version: {current_version}\n"
                    f"Please upgrade: pip install --upgrade python-binance"
                )
                logger.error(error_msg)
                raise RuntimeError(error_msg)
            logger.info(f"python-binance version {current_version} validated (>= {required_version} required)")
        except AttributeError:
            logger.warning("Could not determine python-binance version. Proceeding with caution.")
        
        self.client = TelegramClient(f'session_{phone}', api_id, api_hash)
        self.phone = phone
        self.channel_id = channel_id
        self.parser = TelegramSignalParser()
        self.channel = None
        self.trading_api_key = trading_api_key
        self.trading_secret_key = trading_secret_key
        self.monitoring_api_key = monitoring_api_key
        self.monitoring_secret_key = monitoring_secret_key
        self.trading_client = None
        self.monitoring_client = None
        self.trading_mode = trading_mode.lower()
        self.max_open_positions = int(os.getenv('MAX_OPEN_POSITIONS', 1))
        self.max_open_entry_orders = int(os.getenv('MAX_OPEN_ENTRY_ORDERS', 1))
        self.min_balance = float(os.getenv('MIN_BALANCE', 1.0))
        self.margin_threshold = float(os.getenv('MARGIN_THRESHOLD', 0.2))
        self.deduplication_window = int(os.getenv('DEDUPLICATION_WINDOW', 600))
        self.target_margin_per_trade = float(os.getenv('TARGET_MARGIN_PER_TRADE', 1.0))
        self.default_leverage = int(os.getenv('DEFAULT_LEVERAGE', 20))
        self.max_leverage = int(os.getenv('MAX_LEVERAGE', 20))
        self.sl_percentage = float(os.getenv('SL_PERCENTAGE', 0.07))
        self.trailing_sl_trigger = float(os.getenv('TRAILING_SL_TRIGGER', 0.05))
        self.max_total_notional = float(os.getenv('MAX_TOTAL_NOTIONAL', 41.0))
        
        # --- IMPROVEMENT 1: MINIMUM RISK:REWARD CHECK ---
        self.min_risk_reward = float(os.getenv('MIN_RISK_REWARD', 1.5)) # Minimum RR of 1:1.5
        
        # --- IMPROVEMENT 8: ADJUSTED ROI TARGETS (USED IN IMPROVED TP STRATEGY) ---
        # The document suggests TP1=0.5% & TP2=1.8-2.5%. We will use these for the ROI settings.
        # Note: These are ROI percentages, NOT price percentages.
        self.tp1_roi = float(os.getenv('TP1_ROI', 0.5))  # 50% ROI
        self.tp2_roi = float(os.getenv('TP2_ROI', 2.0))  # 200% ROI (using 2.0 to be closer to 1.8-2.5% recommendation)
        
        self.symbol_locks = {}
        # Semaphore to limit concurrent open-position slots
        # Use BoundedSemaphore so releasing more than acquired raises an error
        self.open_positions_semaphore = asyncio.BoundedSemaphore(self.max_open_positions)
        # Track which symbols currently hold a reserved slot (so we release correctly)
        self.symbols_holding_slots: set[str] = set()
        self.placed_order_ids: dict[str, Set[int]] = {}
        self.tp_orders: dict[str, dict[str, int]] = {}
        self.sl_orders: dict[str, int] = {}
        self.global_lock = asyncio.Lock()
        self.processed_signals = {}
        self.signal_queue = []
        self.executor = ThreadPoolExecutor(max_workers=5)
        self.signal_count_per_symbol: dict[str, int] = {}
        self.monitor_tasks: dict[str, asyncio.Task] = {}
        self.position_info: dict[str, dict] = {}
        self.leverage_set = {}
        self.blocked = {}
        self.valid_symbols = set()
        self.app = Flask(__name__)
        self.web_thread = None
        # State persistence
        self.state_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'state.json')
        self.autosave_task: Optional[asyncio.Task] = None
        self.autosave_interval = int(os.getenv('STATE_AUTOSAVE_INTERVAL', 30))  # seconds
        
        # --- API Rate Limiting Configuration ---
        self.api_request_delay = float(os.getenv('API_REQUEST_DELAY', 0.5))  # Minimum 0.5s between API requests
        self.max_requests_per_minute = int(os.getenv('MAX_REQUESTS_PER_MINUTE', 60))  # Max 60 requests per minute
        self.last_api_request_time: float = 0.0
        self.api_request_timestamps: List[float] = []  # Track request times for rate limiting

        # Risk management: Consecutive losses pause (2 losses = 4hr pause)
        self.max_consecutive_losses = int(os.getenv('MAX_CONSECUTIVE_LOSSES', 2))
        self.consecutive_loss_pause_hours = float(os.getenv('CONSECUTIVE_LOSS_PAUSE_HOURS', 4))
        self.consecutive_losses_count = 0
        self.last_loss_time: Optional[float] = None
        self.trading_paused_until: Optional[float] = None
        self.count_losses_on_manual_close = os.getenv('COUNT_LOSSES_ON_MANUAL_CLOSE', 'False').lower() == 'true'
        
        # Trade frequency filter: max trades per time window
        self.max_trades_per_hour = int(os.getenv('MAX_TRADES_PER_HOUR', 10))
        self.recent_trades: List[float] = []  # timestamps of trades opened
        
        # --- IMPROVEMENT 3: TIGHTEN VOLATILITY FILTER CONFIG ---
        self.max_atr_percent = float(os.getenv('MAX_ATR_PERCENT', 2.5))          # Was 5.0, now 2.5
        self.max_volatility_percent = float(os.getenv('MAX_VOLATILITY_PERCENT', 1.8)) # Was 3.0, now 1.8
        self.max_spread_percent = float(os.getenv('MAX_SPREAD_PERCENT', 0.05))   # Was 0.1, now 0.05
        self.min_funding_rate = float(os.getenv('MIN_FUNDING_RATE', -0.001))
        self.max_funding_rate = float(os.getenv('MAX_FUNDING_RATE', 0.001))

        # --- IMPROVEMENT 4 & 5: NEW FILTER CONFIGS ---
        self.max_candle_wick_percent = float(os.getenv('MAX_CANDLE_WICK_PERCENT', 1.0)) # 1% for wick filter
        self.candle_wick_skip_minutes = int(os.getenv('CANDLE_WICK_SKIP_MINUTES', 5)) # Skip trade for 5 minutes

        self.time_filter_buffer_minutes = int(os.getenv('TIME_FILTER_BUFFER_MINUTES', 5)) # 5 minutes before/after major candles/volume
        self.dead_liquidity_start_utc = int(os.getenv('DEAD_LIQUIDITY_START_UTC', 3)) # 3AM UTC
        self.dead_liquidity_end_utc = int(os.getenv('DEAD_LIQUIDITY_END_UTC', 5)) # 5AM UTC
        
        # --- IMPROVEMENT 6: COOLDOWN CONFIG ---
        self.trade_cooldown_seconds = int(os.getenv('TRADE_COOLDOWN_SECONDS', 300)) # 300s = 5 minutes
        self.last_trade_closed_time: Optional[float] = None

        # --- IMPROVEMENT 7: VOLUME CONFIRMATION CONFIG ---
        self.enable_volume_filter = os.getenv('ENABLE_VOLUME_FILTER', 'False').lower() == 'true'  # Disabled by default
        self.min_volume_increase_percent = float(os.getenv('MIN_VOLUME_INCREASE_PERCENT', 10.0)) # 10% higher than previous candle
        self.max_orderbook_spread_percent = float(os.getenv('MAX_ORDERBOOK_SPREAD_PERCENT', 0.1)) # 0.1% orderbook spread (used as secondary spread check)
        
        # --- FILTER ENABLE/DISABLE TOGGLES ---
        self.enable_risk_reward_filter = os.getenv('ENABLE_RISK_REWARD_FILTER', 'True').lower() == 'true'
        self.enable_volatility_filter = os.getenv('ENABLE_VOLATILITY_FILTER', 'True').lower() == 'true'
        self.enable_spread_filter = os.getenv('ENABLE_SPREAD_FILTER', 'True').lower() == 'true'
        self.enable_time_filter = os.getenv('ENABLE_TIME_FILTER', 'True').lower() == 'true'
        self.enable_trend_filter = os.getenv('ENABLE_TREND_FILTER', 'True').lower() == 'true'
        self.enable_candle_wick_filter = os.getenv('ENABLE_CANDLE_WICK_FILTER', 'True').lower() == 'true'
        self.enable_volume_spike_detection = os.getenv('ENABLE_VOLUME_SPIKE_DETECTION', 'True').lower() == 'true'
        
        # Technical Indicator Filter Settings (STRATEGY 1)
        self.enable_technical_indicator_filter = os.getenv('ENABLE_TECHNICAL_INDICATOR_FILTER', 'False').lower() == 'true'
        self.rsi_overbought = float(os.getenv('RSI_OVERBOUGHT', 70))
        self.rsi_oversold = float(os.getenv('RSI_OVERSOLD', 30))
        self.require_ema_alignment = os.getenv('REQUIRE_EMA_ALIGNMENT', 'True').lower() == 'true'
        
        # User's Simplified RSI Filter (alternative to technical_indicator_filter)
        self.enable_rsi_filter = os.getenv('ENABLE_RSI_FILTER', 'False').lower() == 'true'
        self.ema_period = int(os.getenv('EMA_PERIOD', 20))
        
        # Dynamic Position Sizing Settings (STRATEGY 2)
        self.enable_dynamic_sizing = os.getenv('ENABLE_DYNAMIC_SIZING', 'False').lower() == 'true'
        self.enable_atr_sizing = os.getenv('ENABLE_ATR_SIZING', 'False').lower() == 'true'
        self.max_risk_per_trade_percent = float(os.getenv('MAX_RISK_PER_TRADE_PERCENT', 1.0))
        
        # Adaptive SL/TP Settings (STRATEGY 4)
        self.enable_atr_sl = os.getenv('ENABLE_ATR_SL', 'False').lower() == 'true'
        self.atr_sl_multiplier = float(os.getenv('ATR_SL_MULTIPLIER', 1.5))
        self.breakeven_trigger_pct = float(os.getenv('BREAKEVEN_TRIGGER_PCT', 0.5))

        # Direction Validation Settings (NEW - Prevent wrong-direction signals)
        self.enable_direction_validation = os.getenv('ENABLE_DIRECTION_VALIDATION', 'False').lower() == 'true'
        self.direction_rsi_period = int(os.getenv('DIRECTION_RSI_PERIOD', 14))
        self.direction_ema_period = int(os.getenv('DIRECTION_EMA_PERIOD', 20))
        self.direction_macd_fast = int(os.getenv('DIRECTION_MACD_FAST', 12))
        self.direction_macd_slow = int(os.getenv('DIRECTION_MACD_SLOW', 26))
        self.direction_macd_signal = int(os.getenv('DIRECTION_MACD_SIGNAL', 9))
        self.direction_rsi_bullish_threshold = float(os.getenv('DIRECTION_RSI_BULLISH_THRESHOLD', 50))
        self.direction_rsi_bearish_threshold = float(os.getenv('DIRECTION_RSI_BEARISH_THRESHOLD', 50))
        self.direction_klines_limit = int(os.getenv('DIRECTION_KLINES_LIMIT', 100))
        self.direction_alert_on_skip = os.getenv('DIRECTION_ALERT_ON_SKIP', 'True').lower() == 'true'

        # Cache for last candle close/volume for filters
        self.candle_cache: dict[str, dict] = {}
        
        # --- Trading Journal ---
        self.journal_file = 'Trading_Journal.xlsx'
        self.journal_lock = asyncio.Lock()
        self._initialize_journal()
        
        # Asset Performance Tracking & Blacklisting (STRATEGY 3)
        self.symbol_performance = {}  # {symbol: {'wins': int, 'losses': int, 'total_pnl': float, 'last_updated': timestamp}}
        self.enable_symbol_blacklist = os.getenv('ENABLE_SYMBOL_BLACKLIST', 'False').lower() == 'true'
        self.min_symbol_win_rate = float(os.getenv('MIN_SYMBOL_WIN_RATE', 0.4))  # 40% minimum
        self.min_trades_for_blacklist = int(os.getenv('MIN_TRADES_FOR_BLACKLIST', 5))
        self.blacklist_duration_minutes = int(os.getenv('BLACKLIST_DURATION_MINUTES', 1440))  # 24 hours
        self.symbol_blacklist = {}  # {symbol: {'until': timestamp, 'reason': str}}



        # PHASE 3: Initialize Telegram Alert Manager
        alert_bot_token = os.getenv('ALERT_BOT_TOKEN')
        alert_chat_id = os.getenv('ALERT_CHAT_ID')
        alert_level = os.getenv('ALERT_LEVEL', 'ALL')  # CRITICAL, WARNING, or ALL
        self.alerts = AlertManager(alert_bot_token, alert_chat_id, alert_level)

        # Load persisted state if exists
        try:
            self._load_state()
        except Exception:
            logger.exception("Failed to load persisted state at startup")

    async def cleanup(self):
        """Clean up resources before shutting down without canceling orders."""
        # PHASE 3: Send shutdown notification
        try:
            await self.alerts.send_alert(
                "WARNING",
                "🛑 Bot Shutting Down",
                f"<b>AutoTrade Bot is stopping...</b>\n\n"
                f"<b>Status:</b>\n"
                f"• All pending orders remain open\n"
                f"• Positions remain active\n"
                f"• Monitoring tasks stopped\n\n"
                f"<b>⚠️ Manual monitoring required</b>\n\n"
                f"<i>Stopped at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>"
            )
            logger.info("Shutdown notification sent to Telegram")
        except Exception as e:
            logger.error(f"Failed to send shutdown notification: {e}")
        
        # Stop autosave task first
        try:
            await self._stop_autosave()
        except Exception:
            logger.exception("Error stopping autosave task")
        
        # CRITICAL FIX: Save final state BEFORE shutting down executor
        # This prevents "RuntimeError: cannot schedule new futures after shutdown"
        try:
            await self._save_state()
            logger.info("Saved state to disk during cleanup")
        except Exception:
            logger.exception("Failed to save state during cleanup")
        
        # Disconnect Telegram client
        try:
            await self.client.disconnect()
        except Exception:
            logger.exception("Error disconnecting Telegram client")
        
        # NOW shut down executor (after state is saved)
        logger.info("Shutting down ThreadPoolExecutor...")
        self.executor.shutdown(wait=True)
        
        # Web server cleanup
        if hasattr(self, 'web_thread') and self.web_thread is not None and self.web_thread.is_alive():
            logger.info("Shutting down web server...")
            # Flask app shutdown is handled by stopping the thread
        
        logger.info("Bot cleanup completed. Pending orders remain open.")

    async def initialize_binance_clients(self):
        for client_type, api_key, secret_key in [
            ('trading', self.trading_api_key, self.trading_secret_key),
            ('monitoring', self.monitoring_api_key, self.monitoring_secret_key)
        ]:
            if not api_key or not secret_key:
                logger.error(f"{client_type.capitalize()} Binance API key or secret key is missing.")
                raise ValueError(f"{client_type.capitalize()} Binance API key or secret key is missing.")
            if not (isinstance(api_key, str) and isinstance(secret_key, str) and len(api_key) > 10 and len(secret_key) > 10):
                logger.error(f"{client_type.capitalize()} Binance API key or secret key appears invalid (too short or incorrect type).")
                raise ValueError(f"{client_type.capitalize()} Binance API key or secret key is invalid.")
            
            try:
                client = BinanceClient(api_key, secret_key)
                await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(self.executor, client.get_account),
                    timeout=15
                )
                if client_type == 'trading':
                    self.trading_client = client
                else:
                    self.monitoring_client = client
                logger.info(f"Successfully authenticated with {client_type} Binance API.")
            except asyncio.TimeoutError:
                logger.error(f"{client_type.capitalize()} Binance API authentication timed out after 15 seconds.")
                raise
            except BinanceAPIException as e:
                logger.error(f"{client_type.capitalize()} Binance API authentication failed: {e}")
                raise
            except Exception as e:
                logger.error(f"Failed to initialize {client_type} Binance client: {e}")
                raise
        # After clients are initialized, start autosave task to persist state periodically
        try:
            await self._start_autosave()
        except Exception:
            logger.exception("Failed to start autosave task")
        # Reconcile persisted state with live exchange data (positions and open orders)
        try:
            await self._reconcile_state_with_exchange()
        except Exception:
            logger.exception("Failed to reconcile persisted state with exchange")
        
        # Load historical performance for blacklisting (STRATEGY 3)
        if self.enable_symbol_blacklist:
            try:
                asyncio.create_task(self._load_historical_performance())
            except Exception:
                logger.exception("Failed to start performance loading task")

    def _initialize_journal(self):
        """Initialize the trading journal file if it doesn't exist."""
        if not os.path.exists(self.journal_file):
            df = pd.DataFrame(columns=[
                'Symbol', 'Direction', 'Entry Price', 'Entry Time',
                'TP1 Filled Time', 'SL Moved to Entry Time', 'Close Time', 'Close Price', 'PNL', 'Status'
            ])
            df.to_excel(self.journal_file, index=False, sheet_name='Trades')
            logger.info(f"Initialized trading journal: {self.journal_file}")

    # --- STRATEGY 3: Asset Performance Tracking & Blacklisting ---
    async def _load_historical_performance(self):
        """Load historical trade performance from Trading_Journal.xlsx"""
        try:
            await asyncio.sleep(2)  # Wait for journal initialization
            
            if not os.path.exists(self.journal_file):
                logger.info("No journal file found, starting with empty performance tracking")
                return
            
            df = await asyncio.get_running_loop().run_in_executor(
                self.executor,
                pd.read_excel,
                self.journal_file
            )
            
            for symbol in df['Symbol'].unique():
                symbol_trades = df[df['Symbol'] == symbol]
                wins = len(symbol_trades[symbol_trades['Status'].str.contains('Win|TP', na=False, case=False)])
                losses = len(symbol_trades[symbol_trades['Status'].str.contains('Loss|SL', na=False, case=False)])
                total_pnl = symbol_trades['PNL'].sum() if 'PNL' in df.columns else 0.0
                
                self.symbol_performance[symbol] = {
                    'wins': wins,
                    'losses': losses,
                    'total_pnl': total_pnl,
                    'last_updated': time.time()
                }
            
            logger.info(f"Loaded performance data for {len(self.symbol_performance)} symbols from journal")
            
        except Exception as e:
            logger.error(f"Error loading historical performance: {e}")

    async def check_symbol_blacklist(self, symbol: str, direction: str) -> bool:
        """Check if symbol should be blacklisted based on performance"""
        if not self.enable_symbol_blacklist:
            return True
        
        # Check if already blacklisted
        if symbol in self.symbol_blacklist:
            if time.time() < self.symbol_blacklist[symbol]['until']:
                logger.warning(f"{symbol} is blacklisted until {datetime.fromtimestamp(self.symbol_blacklist[symbol]['until']).strftime('%Y-%m-%d %H:%M')} - Reason: {self.symbol_blacklist[symbol]['reason']}")
                return False
            else:
                # Blacklist expired
                del self.symbol_blacklist[symbol]
                logger.info(f"{symbol} blacklist expired, allowing trade")
        
        # Check performance
        if symbol not in self.symbol_performance:
            return True  # No history, allow trade
        
        perf = self.symbol_performance[symbol]
        total_trades = perf['wins'] + perf['losses']
        
        if total_trades < self.min_trades_for_blacklist:
            return True  # Not enough data
        
        win_rate = perf['wins'] / total_trades if total_trades > 0 else 0
        
        # Blacklist if win rate too low
        if win_rate < self.min_symbol_win_rate:
            blacklist_until = time.time() + (self.blacklist_duration_minutes * 60)
            self.symbol_blacklist[symbol] = {
                'until': blacklist_until,
                'reason': f"Low win rate: {win_rate:.1%} ({perf['wins']}/{total_trades} wins)"
            }
            logger.warning(f"{symbol} BLACKLISTED for {self.blacklist_duration_minutes} minutes - Win rate: {win_rate:.1%}")
            
            await self.alerts.send_alert(
                "WARNING",
                f"🚫 Symbol Blacklisted: {symbol}",
                f"<b>Symbol temporarily blocked</b>\n\n"
                f"• Win Rate: {win_rate:.1%} (below {self.min_symbol_win_rate:.1%})\n"
                f"• Record: {perf['wins']}W - {perf['losses']}L\n"
                f"• Total PNL: {perf['total_pnl']:.2f} USDT\n"
                f"• Duration: {self.blacklist_duration_minutes} minutes\n\n"
                f"ℹ️ Will re-evaluate after cooldown period"
            )
            return False
        
        return True

    async def update_journal(self, trade_data: dict):
        """Update the trading journal with new trade information."""
        async with self.journal_lock:
            try:
                # Use a function that can be run in a thread to avoid blocking
                def _write_to_excel():
                    # Load the existing workbook and select the active sheet
                    book = load_workbook(self.journal_file)
                    sheet = book['Trades']

                    symbol = trade_data.get('Symbol', '')
                    update_fields = trade_data.copy()
                    del update_fields['Symbol']

                    # Find the row for this symbol that is not closed
                    updated = False
                    for row_num in range(2, sheet.max_row + 1):  # Start from 2 to skip header
                        if sheet.cell(row=row_num, column=1).value == symbol and sheet.cell(row=row_num, column=8).value == 'Open':  # Status 'Open'
                            # Update the row with new data
                            for col, (key, cell_name) in enumerate([('Direction', 2), ('Entry Price', 3), ('Entry Time', 4), ('TP1 Filled Time', 5), ('SL Moved to Entry Time', 6), ('Close Time', 7), ('Status', 8)], start=2):
                                if key in update_fields:
                                    sheet.cell(row=row_num, column=col).value = update_fields[key]
                            updated = True
                            break

                    # If no existing row found, create a new one (for initial entry)
                    if not updated:
                        new_row = [
                            symbol,
                            trade_data.get('Direction', ''),
                            trade_data.get('Entry Price', ''),
                            trade_data.get('Entry Time', ''),
                            trade_data.get('TP1 Filled Time', ''),
                            trade_data.get('SL Moved to Entry Time', ''),
                            trade_data.get('Close Time', ''),
                            trade_data.get('Status', '')
                        ]
                        sheet.append(new_row)

                    # Save the workbook
                    book.save(self.journal_file)
                    logger.info(f"Updated trading journal for {symbol}")

                # Run the synchronous file I/O in a separate thread
                await asyncio.get_running_loop().run_in_executor(self.executor, _write_to_excel)

            except Exception as e:
                logger.error(f"Failed to update trading journal: {e}", exc_info=True)

    def start_web_server(self):
        """Start a Flask web server in a separate thread to monitor trades."""
        def run_flask():
            # Define the HTML template for the dashboard
            dashboard_template = """
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <meta http-equiv="refresh" content="10">
                <title>Trade Monitoring Dashboard</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 20px; }
                    h1 { color: #333; }
                    table { width: 100%; border-collapse: collapse; margin-top: 20px; }
                    th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                    th { background-color: #f2f2f2; }
                    tr:nth-child(even) { background-color: #f9f9f9; }
                    .no-data { color: #888; font-style: italic; }
                </style>
            </head>
            <body>
                <h1>Trade Monitoring Dashboard</h1>
                <h2>Current Positions</h2>
                {% if positions %}
                <table>
                    <tr>
                        <th>Symbol</th>
                        <th>Quantity</th>
                        <th>Average Entry Price</th>
                    </tr>
                    {% for pos in positions %}
                    <tr>
                        <td>{{ pos.symbol }}</td>
                        <td>{{ pos.quantity }}</td>
                        <td>{{ pos.avg_entry_price }}</td>
                    </tr>
                    {% endfor %}
                </table>
                {% else %}
                <p class="no-data">No open positions.</p>
                {% endif %}

                <h2>Open Orders</h2>
                {% if orders %}
                <table>
                    <tr>
                        <th>Symbol</th>
                        <th>Order ID</th>
                        <th>Type</th>
                    </tr>
                    {% for order in orders %}
                    <tr>
                        <td>{{ order.symbol }}</td>
                        <td>{{ order.order_id }}</td>
                        <td>{{ order.type }}</td>
                    </tr>
                    {% endfor %}
                </table>
                {% else %}
                <p class="no-data">No open orders.</p>
                {% endif %}
            </body>
            </html>
            """

            @self.app.route('/')
            def dashboard():
                # Prepare data for positions
                positions = []
                for symbol, info in self.position_info.items():
                    positions.append({
                        'symbol': symbol,
                        'quantity': info.get('total_qty', 0.0),
                        'avg_entry_price': info.get('avg_entry_price', 0.0)
                    })

                # Prepare data for orders (TP and SL)
                orders = []
                for symbol, tp_dict in self.tp_orders.items():
                    for tp_type, order_id in tp_dict.items():
                        orders.append({
                            'symbol': symbol,
                            'order_id': order_id,
                            'type': f"Take Profit ({tp_type})"
                        })
                for symbol, order_id in self.sl_orders.items():
                    orders.append({
                        'symbol': symbol,
                        'order_id': order_id,
                        'type': "Stop Loss"
                    })

                return render_template_string(dashboard_template, positions=positions, orders=orders)

            @self.app.route('/api/trades')
            def api_trades():
                positions = []
                for symbol, info in self.position_info.items():
                    positions.append({
                        'symbol': symbol,
                        'quantity': info.get('total_qty', 0.0),
                        'avg_entry_price': info.get('avg_entry_price', 0.0)
                    })

                orders = []
                for symbol, tp_dict in self.tp_orders.items():
                    for tp_type, order_id in tp_dict.items():
                        orders.append({
                            'symbol': symbol,
                            'order_id': order_id,
                            'type': f"Take Profit ({tp_type})"
                        })
                for symbol, order_id in self.sl_orders.items():
                    orders.append({
                        'symbol': symbol,
                        'order_id': order_id,
                        'type': "Stop Loss"
                    })

                return jsonify({'positions': positions, 'orders': orders})

            serve(self.app, host='0.0.0.0', port=5000, threads=8)

        self.web_thread = threading.Thread(target=run_flask, daemon=True)
        self.web_thread.start()

    async def connect(self):
        try:
            logger.info(f"Attempting to connect with phone: {self.phone}")
            await self.client.connect()
            if not await self.client.is_user_authorized():
                logger.info("User not authorized. Requesting SMS code...")
                sent_code = await self.client.send_code_request(self.phone)
                code = input("Please enter the code you received: ").strip()
                try:
                    await self.client.sign_in(self.phone, code)
                except SessionPasswordNeededError:
                    password = input("Two-factor authentication enabled. Enter password: ")
                    await self.client.sign_in(password=password)
                logger.info("Successfully signed in!")
            else:
                logger.info("User already authorized")
            logger.info("Connected to Telegram successfully!")
            await self.initialize_binance_clients()
            await self.cache_valid_symbols()
            self.start_web_server()
            logger.info("Web dashboard started at http://0.0.0.0:5000 (accessible from network)")
            
            # PHASE 2: Start position reconciliation task
            asyncio.create_task(self.reconcile_positions_periodic())
            logger.info("Position reconciliation task started (checks every 5 minutes)")
            
            # PHASE 3: Send startup test message to verify alerts are working
            await self.alerts.send_alert(
                "INFO",
                "🤖 Bot Started Successfully",
                f"<b>AutoTrade Bot is now running!</b>\n\n"
                f"<b>Status:</b>\n"
                f"• Trading Mode: {self.trading_mode.upper()}\n"
                f"• Max Positions: {self.max_open_positions}\n"
                f"• Max Leverage: {self.max_leverage}X\n"
                f"• Alert Level: {self.alerts.alert_level}\n\n"
                f"<b>✅ All systems operational</b>\n"
                f"<b>✅ Telegram alerts working</b>\n\n"
                f"<i>Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>"
            )
            logger.info("Startup test message sent to Telegram")
        except FloodWaitError as e:
            logger.error(f"Rate limit exceeded. Waiting {e.seconds} seconds...")
            await asyncio.sleep(e.seconds)
            raise
        except Exception as e:
            logger.error(f"Failed to connect: {e}")
            await self.client.disconnect()
            raise

    async def cache_valid_symbols(self):
        """Cache valid trading symbols from Binance Futures."""
        try:
            exchange_info = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_exchange_info),
                timeout=15
            )
            self.valid_symbols = {s['symbol'] for s in exchange_info['symbols'] if s['status'] == 'TRADING'}
            logger.info(f"Cached {len(self.valid_symbols)} valid trading symbols.")
        except asyncio.TimeoutError:
            logger.error("Timeout caching valid symbols after 15 seconds.")
            self.valid_symbols = set()
        except Exception as e:
            logger.error(f"Error caching valid symbols: {e}")
            self.valid_symbols = set()

    async def get_channel(self):
        try:
            logger.info(f"Attempting to resolve channel with identifier: {self.channel_id}")
            if self.channel_id.startswith('@'):
                logger.debug(f"Resolving public channel by username: {self.channel_id}")
                self.channel = await self.client.get_entity(self.channel_id)
            else:
                try:
                    channel_id, access_hash = self.channel_id.split(':')
                    channel_id = int(channel_id)
                    access_hash = int(access_hash)
                    logger.debug(f"Resolving channel with ID: {channel_id}, Access Hash: {access_hash}")
                    self.channel = InputPeerChannel(channel_id, access_hash)
                    self.channel = await self.client.get_entity(self.channel)
                except ValueError as e:
                    logger.error(f"Invalid CHANNEL_ID format: {self.channel_id}. Expected format 'channel_id:access_hash' or '@username'. Error: {e}")
                    raise
                except (PeerIdInvalidError, ChatIdInvalidError) as e:
                    logger.error(f"Failed to resolve channel ID: {channel_id}. Ensure the bot is a member of the channel and the ID/access hash are correct. Error: {e}")
                    raise
            full_channel = await self.client.get_entity(self.channel)
            channel_name = getattr(full_channel, 'title', 'Unknown')
            channel_id_resolved = getattr(full_channel, 'id', 'Unknown')
            logger.info(f"Channel resolved successfully: {channel_name} (ID: -100{channel_id_resolved})")
            return full_channel
        except Exception as e:
            logger.error(f"Error resolving channel: {e}")
            raise

    async def validate_symbol(self, symbol: str) -> bool:
        """Validate if the symbol is tradable on Binance Futures."""
        try:
            if symbol in self.valid_symbols:
                return True
            exchange_info = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_exchange_info),
                timeout=15
            )
            for s in exchange_info['symbols']:
                if s['symbol'] == symbol and s['status'] == 'TRADING':
                    self.valid_symbols.add(symbol)
                    return True
            logger.warning(f"Symbol {symbol} is not a valid or tradable pair on Binance Futures.")
            return False
        except asyncio.TimeoutError:
            logger.error(f"Timeout validating symbol {symbol} after 15 seconds.")
            return False
        except Exception as e:
            logger.error(f"Error validating symbol {symbol}: {e}")
            return False

    async def get_open_positions_count(self) -> int:
        try:
            positions = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_position_information),
                timeout=15
            )
            open_positions = sum(1 for pos in positions if float(pos['positionAmt']) != 0)
            logger.info(f"Current open positions: {open_positions}")
            return open_positions
        except asyncio.TimeoutError:
            logger.error("Timeout fetching open positions count after 15 seconds.")
            return 0
        except BinanceAPIException as e:
            action = await self._handle_api_exception(e, "fetching open positions count")
            if action == 'abort':
                logger.error("Aborting open positions count due to critical API error")
            return 0
        except Exception as e:
            logger.error(f"Error fetching open positions count: {e}")
            return 0

    async def cancel_open_entry_orders(self, symbol: str):
        try:
            open_orders = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, partial(self.trading_client.futures_get_open_orders, symbol=symbol)),
                timeout=15
            )
            entry_orders = [order for order in open_orders if order['reduceOnly'] is False]
            for order in entry_orders:
                order_id = order['orderId']
                await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(self.executor, partial(
                        self.trading_client.futures_cancel_order,
                        symbol=symbol,
                        orderId=order_id
                    )),
                    timeout=15
                )
                self.placed_order_ids[symbol].discard(order_id)
                logger.info(f"Canceled open entry order {order_id} for {symbol}")
            try:
                await self._save_state()
            except Exception:
                logger.exception("Failed to save state after canceling open entry orders")
        except asyncio.TimeoutError:
            logger.warning(f"Timeout canceling open entry orders for {symbol} after 15 seconds.")
        except BinanceAPIException as e:
            logger.warning(f"Failed to cancel open entry orders for {symbol}: {e}")
        except Exception as e:
            logger.error(f"Error canceling open entry orders for {symbol}: {e}")

    async def cancel_all_pending_orders(self, symbol: str):
        """Cancel all pending TP, SL, and entry orders for the symbol."""
        try:
            # Cancel remaining TP orders (e.g., TP1 if not filled)
            await self.cancel_tp_orders(symbol)

            # Cancel SL order if it exists and log precise result
            sl_result = await self.cancel_sl_order(symbol)
            if sl_result == 'canceled':
                logger.info(f"SL order canceled for {symbol}")
            elif sl_result == 'filled':
                logger.info(f"SL order for {symbol} already processed: FILLED")
            elif sl_result == 'not_found':
                logger.info(f"SL order for {symbol} not found (likely already processed)")
            elif sl_result == 'no_sl':
                logger.debug(f"No SL order recorded for {symbol}")
            else:
                logger.info(f"SL cancellation result for {symbol}: {sl_result}")
            
            # Cancel all open entry orders
            await self.cancel_open_entry_orders(symbol)
            logger.info(f"All entry orders canceled for {symbol}")
        except Exception as e:
            logger.error(f"Error while canceling all pending orders for {symbol}: {e}")

    async def listen_to_channel(self):
        while True:
            try:
                channel_entity = await self.get_channel()
                @self.client.on(events.NewMessage(chats=self.channel))
                async def message_handler(event):
                    message_timestamp = event.message.date.timestamp()
                    signal = self.parser.parse_message(event.raw_text, message_timestamp)
                    if signal:
                        signal_key = (signal.coin, tuple(signal.entry_prices))
                        dedup_key = (signal.coin, tuple(signal.entry_prices))
                        current_time = time.time()
                        if dedup_key in self.processed_signals:
                            last_timestamp, last_signal_key = self.processed_signals[dedup_key]
                            if last_signal_key == signal_key and (current_time - last_timestamp) < self.deduplication_window:
                                logger.info(f"Duplicate signal detected for {signal.coin} with entries {signal.entry_prices}. Skipping.")
                                return
                        self.processed_signals[dedup_key] = (current_time, signal_key)
                        self.processed_signals = {k: v for k, v in self.processed_signals.items() if current_time - v[0] < 3600}
                        logger.info(f"Valid signal detected:\n{signal}")
                        if signal.is_take_profit:
                            # Check for "Closed due to opposite direction" - this is the ONLY TP-type signal we care about
                            if "Closed due to opposite direction" in signal.message:
                                has_position = await self.has_symbol_position(signal.coin)
                                if not has_position:
                                    logger.info(f"Received 'Closed due to opposite direction' signal for {signal.coin}. No position open, canceling open entry orders.")
                                    await self.cancel_open_entry_orders(signal.coin)
                            else:
                                # Ignore all regular TP signals - bot has its own TP1/TP2 system
                                logger.info(f"Ignoring TP signal for {signal.coin} (profit {signal.profit}%). Bot uses own TP1/TP2 system.")
                                return  # Skip all processing for regular TP signals
                        else:
                            # NEW: Direction Validation Filter - Check if signal direction aligns with market trend
                            if not await self.validate_signal_direction(signal.coin, signal.direction):
                                logger.info(f"Skipping signal for {signal.coin} due to direction validation failure")
                                return  # Skip this signal
                            
                            success = await self.handle_signal(signal)
                            if not success:
                                self.signal_queue.append((signal, time.time()))
                                logger.info(f"Signal for {signal.coin} queued for retry due to limits.")

                    if self.signal_queue:
                        await self.retry_queued_signals()

                channel_name = getattr(channel_entity, 'title', 'Unknown')
                logger.info(f"Listening to channel {channel_name}...")
                await self.client.run_until_disconnected()
                logger.warning("Disconnected from Telegram. Reconnecting in 10 seconds...")
            except Exception as e:
                logger.error(f"Error in listen_to_channel: {e}")
            await asyncio.sleep(10)
            logger.info("Reconnecting to Telegram...")
            await self.connect()

    async def retry_queued_signals(self):
        current_time = time.time()
        retry_interval = 60
        max_queue_age = self.deduplication_window

        new_queue = []
        for signal, queue_time in self.signal_queue:
            if current_time - queue_time > max_queue_age:
                logger.info(f"Discarding old signal for {signal.coin} from queue (age: {current_time - queue_time:.0f}s).")
                continue

            logger.info(f"Retrying queued signal for {signal.coin}...")
            # --- IMPROVEMENT 6: ADD COOLDOWN CHECK ON RETRY ---
            if not self.check_trade_cooldown():
                logger.info(f"Trade cooldown active. Keeping signal for {signal.coin} in queue.")
                new_queue.append((signal, queue_time))
                continue

            success = await self.handle_signal(signal)
            if not success:
                new_queue.append((signal, queue_time))
                logger.info(f"Signal for {signal.coin} still exceeds limits. Keeping in queue.")

        self.signal_queue = new_queue

    async def get_open_entry_orders_count(self):
        try:
            open_orders = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_get_open_orders),
                timeout=15
            )
            entry_orders = [order for order in open_orders if order['reduceOnly'] is False]
            logger.info(f"Current open entry orders: {len(entry_orders)} - Details: {[order['orderId'] for order in entry_orders]}")
            return len(entry_orders)
        except asyncio.TimeoutError:
            logger.error("Timeout fetching open orders count after 15 seconds.")
            return 0
        except BinanceAPIException as e:
            action = await self._handle_api_exception(e, "fetching open entry orders count")
            if action == 'abort':
                logger.error("Aborting open entry orders count due to critical API error")
            return 0
        except Exception as e:
            logger.error(f"Error fetching open orders count: {e}")
            return 0

    async def has_symbol_position(self, symbol: str) -> bool:
        try:
            positions = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_position_information),
                timeout=15
            )
            has_position = any(pos['symbol'] == symbol and float(pos['positionAmt']) != 0 for pos in positions)
            logger.debug(f"Checked position for {symbol}: {'Yes' if has_position else 'No'}")
            return has_position
        except asyncio.TimeoutError:
            logger.error(f"Timeout checking position for {symbol} after 15 seconds.")
            return False
        except BinanceAPIException as e:
            action = await self._handle_api_exception(e, "checking position", symbol)
            if action == 'abort':
                logger.error(f"Aborting position check for {symbol} due to critical API error")
            return False
        except Exception as e:
            logger.error(f"Error checking position for {symbol}: {e}")
            return False

    async def _throttle_api_request(self):
        """Enforce rate limiting to prevent API bans.
        
        Ensures:
        1. Minimum delay between consecutive requests
        2. Maximum requests per minute limit
        """
        current_time = time.time()
        
        # Enforce minimum delay between requests
        time_since_last_request = current_time - self.last_api_request_time
        if time_since_last_request < self.api_request_delay:
            delay = self.api_request_delay - time_since_last_request
            await asyncio.sleep(delay)
            current_time = time.time()
        
        # Clean up old timestamps (older than 1 minute)
        cutoff_time = current_time - 60
        self.api_request_timestamps = [t for t in self.api_request_timestamps if t > cutoff_time]
        
        # Check if we've hit the per-minute limit
        if len(self.api_request_timestamps) >= self.max_requests_per_minute:
            # Wait until the oldest request is more than 1 minute old
            oldest_request = self.api_request_timestamps[0]
            wait_time = 60 - (current_time - oldest_request)
            if wait_time > 0:
                logger.warning(f"API rate limit approaching. Waiting {wait_time:.1f}s before next request.")
                await asyncio.sleep(wait_time)
                current_time = time.time()
        
        # Record this request
        self.last_api_request_time = current_time
        self.api_request_timestamps.append(current_time)

    async def _handle_api_exception(self, e: BinanceAPIException, context: str, symbol: str = None) -> str:
        """Centralized handler for BinanceAPIException errors.
        
        Args:
            e: The BinanceAPIException that was raised
            context: Description of what operation failed (e.g., "fetching position info")
            symbol: Optional symbol for context
            
        Returns:
            Action to take: 'retry', 'abort', or 'continue'
        """
        error_code = getattr(e, 'code', None)
        error_msg = str(e)
        symbol_str = f" for {symbol}" if symbol else ""
        
        # Permission/Authentication errors (CRITICAL)
        if error_code == -2015:
            logger.error(f"CRITICAL: API Permission Error{symbol_str} while {context}: {error_msg}")
            await self.alerts.send_alert(
                "CRITICAL",
                "API Permission Error",
                f"Monitoring API key lacks permissions while {context}{symbol_str}.\n"
                f"Error: {error_msg}\n\n"
                f"ACTION REQUIRED:\n"
                f"1. Check Binance API Management\n"
                f"2. Ensure MONITORING_API_KEY has 'Enable Futures' permission\n"
                f"3. Verify IP restrictions match bot's IP"
            )
            return 'abort'
        
        # Rate limit errors (WARNING - retry with backoff)
        elif error_code in [-1003, -1013]:
            logger.warning(f"Rate limit hit{symbol_str} while {context}: {error_msg}")
            await self.alerts.send_alert(
                "WARNING",
                "API Rate Limit",
                f"Rate limit exceeded while {context}{symbol_str}. Bot will retry with backoff."
            )
            return 'retry'
        
        # Timestamp sync errors (WARNING - retry)
        elif error_code == -1021:
            logger.warning(f"Timestamp sync error{symbol_str} while {context}: {error_msg}")
            return 'retry'
        
        # Order would immediately trigger (INFO - expected in some cases)
        elif error_code == -2022:
            logger.info(f"Order would immediately trigger{symbol_str} while {context}: {error_msg}")
            return 'continue'
        
        # Order not found (may be filled/cancelled)
        elif error_code == -2011:
            logger.info(f"Order not found{symbol_str} while {context}: {error_msg}")
            return 'continue'
        
        # Invalid symbol or other client errors
        elif error_code in [-1121, -1100]:
            logger.error(f"Invalid request{symbol_str} while {context}: {error_msg}")
            return 'abort'
        
        # Unknown/Other API errors
        else:
            logger.error(f"API Error (code {error_code}){symbol_str} while {context}: {error_msg}")
            return 'retry'

    async def get_position_quantity(self, symbol: str, direction: str) -> float:
        try:
            # Apply rate limiting before API call
            await self._throttle_api_request()
            
            positions = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_position_information),
                timeout=15
            )
            position = next((pos for pos in positions if pos['symbol'] == symbol), None)
            if not position:
                logger.warning(f"No position found for {symbol}")
                return 0.0
            position_amt = float(position['positionAmt'])
            if direction == SIDE_BUY and position_amt > 0:
                return position_amt
            elif direction == SIDE_SELL and position_amt < 0:
                return abs(position_amt)
            else:
                logger.warning(f"No matching position for {symbol} in direction {direction}")
                return 0.0
        except asyncio.TimeoutError:
            logger.error(f"Timeout fetching position quantity for {symbol} after 15 seconds.")
            return 0.0
        except BinanceAPIException as e:
            action = await self._handle_api_exception(e, "fetching position quantity", symbol)
            if action == 'abort':
                logger.error(f"Aborting position quantity fetch for {symbol} due to critical API error")
            return 0.0
        except Exception as e:
            logger.error(f"Error fetching position quantity for {symbol}: {e}")
            return 0.0


    async def close_existing_position_market(self, symbol: str, close_side: str, quantity: float, cancel_orders_first: bool = True) -> bool:
        """Close an existing position immediately using a market reduce-only order.

        Args:
            symbol: trading symbol (e.g., 'BTCUSDT')
            close_side: side to place the order (SIDE_BUY or SIDE_SELL)
            quantity: quantity to close (already positive)
            cancel_orders_first: if True, cancel all TP/SL orders before closing position (default: True)

        Returns:
            True if order was successfully placed (or already closed), False on failure.
        """
        try:
            if quantity <= 0:
                logger.info(f"No quantity to close for {symbol}")
                return True

            # CRITICAL FIX: Cancel all protective orders BEFORE placing market close order
            # This prevents race conditions where TP/SL orders are active while position is being closed
            if cancel_orders_first:
                logger.info(f"Canceling all protective orders for {symbol} before closing position")
                try:
                    await self.cancel_all_pending_orders(symbol)
                    logger.info(f"Successfully canceled all orders for {symbol} before position close")
                except Exception as e:
                    logger.warning(f"Error canceling orders before closing {symbol}: {e}. Proceeding with position close anyway.")

            # Attempt to place a market reduce-only order to close position
            logger.info(f"Placing market reduce-only order to close {symbol}: side={close_side}, qty={quantity}")
            # Use the trading client to send the order in a thread to avoid blocking
            close_order = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, partial(
                    self.trading_client.futures_create_order,
                    symbol=symbol,
                    side=close_side,
                    type='MARKET',
                    quantity=float(quantity),
                    reduceOnly=True
                )),
                timeout=15
            )
            order_id = close_order.get('orderId')
            logger.info(f"Close market order placed for {symbol}, orderId={order_id}")

            # Persist state and record placed order id for tracking
            try:
                if symbol not in self.placed_order_ids:
                    self.placed_order_ids[symbol] = set()
                if order_id:
                    self.placed_order_ids[symbol].add(order_id)
                await self._save_state()
            except Exception:
                logger.exception("Failed to update state after placing close market order")

            # --- IMPROVEMENT 6: Record trade closure time ---
            self.last_trade_closed_time = time.time()
            logger.info(f"Recorded last trade closed time: {self.last_trade_closed_time}")

            return True
        except asyncio.TimeoutError:
            logger.error(f"Timeout placing close market order for {symbol} after 15 seconds.")
            return False
        except BinanceAPIException as e:
            logger.error(f"Binance API error placing close market order for {symbol}: {e}")
            return False
        except Exception as e:
            logger.error(f"Unexpected error closing position for {symbol}: {e}")
            return False

    async def check_balance(self) -> tuple[bool, float]:
        try:
            account = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_account),
                timeout=15
            )
            available_balance = float(account['availableBalance'])
            logger.info(f"Available balance: {available_balance:.2f} USDT")
            if available_balance < self.min_balance:
                logger.warning(f"Available balance {available_balance:.2f} USDT below minimum {self.min_balance:.2f} USDT.")
                return False, available_balance
            return True, available_balance
        except asyncio.TimeoutError:
            logger.error("Timeout checking balance after 15 seconds.")
            return False, 0.0
        except Exception as e:
            logger.error(f"Error checking balance: {e}")
            return False, 0.0

    async def check_margin(self, symbol: str, leverage: int, quantity: float, current_price: float) -> tuple[bool, float]:
        try:
            account = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_account),
                timeout=15
            )
            available_balance = float(account['totalWalletBalance'])
            total_initial_margin = float(account['totalInitialMargin'])
            total_margin_balance = float(account['totalMarginBalance'])

            notional_value = quantity * current_price
            required_margin = notional_value / leverage

            logger.info(f"Margin check for {symbol}: Available balance: {available_balance:.2f}, Required margin: {required_margin:.2f} USDT")
            if available_balance < required_margin:
                logger.info(f"Insufficient balance for {symbol}. Required: {required_margin:.2f} USDT, Available: {available_balance:.2f} USDT")
                return False, required_margin
            new_total_margin = total_initial_margin + required_margin
            margin_ratio = new_total_margin / total_margin_balance
            if margin_ratio > (1 - self.margin_threshold):
                logger.info(f"Margin ratio too high for {symbol}. Ratio: {margin_ratio:.2%}, Threshold: {1 - self.margin_threshold:.2%}")
                return False, required_margin
            return True, required_margin
        except asyncio.TimeoutError:
            logger.error(f"Timeout checking margin for {symbol} after 15 seconds.")
            return False, 0.0
        except Exception as e:
            logger.error(f"Error checking margin for {symbol}: {e}")
            return False, 0.0

    async def check_trading_paused(self) -> bool:
        """Check if trading is paused due to consecutive losses."""
        if self.trading_paused_until is None:
            return False
        current_time = time.time()
        if current_time < self.trading_paused_until:
            remaining_secs = self.trading_paused_until - current_time
            logger.warning(f"Trading paused for {remaining_secs:.0f} more seconds due to consecutive losses")
            return True
        else:
            # Pause expired, reset counter
            self.trading_paused_until = None
            self.consecutive_losses_count = 0
            logger.info("Trading pause expired. Loss counter reset.")
            try:
                await self._save_state()
            except Exception:
                logger.exception("Failed to save state after pause expiry")
            return False

    def check_trade_frequency(self) -> bool:
        """Check if number of trades in last hour exceeds limit."""
        current_time = time.time()
        one_hour_ago = current_time - 3600
        
        # Remove trades older than 1 hour
        self.recent_trades = [ts for ts in self.recent_trades if ts > one_hour_ago]
        
        if len(self.recent_trades) >= self.max_trades_per_hour:
            logger.warning(f"Trade frequency limit reached: {len(self.recent_trades)} trades in last hour. Max: {self.max_trades_per_hour}")
            return False
        return True

    # --- IMPROVEMENT 6: ADD COOLDOWN CHECK ---
    def check_trade_cooldown(self) -> bool:
        """Check if enough time has passed since the last trade was closed."""
        if self.last_trade_closed_time is None:
            return True
        
        current_time = time.time()
        time_since_last_trade = current_time - self.last_trade_closed_time
        
        if time_since_last_trade < self.trade_cooldown_seconds:
            remaining = self.trade_cooldown_seconds - time_since_last_trade
            logger.warning(f"Trade cooldown active: {remaining:.0f}s remaining. Skipping signal.")
            return False
        
        return True

    def record_trade_opened(self):
        """Record that a new trade was opened (for frequency tracking)."""
        self.recent_trades.append(time.time())
        logger.info(f"Trade recorded. Total trades in last hour: {len(self.recent_trades)}")

    def record_trade_loss(self):
        """Record a losing trade and check if pause should be triggered."""
        self.consecutive_losses_count += 1
        self.last_loss_time = time.time()
        logger.warning(f"Loss recorded. Consecutive losses: {self.consecutive_losses_count}/{self.max_consecutive_losses}")
        
        if self.consecutive_losses_count >= self.max_consecutive_losses:
            self.trading_paused_until = time.time() + (self.consecutive_loss_pause_hours * 3600)
            logger.error(f"Max consecutive losses ({self.max_consecutive_losses}) reached. Pausing trading for {self.consecutive_loss_pause_hours} hours until {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(self.trading_paused_until))}")
            try:
                asyncio.create_task(self._save_state())
            except Exception:
                logger.exception("Failed to save state after loss pause trigger")

    def record_trade_win(self):
        """Record a winning trade and reset loss counter."""
        if self.consecutive_losses_count > 0:
            logger.info(f"Win recorded. Resetting loss counter from {self.consecutive_losses_count} to 0")
            self.consecutive_losses_count = 0
            try:
                asyncio.create_task(self._save_state())
            except Exception:
                logger.exception("Failed to save state after win")

    # --- NEW: Get Klines for Filter Calculation ---
    async def get_klines_data(self, symbol: str, interval: str, limit: int = 100):
        try:
            klines = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(
                    self.executor,
                    partial(self.monitoring_client.futures_klines, symbol=symbol, interval=interval, limit=limit)
                ),
                timeout=10
            )
            # Structure: [timestamp, open, high, low, close, volume, close_time, quote_asset_vol, num_trades, taker_buy_base_vol, taker_buy_quote_vol, ignore]
            data = {
                'open': np.array([float(k[1]) for k in klines]),
                'high': np.array([float(k[2]) for k in klines]),
                'low': np.array([float(k[3]) for k in klines]),
                'close': np.array([float(k[4]) for k in klines]),
                'volume': np.array([float(k[5]) for k in klines]),
                'close_time': np.array([k[6] for k in klines]),
                'num_trades': np.array([k[8] for k in klines])
            }
            return data
        except asyncio.TimeoutError:
            logger.error(f"Timeout fetching {interval} klines for {symbol}")
            return None
        except Exception as e:
            logger.error(f"Error fetching klines for {symbol}: {e}")
            return None

    # --- IMPROVEMENT 2: Trend Filter Check ---
    async def check_trend_filter(self, symbol: str, direction: str) -> bool:
        """Check for 50 EMA, RSI, and MACD trend confirmation."""
        # Use 1-hour klines for stable trend confirmation (50 EMA, RSI)
        klines_data = await self.get_klines_data(symbol, '1h', limit=50) # Need at least 50 for 50 EMA
        if not klines_data:
            logger.warning(f"Could not fetch klines for {symbol}. Skipping trend filter.")
            return True

        closes = klines_data['close'].tolist()
        
        # 1. 50 EMA
        ema_50 = calculate_ema(closes, 50)
        current_price = closes[-1]
        
        if ema_50 is None:
            logger.warning(f"Insufficient data for 50 EMA on {symbol}. Skipping trend filter.")
            return True

        trend_ok = False
        if direction == SIDE_BUY:
            # LONG: Price > 50 EMA
            if current_price > ema_50:
                logger.debug(f"{symbol} LONG trend OK: Price ({current_price:.4f}) > 50 EMA ({ema_50:.4f})")
                trend_ok = True
            else:
                logger.debug(f"{symbol} LONG trend FAIL: Price ({current_price:.4f}) < 50 EMA ({ema_50:.4f})")
                return False
        elif direction == SIDE_SELL:
            # SHORT: Price < 50 EMA
            if current_price < ema_50:
                logger.debug(f"{symbol} SHORT trend OK: Price ({current_price:.4f}) < 50 EMA ({ema_50:.4f})")
                trend_ok = True
            else:
                logger.debug(f"{symbol} SHORT trend FAIL: Price ({current_price:.4f}) > 50 EMA ({ema_50:.4f})")
                return False
        
        # 2. RSI (Use 14 period default)
        rsi = calculate_rsi(closes)
        if rsi is None:
            logger.warning(f"Insufficient data for RSI on {symbol}. Continuing.")
        else:
            if direction == SIDE_BUY:
                # LONG: RSI > 50
                if rsi > 50:
                    logger.debug(f"{symbol} LONG RSI OK: {rsi:.2f} > 50")
                else:
                    logger.debug(f"{symbol} LONG RSI FAIL: {rsi:.2f} < 50")
                    return False
            elif direction == SIDE_SELL:
                # SHORT: RSI < 50
                if rsi < 50:
                    logger.debug(f"{symbol} SHORT RSI OK: {rsi:.2f} < 50")
                else:
                    logger.debug(f"{symbol} SHORT RSI FAIL: {rsi:.2f} > 50")
                    return False
        
        # 3. MACD Histogram (Rising/Falling)
        macd_result = calculate_macd_histogram(closes)
        if macd_result is None:
            logger.warning(f"Insufficient data for MACD on {symbol}. Continuing.")
        else:
            macd_status, macd_value = macd_result
            if direction == SIDE_BUY:
                # LONG: MACD histogram rising
                if macd_status == 'RISING':
                    logger.debug(f"{symbol} LONG MACD OK: {macd_status}")
                else:
                    logger.debug(f"{symbol} LONG MACD FAIL: {macd_status}")
                    return False
            elif direction == SIDE_SELL:
                # SHORT: MACD histogram falling
                if macd_status == 'FALLING':
                    logger.debug(f"{symbol} SHORT MACD OK: {macd_status}")
                else:
                    logger.debug(f"{symbol} SHORT MACD FAIL: {macd_status}")
                    return False

        logger.info(f"{symbol} passed Trend Filter.")
        return True

    # --- IMPROVEMENT 7: Volume & Orderbook Filter Check ---
    async def check_volume_confirmation(self, symbol: str) -> bool:
        """Check if Volume > 10% higher than previous candle AND Orderbook Spread < 0.1%."""
        
        # 1. Volume Confirmation (Use 5-minute candles for recent activity)
        klines_data = await self.get_klines_data(symbol, '5m', limit=2)
        if klines_data is None or len(klines_data['volume']) < 2:
            logger.warning(f"Could not fetch 5m klines for volume check on {symbol}. Skipping volume filter.")
            # Do NOT fail if data fetch fails, as this is secondary confirmation
        else:
            current_volume = klines_data['volume'][-1]
            previous_volume = klines_data['volume'][-2]
            
            if previous_volume > 0:
                volume_increase = (current_volume - previous_volume) / previous_volume * 100
                if volume_increase < self.min_volume_increase_percent:
                    logger.warning(f"{symbol} Volume FAIL: Current volume ({current_volume:.0f}) is only {volume_increase:.2f}% higher than previous ({previous_volume:.0f}). Min: {self.min_volume_increase_percent}%. Skipping signal.")
                    return False
                else:
                    logger.debug(f"{symbol} Volume OK: {volume_increase:.2f}% increase.")
            else:
                logger.warning(f"{symbol} Previous volume is zero. Skipping volume increase check.")

        # 2. Orderbook Spread Confirmation
        try:
            order_book = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(
                    self.executor,
                    partial(self.monitoring_client.futures_order_book, symbol=symbol, limit=5)
                ),
                timeout=5
            )
            bid_price = float(order_book['bids'][0][0]) if order_book['bids'] else 0
            ask_price = float(order_book['asks'][0][0]) if order_book['asks'] else 0
            mid_price = (bid_price + ask_price) / 2 if bid_price > 0 and ask_price > 0 else 0
            
            if mid_price > 0:
                spread_percent = ((ask_price - bid_price) / mid_price * 100)
                if spread_percent > self.max_orderbook_spread_percent:
                    logger.warning(f"{symbol} Orderbook Spread FAIL: {spread_percent:.4f}% > {self.max_orderbook_spread_percent}%. Skipping signal.")
                    return False
                else:
                    logger.debug(f"{symbol} Orderbook Spread OK: {spread_percent:.4f}%")
            else:
                logger.warning(f"{symbol} Mid price is zero. Skipping orderbook spread check.")
                
        except asyncio.TimeoutError:
            logger.error(f"Timeout fetching order book for {symbol}")
            return False # Fail trade if critical market data is unavailable
        except Exception as e:
            logger.error(f"Error fetching order book for {symbol}: {e}")
            return False # Fail trade if critical market data is unavailable

        logger.info(f"{symbol} passed Volume Confirmation Filter.")
        return True

    # --- IMPROVEMENT 4: Candle-Wick Filter Check ---
    async def check_candle_wick_filter(self, symbol: str) -> bool:
        """Check if the last candle wick exceeds 1% and set a temporary block."""
        
        # Use 1-minute klines for the most recent candle check
        klines_data = await self.get_klines_data(symbol, '1m', limit=2)
        if klines_data is None or len(klines_data['close']) < 1:
            logger.warning(f"Could not fetch 1m klines for wick check on {symbol}. Skipping wick filter.")
            return True
            
        last_high = klines_data['high'][-1]
        last_low = klines_data['low'][-1]
        last_open = klines_data['open'][-1]
        last_close = klines_data['close'][-1]
        
        # True body size (absolute difference between open and close)
        body_size = abs(last_close - last_open)
        # Total candle range
        total_range = last_high - last_low
        
        # Calculate the size of the wicks
        # Upper wick is High - max(Open, Close)
        upper_wick = last_high - max(last_open, last_close)
        # Lower wick is min(Open, Close) - Low
        lower_wick = min(last_open, last_close) - last_low
        
        # Percentage of the biggest wick relative to the last close price
        # Check if the biggest wick is > 1% of the close price
        max_wick = max(upper_wick, lower_wick)
        if last_close > 0:
            max_wick_percent = (max_wick / last_close) * 100
        else:
            max_wick_percent = 0
        
        if max_wick_percent > self.max_candle_wick_percent:
            # Set a temporary block for this symbol
            block_until = time.time() + (self.candle_wick_skip_minutes * 60)
            self.blocked[symbol] = {'until': block_until, 'reason': 'LONG_WICK'}
            logger.warning(f"{symbol} Candle Wick FAIL: Max wick {max_wick_percent:.2f}% > {self.max_candle_wick_percent}%. Blocking until {datetime.fromtimestamp(block_until).strftime('%H:%M:%S')}")
            # Cache the candle close time to compare against time filters
            self.candle_cache[symbol] = {'close_time': time.time(), 'is_major_candle': True}
            return False
            
        logger.debug(f"{symbol} passed Candle Wick Filter. Max wick: {max_wick_percent:.2f}%")
        
        # Also, check if this was a major volume spike for time filter
        if len(klines_data['volume']) > 1 and klines_data['volume'][-2] > 0:
             volume_increase = (klines_data['volume'][-1] - klines_data['volume'][-2]) / klines_data['volume'][-2] * 100
             is_major_volume_spike = volume_increase > 50.0 # Arbitrary high percentage to denote a "huge volume spike"
        else:
             is_major_volume_spike = False

        self.candle_cache[symbol] = {
            'close_time': time.time(), 
            'is_major_candle': total_range > (last_close * 0.5/100) and body_size > 0, # Consider a large-range candle as "major"
            'is_major_volume_spike': is_major_volume_spike
        }
        return True

    # --- IMPROVEMENT 5: Time-Based Filter Check ---
    async def check_time_filters(self, symbol: str) -> bool:
        current_time = time.time()
        
        # 1. Dead Liquidity Time
        current_utc_hour = datetime.now(timezone.utc).hour
        if self.dead_liquidity_start_utc <= current_utc_hour < self.dead_liquidity_end_utc:
            logger.warning(f"Time Filter FAIL: Dead liquidity window (UTC {self.dead_liquidity_start_utc}–{self.dead_liquidity_end_utc}). Skipping signal for {symbol}.")
            return False

        # 2. Check for recent major candle or volume spike (uses cached info)
        if symbol in self.candle_cache:
            last_candle_close_time = self.candle_cache[symbol].get('close_time', 0)
            is_major_candle = self.candle_cache[symbol].get('is_major_candle', False)
            is_major_volume_spike = self.candle_cache[symbol].get('is_major_volume_spike', False)
            
            time_since_close = current_time - last_candle_close_time
            buffer_seconds = self.time_filter_buffer_minutes * 60

            if time_since_close < buffer_seconds:
                if is_major_candle:
                    logger.warning(f"Time Filter FAIL: Just after strong move ({time_since_close:.0f}s ago). Skipping signal for {symbol}.")
                    return False
                
                if is_major_volume_spike:
                    logger.warning(f"Time Filter FAIL: Just after huge volume spike ({time_since_close:.0f}s ago). Skipping signal for {symbol}.")
                    return False
        
        # 3. News Event Filter (Approximation: check for sudden extreme volume/volatility spike in 1m candles)
        # We'll rely on the existing volatility/ATR check, but add a 1m volume check as a proxy for news/major volatility events.
        try:
            klines_1m = await self.get_klines_data(symbol, '1m', limit=10)
            if klines_1m and len(klines_1m['volume']) > 5:
                # Calculate average volume over the last 5 minutes (excluding the current incomplete candle if any)
                recent_volumes = klines_1m['volume'][-6:-1]
                avg_volume = np.mean(recent_volumes)
                current_volume = klines_1m['volume'][-1]
                
                # If current volume is X times the average of the last few minutes, treat as major volatility
                if avg_volume > 0 and current_volume > avg_volume * 5: # 5x volume spike in the last minute
                    # Set a temporary block for 5 minutes, proxy for a news event
                    block_until = current_time + (self.time_filter_buffer_minutes * 60)
                    self.blocked[symbol] = {'until': block_until, 'reason': 'NEWS_VOLATILITY'}
                    logger.warning(f"Time Filter FAIL: Possible news event (5x volume spike). Blocking until {datetime.fromtimestamp(block_until).strftime('%H:%M:%S')}")
                    return False
        except Exception as e:
            logger.warning(f"Error checking 1m volume for news proxy: {e}")

        logger.info(f"{symbol} passed Time Filters.")
        return True

    async def get_symbol_volatility(self, symbol: str) -> dict:
        """Compute volatility metrics: ATR, 3-min volatility, bid-ask spread, funding rate."""
        try:
            volatility_data = {}
            
            # Get 1-hour klines to calculate ATR
            klines_1h = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(
                    self.executor,
                    partial(self.monitoring_client.futures_klines, symbol=symbol, interval='1h', limit=14)
                ),
                timeout=10
            )
            highs = [float(kline[2]) for kline in klines_1h]
            lows = [float(kline[3]) for kline in klines_1h]
            closes = [float(kline[4]) for kline in klines_1h]
            
            # Calculate ATR (Average True Range)
            tr_values = []
            for i in range(1, len(closes)):
                tr = max(
                    highs[i] - lows[i],
                    abs(highs[i] - closes[i-1]),
                    abs(lows[i] - closes[i-1])
                )
                tr_values.append(tr)
            atr = sum(tr_values) / len(tr_values) if tr_values else 0
            atr_percent = (atr / closes[-1] * 100) if closes[-1] > 0 else 0
            volatility_data['atr_percent'] = atr_percent
            
            # Get 3-minute klines for volatility
            klines_3m = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(
                    self.executor,
                    partial(self.monitoring_client.futures_klines, symbol=symbol, interval='3m', limit=20)
                ),
                timeout=10
            )
            closes_3m = [float(kline[4]) for kline in klines_3m]
            
            # Calculate 3-min volatility (standard deviation of returns)
            if len(closes_3m) > 1:
                returns = [(closes_3m[i] / closes_3m[i-1] - 1) for i in range(1, len(closes_3m))]
                volatility_3m = (np.std(returns)) * 100 # Using numpy for standard deviation
                volatility_data['volatility_3m_percent'] = volatility_3m
            else:
                volatility_data['volatility_3m_percent'] = 0
            
            # Get order book for bid-ask spread (redundant check, also in volume filter, but keep for consistency with old code)
            order_book = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(
                    self.executor,
                    partial(self.monitoring_client.futures_order_book, symbol=symbol, limit=5)
                ),
                timeout=10
            )
            bid_price = float(order_book['bids'][0][0]) if order_book['bids'] else 0
            ask_price = float(order_book['asks'][0][0]) if order_book['asks'] else 0
            mid_price = (bid_price + ask_price) / 2 if bid_price > 0 and ask_price > 0 else closes[-1]
            spread_percent = ((ask_price - bid_price) / mid_price * 100) if mid_price > 0 else 0
            volatility_data['spread_percent'] = spread_percent
            
            # Get funding rate
            funding_rate = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(
                    self.executor,
                    partial(self.monitoring_client.futures_funding_rate, symbol=symbol)
                ),
                timeout=10
            )
            funding_rate_val = float(funding_rate[-1]['fundingRate']) if funding_rate else 0
            volatility_data['funding_rate'] = funding_rate_val
            
            logger.debug(f"{symbol} volatility: ATR={atr_percent:.2f}%, Vol3m={volatility_data['volatility_3m_percent']:.2f}%, Spread={spread_percent:.4f}%, FundingRate={funding_rate_val:.4f}")
            return volatility_data
            
        except asyncio.TimeoutError:
            logger.error(f"Timeout fetching volatility data for {symbol}")
            return {}
        except Exception as e:
            logger.error(f"Error fetching volatility data for {symbol}: {e}")
            return {}

    # --- IMPROVEMENT 3: TIGHTENED VOLATILITY FILTER ---
    async def check_volatility_filter(self, symbol: str) -> bool:
        """Check if volatility and market conditions are within acceptable ranges."""
        vol_data = await self.get_symbol_volatility(symbol)
        
        if not vol_data:
            logger.warning(f"Could not fetch volatility data for {symbol}. Allowing trade.")
            return True
        
        # Check ATR
        if vol_data.get('atr_percent', 0) > self.max_atr_percent:
            logger.warning(f"{symbol} ATR too high: {vol_data['atr_percent']:.2f}% > {self.max_atr_percent}%. Skipping signal.")
            return False
        
        # Check 3-min volatility
        if vol_data.get('volatility_3m_percent', 0) > self.max_volatility_percent:
            logger.warning(f"{symbol} 3-min volatility too high: {vol_data['volatility_3m_percent']:.2f}% > {self.max_volatility_percent}%. Skipping signal.")
            return False
        
        # Check spread (optional)
        if self.enable_spread_filter:
            if vol_data.get('spread_percent', 0) > self.max_spread_percent:
                logger.warning(f"{symbol} bid-ask spread too high: {vol_data['spread_percent']:.4f}% > {self.max_spread_percent}%. Skipping signal.")
                return False
        else:
            logger.debug(f"{symbol} spread filter disabled. Spread: {vol_data.get('spread_percent', 0):.4f}%")
        
        # Check funding rate
        funding_rate = vol_data.get('funding_rate', 0)
        if funding_rate < self.min_funding_rate or funding_rate > self.max_funding_rate:
            logger.warning(f"{symbol} funding rate out of range: {funding_rate:.4f} (min: {self.min_funding_rate}, max: {self.max_funding_rate}). Skipping signal.")
            return False
        
        logger.info(f"{symbol} passed Volatility Filter")
        return True

    # --- STRATEGY 1: Technical Indicator Signal Validation ---
    async def check_technical_indicators_filter(self, symbol: str, direction: str) -> Tuple[bool, dict]:
        """
        Validate signal against RSI, MACD, and EMA indicators.
        Returns (passes_filter, indicator_data)
        
        Filters out:
        - LONG signals when RSI > overbought, MACD not rising, or price below EMA20
        - SHORT signals when RSI < oversold, MACD not falling, or price above EMA20
        """
        try:
            # Fetch 5m klines (100 candles for indicator calculation)
            klines = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(
                    self.executor,
                    partial(self.trading_client.futures_klines, symbol=symbol, interval='5m', limit=100)
                ),
                timeout=15
            )
            
            closes = [float(k[4]) for k in klines]
            current_price = closes[-1]
            
            # Calculate indicators using existing functions
            rsi = calculate_rsi(closes)
            macd_status, macd_value = calculate_macd_histogram(closes)
            ema20 = calculate_ema(closes, 20)
            
            indicator_data = {
                'rsi': rsi,
                'macd_status': macd_status,
                'macd_value': macd_value,
                'ema20': ema20,
                'current_price': current_price
            }
            
            # LONG validation
            if direction == SIDE_BUY:
                if rsi > self.rsi_overbought:
                    logger.warning(f"{symbol} LONG rejected: RSI overbought ({rsi:.1f} > {self.rsi_overbought})")
                    return False, indicator_data
                if macd_status != 'RISING':
                    logger.warning(f"{symbol} LONG rejected: MACD not rising ({macd_status})")
                    return False, indicator_data
                if self.require_ema_alignment and current_price < ema20:
                    logger.warning(f"{symbol} LONG rejected: Price below EMA20 ({current_price:.4f} < {ema20:.4f})")
                    return False, indicator_data
            
            # SHORT validation
            elif direction == SIDE_SELL:
                if rsi < self.rsi_oversold:
                    logger.warning(f"{symbol} SHORT rejected: RSI oversold ({rsi:.1f} < {self.rsi_oversold})")
                    return False, indicator_data
                if macd_status != 'FALLING':
                    logger.warning(f"{symbol} SHORT rejected: MACD not falling ({macd_status})")
                    return False, indicator_data
                if self.require_ema_alignment and current_price > ema20:
                    logger.warning(f"{symbol} SHORT rejected: Price above EMA20 ({current_price:.4f} > {ema20:.4f})")
                    return False, indicator_data
            
            logger.info(f"{symbol} passed Technical Indicators Filter (RSI={rsi:.1f}, MACD={macd_status}, Price vs EMA20={'above' if current_price > ema20 else 'below'})")
            return True, indicator_data
            
        except asyncio.TimeoutError:
            logger.warning(f"Timeout fetching klines for technical indicator filter on {symbol}. Allowing trade.")
            return True, {}  # Fail-open on timeout
        except Exception as e:
            logger.error(f"Error in technical indicators filter for {symbol}: {e}")
            return True, {}  # Fail-open on error

    async def validate_signal(self, signal: TradingSignal) -> bool:
        """Validate signal using RSI, MACD, and EMA."""
        if not self.enable_rsi_filter:
            return True  # Skip if disabled
        
        try:
            klines = await asyncio.get_running_loop().run_in_executor(
                self.executor,
                partial(self.trading_client.futures_klines, symbol=signal.coin, interval='5m', limit=100)
            )
            closes = [float(k[4]) for k in klines]
            
            rsi = calculate_rsi(closes)
            macd_status, _ = calculate_macd_histogram(closes)
            ema = calculate_ema(closes, self.ema_period)
            current_price = closes[-1]
            
            if rsi is None or ema is None:
                logger.warning(f"Insufficient data for validation on {signal.coin}. Skipping trade.")
                return False
            
            direction = signal.direction.upper()
            
            if direction == 'LONG':
                if rsi > self.rsi_overbought or macd_status != 'RISING' or current_price < ema:
                    logger.warning(f"Rejected LONG for {signal.coin}: RSI={rsi:.2f} (overbought={self.rsi_overbought}), MACD={macd_status}, Price below EMA.")
                    return False
            elif direction == 'SHORT':
                if rsi < self.rsi_oversold or macd_status != 'FALLING' or current_price > ema:
                    logger.warning(f"Rejected SHORT for {signal.coin}: RSI={rsi:.2f} (oversold={self.rsi_oversold}), MACD={macd_status}, Price above EMA.")
                    return False
            
            logger.info(f"Signal validated for {signal.coin}")
            return True
        
        except Exception as e:
            logger.error(f"Validation error for {signal.coin}: {e}")
            return False  # Reject on error

    # --- STRATEGY 2: Dynamic Position Sizing with Confidence Score ---
    async def calculate_confidence_score(self, symbol: str, direction: str, indicator_data: dict) -> float:
        """
        Calculate confidence score (0.0 to 1.0) based on indicator alignment.
        Higher score = stronger signal = larger position size.
        
        Args:
            symbol: Trading symbol
            direction: SIDE_BUY or SIDE_SELL
            indicator_data: Dict with RSI, MACD, EMA data from technical filter
        
        Returns:
            Confidence score between 0.0 and 1.0
        """
        score = 0.5  # Base score
        
        if not indicator_data:
            logger.debug(f"{symbol} no indicator data, using base confidence {score:.2f}")
            return score
        
        rsi = indicator_data.get('rsi', 50)
        macd_status = indicator_data.get('macd_status', 'NEUTRAL')
        current_price = indicator_data.get('current_price', 0)
        ema20 = indicator_data.get('ema20', 0)
        
        if direction == SIDE_BUY:
            # RSI in ideal range (40-60 = strong, 30-40 = moderate)
            if 40 <= rsi <= 60:
                score += 0.3
            elif 30 <= rsi < 40:
                score += 0.15
            
            # MACD rising
            if macd_status == 'RISING':
                score += 0.2
            
            # Price above EMA20
            if current_price > 0 and ema20 > 0 and current_price > ema20:
                score += 0.2
        
        elif direction == SIDE_SELL:
            # RSI in ideal range
            if 40 <= rsi <= 60:
                score += 0.3
            elif 60 < rsi <= 70:
                score += 0.15
            
            # MACD falling
            if macd_status == 'FALLING':
                score += 0.2
            
            # Price below EMA20
            if current_price > 0 and ema20 > 0 and current_price < ema20:
                score += 0.2
        
        # Cap at 1.0
        score = min(score, 1.0)
        logger.info(f"{symbol} confidence score: {score:.2f} (RSI={rsi:.1f}, MACD={macd_status})")
        return score

    def round_to_tick_size(self, price: float, tick_size: float) -> float:
        return round(price / tick_size) * tick_size

    def round_to_lot_size(self, quantity: float, step_size: float, precision: int) -> float:
        # Use floor/truncate for safety, ensuring quantity doesn't exceed available lot size
        multiplier = 10**precision
        rounded_quantity = math.floor(quantity * multiplier) / multiplier
        
        # Ensure it's a multiple of step_size
        if step_size > 0:
            steps = math.floor(rounded_quantity / step_size)
            return round(steps * step_size, precision)
        return rounded_quantity

    # --- BULLETPROOFING: Helper Methods ---
    
    async def verify_order_active(self, symbol: str, order_id: int, max_attempts: int = 3) -> bool:
        """Verify an order is active on the exchange"""
        for attempt in range(max_attempts):
            try:
                order = await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(self.executor, partial(
                        self.monitoring_client.futures_get_order,
                        symbol=symbol,
                        orderId=order_id
                    )),
                    timeout=15
                )
                
                if order['status'] in ['NEW', 'PARTIALLY_FILLED']:
                    return True
                else:
                    logger.warning(f"Order {order_id} status: {order['status']}")
                    return False
                    
            except BinanceAPIException as e:
                if e.code == -2011:  # Order not found
                    logger.error(f"Order {order_id} not found on exchange")
                    return False
                logger.warning(f"Attempt {attempt+1}: Error verifying order: {e}")
                if attempt < max_attempts - 1:
                    await asyncio.sleep(1)
            except Exception as e:
                logger.error(f"Error verifying order {order_id}: {e}")
                if attempt < max_attempts - 1:
                    await asyncio.sleep(1)
        
        return False

    async def emergency_close_position(self, symbol: str, close_side: str, quantity: float):
        """Emergency close position if TP/SL placement fails"""
        logger.error(f"🚨 EMERGENCY: Closing position for {symbol} due to failed protective orders")
        
        try:
            # Close at market immediately
            close_order = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, partial(
                    self.trading_client.futures_create_order,
                    symbol=symbol,
                    side=close_side,
                    type='MARKET',
                    quantity=quantity,
                    reduceOnly=True
                )),
                timeout=15
            )
            
            logger.info(f"Emergency close order placed for {symbol}: {close_order['orderId']}")
            
            # PHASE 3: Send CRITICAL alert
            await self.alerts.send_alert(
                "CRITICAL",
                f"EMERGENCY CLOSE: {symbol}",
                f"Position closed at market due to failed protective orders.\n\n"
                f"<b>Details:</b>\n"
                f"• Symbol: {symbol}\n"
                f"• Quantity: {quantity}\n"
                f"• Reason: SL placement failed after all retries",
                symbol=symbol
            )
            
            # Record as emergency close
            await self.update_journal({
                'Symbol': symbol,
                'Close Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Status': 'Emergency Close - Failed TP/SL'
            })
            
            # Cleanup
            self.cleanup_symbol(symbol)
            return True
            
        except Exception as e:
            logger.critical(f"CRITICAL: Emergency close FAILED for {symbol}: {e}")
            
            # PHASE 3: Send CRITICAL alert for failed emergency close
            await self.alerts.send_alert(
                "CRITICAL",
                f"EMERGENCY CLOSE FAILED: {symbol}",
                f"<b>CRITICAL:</b> Emergency close attempt FAILED!\n\n"
                f"<b>Details:</b>\n"
                f"• Symbol: {symbol}\n"
                f"• Error: {str(e)}\n"
                f"• Action: Manual intervention required!",
                symbol=symbol
            )
            return False

    async def cancel_order_safe(self, symbol: str, order_id: int, force_cancel_attempt: bool = False) -> bool:
        """
        Safely cancel an order with proper error handling and retry logic
        
        Args:
            symbol: Trading symbol
            order_id: Order ID to cancel
            force_cancel_attempt: If True, always attempt cancel with trading_client even if 
                                 monitoring_client reports order not found. Use this for timeout
                                 scenarios where API permission mismatches may occur.
        """
        max_retries = 3
        
        for attempt in range(max_retries):
            try:
                # Check if order exists first using monitoring client
                try:
                    order = await asyncio.wait_for(
                        asyncio.get_running_loop().run_in_executor(self.executor, partial(
                            self.monitoring_client.futures_get_order,
                            symbol=symbol,
                            orderId=order_id
                        )),
                        timeout=15
                    )
                    
                    if order['status'] in ['NEW', 'PARTIALLY_FILLED']:
                        # Cancel it using trading client
                        await asyncio.wait_for(
                            asyncio.get_running_loop().run_in_executor(self.executor, partial(
                                self.trading_client.futures_cancel_order,
                                symbol=symbol,
                                orderId=order_id
                            )),
                            timeout=15
                        )
                        logger.info(f"Canceled order {order_id} for {symbol}")
                        return True
                    else:
                        logger.info(f"Order {order_id} already {order['status']}, no need to cancel")
                        return True
                        
                except BinanceAPIException as e:
                    if e.code == -2011:  # Order not found via monitoring client
                        logger.warning(f"Order {order_id} not found via monitoring client — this may indicate API permission mismatch")
                        
                        # If force_cancel_attempt is True, try to cancel anyway using trading client
                        if force_cancel_attempt:
                            logger.info(f"Attempting cancel with trading client anyway (force_cancel_attempt=True)")
                            try:
                                await asyncio.wait_for(
                                    asyncio.get_running_loop().run_in_executor(self.executor, partial(
                                        self.trading_client.futures_cancel_order,
                                        symbol=symbol,
                                        orderId=order_id
                                    )),
                                    timeout=15
                                )
                                logger.info(f"Successfully canceled order {order_id} for {symbol} via trading client")
                                return True
                            except BinanceAPIException as cancel_error:
                                if cancel_error.code == -2011:
                                    # Order truly doesn't exist
                                    logger.info(f"Order {order_id} confirmed not found (already processed)")
                                    return True
                                else:
                                    logger.error(f"Error canceling order {order_id} via trading client: {cancel_error}")
                                    # Don't return False yet, let retry logic handle it
                                    if attempt < max_retries - 1:
                                        backoff_time = 2 ** attempt  # 1s, 2s, 4s
                                        logger.warning(f"Retry {attempt + 1}/{max_retries} for order {order_id} in {backoff_time}s...")
                                        await asyncio.sleep(backoff_time)
                                        continue
                                    return False
                        else:
                            # Not forcing, assume order is already processed
                            logger.info(f"Order {order_id} assumed already processed (not forcing cancel)")
                            return True
                    else:
                        # Other API error - retry
                        logger.warning(f"API error (code {e.code}) canceling order {order_id}: {e}")
                        if attempt < max_retries - 1:
                            backoff_time = 2 ** attempt  # 1s, 2s, 4s
                            logger.warning(f"Retry {attempt + 1}/{max_retries} for order {order_id} in {backoff_time}s...")
                            await asyncio.sleep(backoff_time)
                            continue
                        raise
                    
            except BinanceAPIException as e:
                logger.error(f"API error canceling order {order_id} (attempt {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    backoff_time = 2 ** attempt  # 1s, 2s, 4s
                    logger.warning(f"Retrying in {backoff_time}s...")
                    await asyncio.sleep(backoff_time)
                else:
                    return False
            except Exception as e:
                logger.error(f"Unexpected error canceling order {order_id} (attempt {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    backoff_time = 2 ** attempt  # 1s, 2s, 4s
                    logger.warning(f"Retrying in {backoff_time}s...")
                    await asyncio.sleep(backoff_time)
                else:
                    return False
        
        # All retries exhausted
        logger.error(f"Failed to cancel order {order_id} for {symbol} after {max_retries} attempts")
        return False

    # --- END BULLETPROOFING: Helper Methods ---

    # --- BULLETPROOFING: TP/SL Placement with Verification ---
    
    async def place_and_verify_protective_orders(self, symbol: str, direction: str, 
                                                 opposite_side: str, avg_entry_price: float,
                                                 total_qty: float, leverage: int, roi_targets: list,
                                                 qty_precision: int, price_precision: int, tick_size: float) -> bool:
        """
        Place TP/SL orders with verification and emergency close fallback.
        Returns True if successful, False if emergency close was triggered.
        """
        
        # Calculate SL price
        sl_pct = self.sl_percentage / leverage
        initial_stop_price = avg_entry_price * (1 - sl_pct) if direction == SIDE_BUY else avg_entry_price * (1 + sl_pct)
        
        # Place SL with retry and verification
        sl_order_id = await self.place_sl_order(
            symbol=symbol,
            side=opposite_side,
            stop_price=initial_stop_price,
            qty_precision=qty_precision,
            price_precision=price_precision,
            tick_size=tick_size,
            quantity=total_qty
        )
        
        if not sl_order_id:
            logger.error(f"CRITICAL: Failed to place SL for {symbol} after all attempts!")
            # Emergency: Close position immediately
            await self.emergency_close_position(symbol, opposite_side, total_qty)
            return False
        
        # Place TP orders with verification
        tp_success = await self.place_tp_orders(
            symbol=symbol,
            opposite_side=opposite_side,
            quantity=total_qty,
            targets=roi_targets,
            qty_precision=qty_precision,
            price_precision=price_precision,
            current_price=avg_entry_price,
            tick_size=tick_size
        )
        
        if not tp_success:
            logger.warning(f"WARNING: TP placement had issues for {symbol}, but SL is active")
        
        return True

    # --- END BULLETPROOFING: TP/SL Placement with Verification ---

    # --- PHASE 2: Position Reconciliation ---
    
    async def place_emergency_sl(self, symbol: str, position: dict):
        """Place emergency SL for unprotected position"""
        try:
            position_amt = float(position['positionAmt'])
            entry_price = float(position['entryPrice'])
            
            if position_amt == 0:
                logger.warning(f"Cannot place emergency SL for {symbol}: position is 0")
                return False
            
            # Determine direction and calculate emergency SL (2% from entry)
            if position_amt > 0:  # Long position
                sl_price = entry_price * 0.98  # 2% below entry
                opposite_side = SIDE_SELL
                direction = SIDE_BUY
            else:  # Short position
                sl_price = entry_price * 1.02  # 2% above entry
                opposite_side = SIDE_BUY
                direction = SIDE_SELL
            
            # Get exchange info for precision
            exchange_info = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_exchange_info),
                timeout=15
            )
            symbol_info = next((s for s in exchange_info['symbols'] if s['symbol'] == symbol), None)
            if not symbol_info:
                logger.error(f"Cannot get exchange info for {symbol}")
                return False
            
            qty_precision = symbol_info['quantityPrecision']
            price_precision = symbol_info['pricePrecision']
            tick_size = float(next(f['tickSize'] for f in symbol_info['filters'] if f['filterType'] == 'PRICE_FILTER'))
            
            # Place emergency SL
            sl_order_id = await self.place_sl_order(
                symbol=symbol,
                side=opposite_side,
                stop_price=sl_price,
                qty_precision=qty_precision,
                price_precision=price_precision,
                tick_size=tick_size,
                quantity=abs(position_amt)
            )
            
            if sl_order_id:
                logger.info(f"✓ Emergency SL placed for {symbol} at {sl_price:.{price_precision}f}")
                return True
            else:
                logger.error(f"Failed to place emergency SL for {symbol}")
            return False
                
        except Exception as e:
            logger.error(f"Error placing emergency SL for {symbol}: {e}")
            return False
    
    # --- ALGO ORDER RECONCILIATION AND CANCELLATION (Dec 2025 API Migration) ---
    
    async def reconcile_algo_orders(self, symbols_with_positions: set):
        """
        Fetch open algo orders and reconcile with current positions.
        Cancels orphan algo orders (TP/SL without positions).
        Loads existing algo orders into state dictionaries.
        
        Args:
            symbols_with_positions: Set of symbols that have open positions
        """
        try:
            logger.info("Fetching open algo orders for reconciliation...")
            
            # Fetch all open algo orders
            open_algo_orders = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(
                    self.executor,
                    self.monitoring_client.futures_get_algo_open_orders
                ),
                timeout=15
            )
            
            logger.info(f"Found {len(open_algo_orders)} open algo orders")
            
            for algo_order in open_algo_orders:
                try:
                    symbol = algo_order['symbol']
                    algo_id = algo_order['algoId']
                    order_type = algo_order.get('orderType', '').upper()
                    reduce_only = bool(algo_order.get('reduceOnly'))
                    close_position = bool(algo_order.get('closePosition'))
                    
                    # Skip if not a protective order
                    if not reduce_only and not close_position:
                        continue
                    
                    # Check if this is an orphan (no position exists)
                    if symbol not in symbols_with_positions:
                        try:
                            logger.info(f"Found orphan algo order {algo_id} for {symbol} (no open position). Canceling...")
                            await asyncio.get_running_loop().run_in_executor(
                                self.executor,
                                partial(self.trading_client.futures_cancel_algo_order, symbol=symbol, algoId=algo_id)
                            )
                            logger.info(f"✓ Canceled orphan algo order {algo_id} for {symbol}")
                        except BinanceAPIException as e:
                            if e.code == -2011:  # Already canceled
                                logger.debug(f"Orphan algo order {algo_id} already canceled")
                            else:
                                logger.error(f"Failed to cancel orphan algo order {algo_id}: {e}")
                        except Exception as e:
                            logger.exception(f"Error canceling orphan algo order {algo_id} for {symbol}: {e}")
                        continue
                    
                    # Map to state dictionaries
                    if symbol not in self.tp_orders:
                        self.tp_orders[symbol] = {}
                    
                    if order_type in ['STOP_MARKET', 'STOP']:
                        # This is a Stop Loss order
                        if symbol in self.sl_orders:
                            logger.warning(f"Multiple SL orders found for {symbol}. Keeping newest: {algo_id}")
                        self.sl_orders[symbol] = algo_id
                        logger.info(f"Loaded SL algo order {algo_id} for {symbol}")
                        
                    elif order_type in ['TAKE_PROFIT_MARKET', 'TAKE_PROFIT']:
                        # This is a Take Profit order
                        if 'TP1' not in self.tp_orders[symbol]:
                            self.tp_orders[symbol]['TP1'] = algo_id
                            logger.info(f"Loaded TP1 algo order {algo_id} for {symbol}")
                        elif 'TP2' not in self.tp_orders[symbol]:
                            self.tp_orders[symbol]['TP2'] = algo_id
                            logger.info(f"Loaded TP2 algo order {algo_id} for {symbol}")
                        else:
                            logger.warning(f"Extra TP algo order {algo_id} found for {symbol}. Canceling...")
                            try:
                                await asyncio.get_running_loop().run_in_executor(
                                    self.executor,
                                    partial(self.trading_client.futures_cancel_algo_order, symbol=symbol, algoId=algo_id)
                                )
                                logger.info(f"✓ Canceled extra TP algo order {algo_id}")
                            except Exception as e:
                                logger.error(f"Failed to cancel extra TP algo order: {e}")
                                
                except Exception as e:
                    logger.exception(f"Error processing algo order {algo_order}: {e}")
            
            # Save updated state
            try:
                await self._save_state()
                logger.info("✓ Algo order reconciliation complete")
            except Exception as e:
                logger.exception("Failed to save state after algo reconciliation")
                
        except asyncio.TimeoutError:
            logger.error("Timeout fetching algo orders for reconciliation")
        except Exception as e:
            logger.exception(f"Error in algo order reconciliation: {e}")
    
    async def cancel_algo_orders_for_symbol(self, symbol: str) -> bool:
        """
        Cancel all algo orders (TP/SL) for a specific symbol.
        Uses futures_cancel_algo_order with algoId.
        
        Args:
            symbol: Trading symbol
            
        Returns:
            True if all cancellations succeeded, False otherwise
        """
        success = True
        
        # Cancel SL order
        if symbol in self.sl_orders:
            algo_id = self.sl_orders[symbol]
            try:
                cancel_response = await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(
                        self.executor,
                        partial(self.trading_client.futures_cancel_algo_order, symbol=symbol, algoId=algo_id)
                    ),
                    timeout=15
                )
                
                logger.info(f"✓ Canceled SL algo order {algo_id} for {symbol}")
                del self.sl_orders[symbol]
                
            except BinanceAPIException as e:
                if e.code == -2011:  # Order not found (already canceled)
                    logger.info(f"SL algo order {algo_id} for {symbol} not found (likely already canceled)")
                    del self.sl_orders[symbol]
                else:
                    logger.error(f"Error canceling SL algo order {algo_id} for {symbol}: {e}")
                    success = False
            except asyncio.TimeoutError:
                logger.warning(f"Timeout canceling SL algo order {algo_id} for {symbol}. Assuming canceled.")
                del self.sl_orders[symbol]
            except Exception as e:
                logger.error(f"Error canceling SL algo order for {symbol}: {e}")
                success = False
        
        # Cancel TP orders
        if symbol in self.tp_orders:
            tp_orders_copy = dict(self.tp_orders[symbol])  # Copy to avoid modification during iteration
            
            for tp_name, algo_id in tp_orders_copy.items():
                try:
                    cancel_response = await asyncio.wait_for(
                        asyncio.get_running_loop().run_in_executor(
                            self.executor,
                            partial(self.trading_client.futures_cancel_algo_order, symbol=symbol, algoId=algo_id)
                        ),
                        timeout=15
                    )
                    
                    logger.info(f"✓ Canceled {tp_name} algo order {algo_id} for {symbol}")
                    del self.tp_orders[symbol][tp_name]
                    
                except BinanceAPIException as e:
                    if e.code == -2011:  # Order not found
                        logger.info(f"{tp_name} algo order {algo_id} for {symbol} not found (likely already canceled)")
                        del self.tp_orders[symbol][tp_name]
                    else:
                        logger.error(f"Error canceling {tp_name} algo order {algo_id} for {symbol}: {e}")
                        success = False
                except asyncio.TimeoutError:
                    logger.warning(f"Timeout canceling {tp_name} algo order {algo_id} for {symbol}. Assuming canceled.")
                    del self.tp_orders[symbol][tp_name]
                except Exception as e:
                    logger.error(f"Error canceling {tp_name} algo order for {symbol}: {e}")
                    success = False
            
            # Clean up empty TP orders dict
            if not self.tp_orders[symbol]:
                del self.tp_orders[symbol]
        
        # Save state
        try:
            await self._save_state()
        except Exception as e:
            logger.exception("Failed to save state after canceling algo orders")
        
        return success
    
    async def verify_protective_orders(self, symbol: str) -> bool:
        """
        Verify that protective orders (SL/TP) are active for a symbol.
        Checks algo order status and alerts if orders are not WORKING.
        
        Args:
            symbol: Trading symbol
            
        Returns:
            True if all protective orders are active, False otherwise
        """
        all_active = True
        
        # Verify SL order
        if symbol in self.sl_orders:
            algo_id = self.sl_orders[symbol]
            try:
                sl_status = await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(
                        self.executor,
                        partial(self.monitoring_client.futures_get_algo_order, symbol=symbol, algoId=algo_id)
                    ),
                    timeout=10
                )
                
                status = sl_status.get('status')
                if status != 'WORKING':
                    logger.warning(f"SL for {symbol} not active. Status: {status}")
                    all_active = False
                    
                    # Send alert
                    await self.alerts.send_alert(
                        "WARNING",
                        f"SL Not Active: {symbol}",
                        f"Stop Loss algo order status: {status}\n"
                        f"AlgoId: {algo_id}\n"
                        f"Action: Monitor position manually",
                        symbol=symbol
                    )
                    
            except Exception as e:
                logger.error(f"Error verifying SL for {symbol}: {e}")
                all_active = False
        else:
            logger.warning(f"No SL order found for {symbol}")
            all_active = False
        
        # Verify TP orders
        if symbol in self.tp_orders:
            for tp_name, algo_id in self.tp_orders[symbol].items():
                try:
                    tp_status = await asyncio.wait_for(
                        asyncio.get_running_loop().run_in_executor(
                            self.executor,
                            partial(self.monitoring_client.futures_get_algo_order, symbol=symbol, algoId=algo_id)
                        ),
                        timeout=10
                    )
                    
                    status = tp_status.get('status')
                    if status not in ['WORKING', 'TRIGGERED']:  # TRIGGERED is OK (order executed)
                        logger.warning(f"{tp_name} for {symbol} not active. Status: {status}")
                        all_active = False
                        
                except Exception as e:
                    logger.error(f"Error verifying {tp_name} for {symbol}: {e}")
                    all_active = False
        
        return all_active
    

    async def reconcile_positions_periodic(self):
        """Periodically check all positions have proper TP/SL protection"""
        logger.info("Starting position reconciliation task (runs every 5 minutes)")
        
        while True:
            try:
                await asyncio.sleep(300)  # Every 5 minutes
                
                logger.debug("Running position reconciliation check...")
                
                # Get all open positions
                positions = await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(
                        self.executor, 
                        self.monitoring_client.futures_position_information
                    ),
                    timeout=15
                )
                
                positions_checked = 0
                issues_found = 0
                
                for position in positions:
                    symbol = position['symbol']
                    position_amt = float(position['positionAmt'])
                    
                    # Skip if no position
                    if position_amt == 0:
                        continue
                    
                    positions_checked += 1
                    
                    # Check if we're tracking this position
                    if symbol not in self.position_info:
                        logger.error(f"🚨 Reconciliation: UNTRACKED position found: {symbol} with qty {position_amt}")
                        issues_found += 1
                        
                        # PHASE 3: Send CRITICAL alert for untracked position
                        await self.alerts.send_alert(
                            "CRITICAL",
                            f"UNTRACKED POSITION: {symbol}",
                            f"<b>Position found on exchange but not in bot tracking!</b>\n\n"
                            f"<b>Details:</b>\n"
                            f"• Symbol: {symbol}\n"
                            f"• Quantity: {position_amt}\n"
                            f"• Entry Price: {position['entryPrice']}\n"
                            f"• Possible Cause: Bot restart during fill\n\n"
                            f"<b>Action: Placing emergency SL for protection...</b>",
                            symbol=symbol
                        )
                        
                        # Place emergency SL to protect this untracked position
                        success = await self.place_emergency_sl(symbol, position)
                        if success:
                            logger.info(f"✓ Emergency SL placed for untracked position {symbol}")
                            # Add to tracking so we don't keep alerting
                            entry_price = float(position['entryPrice'])
                            self.position_info[symbol] = {
                                'total_qty': abs(position_amt),
                                'avg_entry_price': entry_price
                            }
                        else:
                            logger.error(f"❌ Failed to place emergency SL for untracked {symbol}")
                            # Send another alert about failure
                            await self.alerts.send_alert(
                                "CRITICAL",
                                f"EMERGENCY SL FAILED: {symbol}",
                                f"<b>MANUAL INTERVENTION REQUIRED!</b>\n\n"
                                f"• Symbol: {symbol}\n"
                                f"• Quantity: {position_amt}\n"
                                f"• Action: Place SL manually on Binance!",
                                symbol=symbol
                            )
                        continue
                    
                    # CRITICAL: Check if SL exists
                    if symbol not in self.sl_orders:
                        logger.error(f"🚨 Reconciliation: Position {symbol} has NO SL order!")
                        issues_found += 1
                        
                        # PHASE 3: Send CRITICAL alert
                        await self.alerts.send_alert(
                            "CRITICAL",
                            f"UNPROTECTED POSITION: {symbol}",
                            f"Position has NO Stop Loss protection!\n\n"
                            f"<b>Details:</b>\n"
                            f"• Symbol: {symbol}\n"
                            f"• Position: {position_amt}\n"
                            f"• Entry Price: {position['entryPrice']}\n"
                            f"• Action: Placing emergency SL...",
                            symbol=symbol
                        )
                        
                        # Place emergency SL
                        success = await self.place_emergency_sl(symbol, position)
                        if success:
                            logger.info(f"✓ Emergency SL placed for unprotected {symbol}")
                        else:
                            logger.error(f"❌ Failed to place emergency SL for {symbol}")
                    else:
                        # Verify SL is still active
                        sl_order_id = self.sl_orders[symbol]
                        sl_active = await self.verify_order_active(symbol, sl_order_id)
                        
                        if not sl_active:
                            logger.error(f"🚨 Reconciliation: SL order {sl_order_id} for {symbol} is not active!")
                            issues_found += 1
                            
                            # PHASE 3: Send CRITICAL alert
                            await self.alerts.send_alert(
                                "CRITICAL",
                                f"SL ORDER INACTIVE: {symbol}",
                                f"Stop Loss order is not active on exchange!\n\n"
                                f"<b>Details:</b>\n"
                                f"• Symbol: {symbol}\n"
                                f"• SL Order ID: {sl_order_id}\n"
                                f"• Status: CANCELED/EXPIRED\n"
                                f"• Action: Placing new emergency SL...",
                                symbol=symbol
                            )
                            
                            # Remove from tracking
                            self.sl_orders.pop(symbol, None)
                            # Place new emergency SL
                            success = await self.place_emergency_sl(symbol, position)
                            if success:
                                logger.info(f"✓ Replaced inactive SL for {symbol}")
                            else:
                                logger.error(f"❌ Failed to replace SL for {symbol}")
                    
                    # Check if at least one TP exists and auto-place if missing
                    if symbol not in self.tp_orders or not self.tp_orders[symbol]:
                        logger.warning(f"⚠️ Reconciliation: Position {symbol} has NO TP orders")
                        
                        # Auto-place missing TP orders using configured ROI targets
                        try:
                            logger.info(f"Attempting to auto-place missing TP orders for {symbol}...")
                            
                            # Get position details
                            entry_price = float(position['entryPrice'])
                            position_amt = abs(float(position['positionAmt']))
                            
                            # Determine direction
                            if float(position['positionAmt']) > 0:
                                opposite_side = 'SELL'
                            else:
                                opposite_side = 'BUY'
                            
                            # Get symbol info for precision
                            symbol_info = await asyncio.wait_for(
                                asyncio.get_running_loop().run_in_executor(
                                    self.executor,
                                    partial(self.trading_client.futures_exchange_info)
                                ),
                                timeout=10
                            )
                            
                            # Find symbol details
                            symbol_data = next((s for s in symbol_info['symbols'] if s['symbol'] == symbol), None)
                            if not symbol_data:
                                logger.error(f"Could not find symbol info for {symbol}")
                                continue
                            
                            # Extract precision
                            qty_precision = symbol_data['quantityPrecision']
                            price_precision = symbol_data['pricePrecision']
                            tick_size = float(next(f['tickSize'] for f in symbol_data['filters'] if f['filterType'] == 'PRICE_FILTER'))
                            
                            # Calculate TP targets from .env settings
                            tp1_roi = float(os.getenv('TP1_ROI', '0.4'))  # Default 0.4%
                            tp2_roi = float(os.getenv('TP2_ROI', '1.0'))  # Default 1.0%
                            
                            if opposite_side == 'SELL':  # LONG position
                                tp1_price = entry_price * (1 + tp1_roi / 100)
                                tp2_price = entry_price * (1 + tp2_roi / 100)
                            else:  # SHORT position
                                tp1_price = entry_price * (1 - tp1_roi / 100)
                                tp2_price = entry_price * (1 - tp2_roi / 100)
                            
                            targets = [tp1_price, tp2_price]
                            
                            # Get current price
                            ticker = await asyncio.wait_for(
                                asyncio.get_running_loop().run_in_executor(
                                    self.executor,
                                    partial(self.trading_client.futures_symbol_ticker, symbol=symbol)
                                ),
                                timeout=10
                            )
                            current_price = float(ticker['price'])
                            
                            # Place TP orders
                            tp_success = await self.place_tp_orders(
                                symbol=symbol,
                                opposite_side=opposite_side,
                                quantity=position_amt,
                                targets=targets,
                                qty_precision=qty_precision,
                                price_precision=price_precision,
                                current_price=current_price,
                                tick_size=tick_size
                            )
                            
                            if tp_success:
                                logger.info(f"✓ Auto-placed missing TP orders for {symbol}")
                                await self.alerts.send_alert(
                                    "INFO",
                                    f"TP Orders Auto-Placed: {symbol}",
                                    f"<b>Reconciliation fixed missing TP orders</b>\\n\\n"
                                    f"• Symbol: {symbol}\\n"
                                    f"• Entry Price: {entry_price:.4f}\\n"
                                    f"• TP1: {tp1_price:.4f} ({tp1_roi}%)\\n"
                                    f"• TP2: {tp2_price:.4f} ({tp2_roi}%)\\n\\n"
                                    f"✅ Position now fully protected",
                                    symbol=symbol
                                )
                            else:
                                logger.error(f"❌ Failed to auto-place TP orders for {symbol}")
                                
                        except Exception as e:
                            logger.error(f"Error auto-placing TP orders for {symbol}: {e}")
                            # Not critical since SL still protects the position
                
                if positions_checked > 0:
                    logger.info(f"Reconciliation complete: Checked {positions_checked} positions, found {issues_found} issues")
                else:
                    logger.debug("Reconciliation: No open positions to check")
                    
            except asyncio.CancelledError:
                logger.info("Position reconciliation task cancelled")
                break
            except Exception as e:
                logger.error(f"Error in position reconciliation: {e}")
                # Continue running despite errors
    
    # --- END PHASE 2: Position Reconciliation ---

    # --- PHASE 2: Async Limit Order Monitoring ---
    
    async def monitor_limit_order_fill(self, symbol: str, order_id: int, direction: str, 
                                       opposite_side: str, entry_price: float, quantity: float,
                                       leverage: int, roi_targets: list, qty_precision: int,
                                       price_precision: int, tick_size: float, signal_timestamp: float):
        """
        Background task to monitor limit order fill (non-blocking).
        Places TP/SL after fill and starts position monitoring.
        """
        logger.info(f"Started monitoring limit order {order_id} for {symbol}")
        
        max_wait_time = 3600  # 1 hour
        check_interval = 10   # Check every 10 seconds
        elapsed = 0
        filled = False
        
        try:
            while elapsed < max_wait_time:
                try:
                    # Check order status
                    order_status = await asyncio.wait_for(
                        asyncio.get_running_loop().run_in_executor(self.executor, partial(
                            self.monitoring_client.futures_get_order,
                            symbol=symbol,
                            orderId=order_id
                        )),
                        timeout=15
                    )
                    
                    if order_status['status'] == 'FILLED':
                        logger.info(f"✓ Limit order {order_id} for {symbol} FILLED!")
                        filled = True
                        
                        # CRITICAL: Wrap entire fill processing in try-catch for safety
                        try:
                            # Get actual fill details
                            avg_entry_price = float(order_status['avgPrice'])
                            total_qty = float(order_status['executedQty'])
                            
                            # Store position info directly (no need to call calculate_position_info)
                            self.position_info[symbol] = {
                                'total_qty': total_qty,
                                'avg_entry_price': avg_entry_price
                            }
                            
                            # Update journal
                            await self.update_journal({
                                'Symbol': symbol,
                                'Direction': direction,
                                'Entry Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                'Entry Price': avg_entry_price,
                                'Quantity': total_qty,
                                'Leverage': leverage,
                                'Entry Type': 'LIMIT',
                                'Signal Timestamp': datetime.fromtimestamp(signal_timestamp).strftime('%Y-%m-%d %H:%M:%S')
                            })
                            
                            # Record trade for frequency tracking
                            self.record_trade_opened()
                            
                            # BULLETPROOFING: Place TP/SL with verification
                            protection_success = await self.place_and_verify_protective_orders(
                                symbol=symbol,
                                direction=direction,
                                opposite_side=opposite_side,
                                avg_entry_price=avg_entry_price,
                                total_qty=total_qty,
                                leverage=leverage,
                                roi_targets=roi_targets,
                                qty_precision=qty_precision,
                                price_precision=price_precision,
                                tick_size=tick_size
                            )
                            
                            if not protection_success:
                                logger.error(f"Failed to place protective orders for {symbol}. Position was emergency closed.")
                                # Cleanup
                                self.cleanup_symbol(symbol)
                                return
                            
                            # Reserve slot
                            self.symbols_holding_slots.add(symbol)
                            logger.info(f"Reserved slot for {symbol} after limit entry")
                            
                            # ALERT: Limit Order Filled
                            try:
                                notional = total_qty * avg_entry_price
                                await self.alerts.send_alert(
                                    "INFO",
                                    f"📈 Entry Filled: {symbol}",
                                    f"<b>Limit order executed!</b>\n\n"
                                    f"• Symbol: {symbol}\n"
                                    f"• Direction: {direction}\n"
                                    f"• Entry Price: {avg_entry_price:.4f}\n"
                                    f"• Quantity: {total_qty:.4f}\n"
                                    f"• Leverage: {leverage}x\n"
                                    f"• Notional: {notional:.2f} USDT\n\n"
                                    f"✅ TP/SL orders placed. Monitoring started.",
                                    symbol=symbol
                                )
                            except Exception as e:
                                logger.error(f"Failed to send limit fill alert for {symbol}: {e}")
                            
                            # Start position monitoring
                            asyncio.create_task(self.monitor_market_position(
                                symbol=symbol,
                                direction=direction,
                                opposite_side=opposite_side,
                                avg_entry_price=avg_entry_price,
                                total_qty=total_qty,
                                qty_precision=qty_precision,
                                price_precision=price_precision,
                                tick_size=tick_size,
                                leverage=leverage  # Add leverage parameter
                            ))
                            
                            logger.info(f"Limit order filled for {symbol}. Monitoring started.")
                            return
                            
                        except Exception as fill_error:
                            # CRITICAL: If ANYTHING goes wrong during fill processing, emergency close!
                            logger.critical(f"🚨 CRITICAL ERROR processing fill for {symbol}: {fill_error}")
                            
                            # Send CRITICAL alert
                            await self.alerts.send_alert(
                                "CRITICAL",
                                f"FILL PROCESSING ERROR: {symbol}",
                                f"<b>CRITICAL:</b> Error during fill processing!\n\n"
                                f"<b>Details:</b>\n"
                                f"• Symbol: {symbol}\n"
                                f"• Order ID: {order_id}\n"
                                f"• Error: {str(fill_error)}\n"
                                f"• Action: Attempting emergency close...",
                                symbol=symbol
                            )
                            
                            # Emergency close the position
                            try:
                                await self.emergency_close_position(symbol, opposite_side, total_qty)
                                logger.info(f"✓ Emergency close successful for {symbol}")
                            except Exception as close_error:
                                logger.critical(f"❌ EMERGENCY CLOSE FAILED for {symbol}: {close_error}")
                                await self.alerts.send_alert(
                                    "CRITICAL",
                                    f"EMERGENCY CLOSE FAILED: {symbol}",
                                    f"<b>MANUAL INTERVENTION REQUIRED!</b>\n\n"
                                    f"• Symbol: {symbol}\n"
                                    f"• Close Error: {str(close_error)}\n"
                                    f"• Action: Close position manually on Binance!",
                                    symbol=symbol
                                )
                            
                            # Cleanup and return
                            self.cleanup_symbol(symbol)
                            return
                        
                    elif order_status['status'] in ['CANCELED', 'REJECTED', 'EXPIRED']:
                        logger.warning(f"Limit order {order_id} for {symbol} was {order_status['status']}")
                        # Cleanup
                        self.cleanup_symbol(symbol)
                        return
                    
                    # Still pending, wait and check again
                    await asyncio.sleep(check_interval)
                    elapsed += check_interval
                    
                except asyncio.TimeoutError:
                    logger.warning(f"Timeout checking limit order status for {symbol}")
                    await asyncio.sleep(check_interval)
                    elapsed += check_interval
                except Exception as e:
                    logger.error(f"Error checking limit order {order_id} for {symbol}: {e}")
                    
                    # PHASE 3: Send CRITICAL alert on monitoring task failure
                    await self.alerts.send_alert(
                        "CRITICAL",
                        f"LIMIT ORDER MONITORING ERROR: {symbol}",
                        f"Monitoring task encountered an error!\n\n"
                        f"<b>Details:</b>\n"
                        f"• Symbol: {symbol}\n"
                        f"• Order ID: {order_id}\n"
                        f"• Error: {str(e)}\n"
                        f"• Status: Retrying...\n\n"
                        f"<b>⚠️ Position may be unprotected if this persists!</b>",
                        symbol=symbol
                    )
                    
                    await asyncio.sleep(check_interval)
                    elapsed += check_interval
            
            # Timeout - order not filled within 1 hour
            if not filled:
                logger.warning(f"Limit order {order_id} for {symbol} not filled within 1 hour. Canceling...")
                # Cancel the order with force_cancel_attempt=True to bypass monitoring client "not found" issues
                cancel_success = await self.cancel_order_safe(symbol, order_id, force_cancel_attempt=True)
                
                if cancel_success:
                    logger.info(f"✓ Limit order {order_id} for {symbol} canceled successfully due to timeout")
                    # Only cleanup after successful cancellation
                    self.cleanup_symbol(symbol)
                else:
                    # CRITICAL: Cancellation failed - DO NOT cleanup!
                    logger.error(f"❌ CRITICAL: Failed to cancel limit order {order_id} for {symbol} after all retries")
                    
                    # Send CRITICAL alert - manual intervention required
                    await self.alerts.send_alert(
                        "CRITICAL",
                        f"🚨 ORDER CANCELLATION FAILED: {symbol}",
                        f"<b>CRITICAL:</b> Limit order could not be canceled after timeout!\\n\\n"
                        f"<b>Details:</b>\\n"
                        f"• Symbol: {symbol}\\n"
                        f"• Order ID: {order_id}\\n"
                        f"• Entry Price: {entry_price}\\n"
                        f"• Quantity: {quantity}\\n"
                        f"• Direction: {direction}\\n\\n"
                        f"<b>⚠️ RISK:</b> Order may still fill on exchange!\\n"
                        f"<b>ACTION REQUIRED:</b>\\n"
                        f"1. Check Binance for order status\\n"
                        f"2. Manually cancel if still active\\n"
                        f"3. Monitor for unexpected fills\\n\\n"
                        f"<b>Bot Status:</b> Keeping order tracked (NOT cleaned up)",
                        symbol=symbol
                    )
                    
                    # DO NOT cleanup - keep tracking the order in case it fills later
                    logger.warning(f"⚠️ Keeping {symbol} in tracking - order may still fill on exchange")
                    return  # Exit without cleanup
                
        except asyncio.CancelledError:
            logger.info(f"Limit order monitoring for {symbol} was cancelled")
            # Cleanup
            self.cleanup_symbol(symbol)
        except Exception as e:
            logger.error(f"Unexpected error monitoring limit order for {symbol}: {e}")
            # Cleanup
            self.cleanup_symbol(symbol)
    
    
    # --- HELPER FUNCTIONS FOR ALGO ORDER MANAGEMENT ---
    
    async def verify_order_active(self, symbol: str, algo_id: int, max_attempts: int = 3) -> bool:
        """Verify an algo order is active on the exchange"""
        for attempt in range(max_attempts):
            try:
                # FIX: Use futures_get_algo_order for algo orders
                order = await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(self.executor, partial(
                        self.monitoring_client.futures_get_algo_order,
                        algoId=algo_id  # Use algoId parameter for algo orders
                    )),
                    timeout=15
                )
                
                # Check algo order status
                if order.get('algoStatus') in ['WORKING', 'NEW']:
                    return True
                else:
                    logger.warning(f"Algo order {algo_id} status: {order.get('algoStatus')}")
                    return False
                    
            except BinanceAPIException as e:
                if e.code == -2011:  # Order not found
                    logger.error(f"Algo order {algo_id} not found on exchange")
                    return False
                logger.warning(f"Attempt {attempt+1}: Error verifying algo order: {e}")
                if attempt < max_attempts - 1:
                    await asyncio.sleep(1)
            except Exception as e:
                logger.error(f"Error verifying algo order {algo_id}: {e}")
                if attempt < max_attempts - 1:
                    await asyncio.sleep(1)
        
        return False

    async def check_order_status(self, symbol: str, algo_id: int) -> bool:
        """Check if an algo order (TP/SL) is filled.
        
        Args:
            symbol: Trading symbol
            algo_id: Algo order ID to check
            
        Returns:
            True if order is filled, False otherwise
        """
        try:
            order = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, partial(
                    self.monitoring_client.futures_get_algo_order,
                    algoId=algo_id
                )),
                timeout=15
            )
            
            # Check if algo order is filled
            algo_status = order.get('algoStatus', '')
            if algo_status == 'FILLED':
                logger.info(f"Algo order {algo_id} for {symbol} is FILLED")
                return True
            elif algo_status in ['CANCELLED', 'EXPIRED']:
                logger.warning(f"Algo order {algo_id} for {symbol} is {algo_status}")
                return False
            else:
                logger.debug(f"Algo order {algo_id} status: {algo_status}")
                return False
                
        except BinanceAPIException as e:
            if e.code == -2013:  # Order does not exist (likely filled and removed)
                logger.info(f"Algo order {algo_id} for {symbol} not found - assuming filled")
                return True
            elif e.code == -2011:  # Order not found
                logger.info(f"Algo order {algo_id} for {symbol} not found - assuming filled")
                return True
            logger.warning(f"Error checking algo order {algo_id}: {e}")
            return False
        except asyncio.TimeoutError:
            logger.warning(f"Timeout checking algo order {algo_id} for {symbol}")
            return False
        except Exception as e:
            logger.error(f"Unexpected error checking algo order {algo_id}: {e}")
            return False

    async def cancel_algo_order_safe(self, symbol: str, algo_id: int) -> bool:
        """Safely cancel an algo order with proper error handling"""
        try:
            # FIX: Use futures_cancel_algo_order for algo orders
            await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, partial(
                    self.trading_client.futures_cancel_algo_order,
                    algoId=algo_id  # Use algoId parameter
                )),
                timeout=15
            )
            logger.info(f"Canceled algo order {algo_id} for {symbol}")
            return True
                
        except BinanceAPIException as e:
            if e.code == -2011:  # Order not found
                logger.info(f"Algo order {algo_id} not found (already processed)")
                return True
            logger.error(f"Error canceling algo order {algo_id}: {e}")
            return False
        except Exception as e:
            logger.error(f"Unexpected error canceling algo order {algo_id}: {e}")
            return False

    async def cancel_all_pending_orders(self, symbol: str):
        """Cancel all pending orders for a symbol (both entry orders and TP/SL algo orders)"""
        logger.info(f"Canceling all pending orders for {symbol}")
        async with self.global_lock:
            # Cancel entry orders (regular limit/market orders)
            if symbol in self.placed_order_ids:
                for order_id in list(self.placed_order_ids[symbol]):
                    try:
                        await asyncio.get_running_loop().run_in_executor(
                            self.executor,
                            partial(self.trading_client.futures_cancel_order, symbol=symbol, orderId=order_id)
                        )
                        logger.info(f"Canceled entry order {order_id} for {symbol}")
                    except BinanceAPIException as e:
                        if e.code == -2011:  # Order not found or already canceled/filled
                            logger.info(f"Entry order {order_id} for {symbol} not found (already processed)")
                        else:
                            logger.error(f"Failed to cancel entry order {order_id} for {symbol}: {e}")
                del self.placed_order_ids[symbol]

            # Cancel TP orders (algo/conditional)
            if symbol in self.tp_orders:
                for tp_key, algo_id in list(self.tp_orders[symbol].items()):
                    try:
                        await asyncio.get_running_loop().run_in_executor(
                            self.executor,
                            partial(self.trading_client.futures_cancel_algo_order, symbol=symbol, algoId=algo_id)
                        )
                        logger.info(f"Canceled {tp_key} for {symbol}")
                    except BinanceAPIException as e:
                        if e.code == -2011:  # Algo order not found or already triggered
                            logger.info(f"Algo order {algo_id} for {symbol} not found (already processed)")
                        else:
                            logger.error(f"Failed to cancel {tp_key} algo order {algo_id} for {symbol}: {e}")
                del self.tp_orders[symbol]

            # Cancel SL order (algo/conditional)
            if symbol in self.sl_orders:
                algo_id = self.sl_orders[symbol]
                try:
                    await asyncio.get_running_loop().run_in_executor(
                        self.executor,
                        partial(self.trading_client.futures_cancel_algo_order, symbol=symbol, algoId=algo_id)
                    )
                    logger.info(f"Canceled SL for {symbol}")
                except BinanceAPIException as e:
                    if e.code == -2011:  # Algo order not found or already triggered
                        logger.info(f"Algo order {algo_id} for {symbol} not found (already processed)")
                    else:
                        logger.error(f"Failed to cancel SL algo order {algo_id} for {symbol}: {e}")
                del self.sl_orders[symbol]

            try:
                await self._save_state()
            except Exception:
                logger.exception("Failed to save state after canceling orders")

    async def emergency_close_position(self, symbol: str, close_side: str, quantity: float):
        """Emergency close position if TP/SL placement fails"""
        logger.error(f"🚨 EMERGENCY: Closing position for {symbol} due to failed protective orders")
        
        # FIX: Cancel ALL pending orders FIRST (Issue #3 - Race Condition Fix)
        try:
            await self.cancel_all_pending_orders(symbol)
            logger.info(f"✓ Canceled all pending orders for {symbol} before emergency close")
        except Exception as e:
            logger.error(f"Error canceling pending orders during emergency close: {e}")
        
        # THEN close at market
        try:
            close_order = await asyncio.wait_for(
                asyncio.get_running_loop().run_in_executor(self.executor, partial(
                    self.trading_client.futures_create_order,
                    symbol=symbol,
                    side=close_side,
                    type='MARKET',
                    quantity=quantity,
                    reduceOnly=True
                )),
                timeout=15
            )
            
            logger.info(f"Emergency close order placed for {symbol}: {close_order['orderId']}")
            
            # Send CRITICAL alert
            await self.alerts.send_alert(
                "CRITICAL",
                f"EMERGENCY CLOSE: {symbol}",
                f"Position closed at market due to failed protective orders.\\n\\n"
                f"<b>Details:</b>\\n"
                f"• Symbol: {symbol}\\n"
                f"• Quantity: {quantity}\\n"
                f"• Reason: SL placement failed after all retries",
                symbol=symbol
            )
            
            # Record as emergency close
            await self.update_journal({
                'Symbol': symbol,
                'Close Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Status': 'Emergency Close - Failed TP/SL'
            })
            
            # Cleanup
            self.cleanup_symbol(symbol)
            return True
            
        except Exception as e:
            logger.critical(f"CRITICAL: Emergency close FAILED for {symbol}: {e}")
            
            # Send CRITICAL alert for failed emergency close
            await self.alerts.send_alert(
                "CRITICAL",
                f"EMERGENCY CLOSE FAILED: {symbol}",
                f"<b>CRITICAL:</b> Emergency close attempt FAILED!\\n\\n"
                f"<b>Details:</b>\\n"
                f"• Symbol: {symbol}\\n"
                f"• Error: {str(e)}\\n"
                f"• Action: Manual intervention required!",
                symbol=symbol
            )
            return False

    def cleanup_symbol(self, symbol: str):
        """Clean up all tracking data for a symbol after position close"""
        # Remove from position tracking
        self.position_info.pop(symbol, None)
        
        # Remove TP/SL orders
        self.tp_orders.pop(symbol, None)
        self.sl_orders.pop(symbol, None)
        
        # Remove from placed orders
        self.placed_order_ids.pop(symbol, None)
        
        # Release position slot if held
        if symbol in self.symbols_holding_slots:
            self.symbols_holding_slots.remove(symbol)
            try:
                self.open_positions_semaphore.release()
                logger.info(f"Released position slot for {symbol}")
            except ValueError:
                # Semaphore wasn't acquired, ignore
                pass
        
        logger.info(f"Cleaned up tracking data for {symbol}")

    # --- END HELPER FUNCTIONS ---


    async def place_sl_order(self, symbol: str, side: str, stop_price: float, qty_precision: int, price_precision: int, tick_size: float, quantity: float):
        try:
            # Check if SL price is valid (must be achievable)
            if stop_price <= 0:
                logger.error(f"Invalid SL price {stop_price} for {symbol}. Cannot place SL order.")
                return None
            
            stop_price = self.round_to_tick_size(stop_price, tick_size)
            stop_price_str = "{:.{}f}".format(stop_price, price_precision)
            qty_str = "{:.{}f}".format(quantity, qty_precision)

            # BULLETPROOFING: Increased retries to 5 with exponential backoff
            for attempt in range(5):
                try:
                    # FIX: Use Algo Order API for STOP_MARKET orders (Binance API update Dec 2025)
                    # API error -4120 requires using futures_create_algo_order instead of futures_create_order
                    sl_order = await asyncio.wait_for(
                        asyncio.get_running_loop().run_in_executor(self.executor, partial(
                            self.trading_client.futures_create_algo_order,  # Changed from futures_create_order
                            symbol=symbol,
                            side=side,
                            type='STOP_MARKET',
                            algoType='CONDITIONAL',  # Required for algo orders
                            triggerPrice=stop_price_str,  # Changed from stopPrice
                            closePosition='true',  # Automatically closes entire position
                            workingType='MARK_PRICE'
                            # No 'price' parameter - not allowed with STOP_MARKET + closePosition
                            # No 'timeInForce' - not applicable for algo orders
                        )),
                        timeout=15
                    )
                    # FIX: Algo orders return 'algoId' instead of 'orderId'
                    order_id = sl_order.get('algoId') or sl_order.get('orderId')
                    if not order_id:
                        logger.error(f"No order ID returned from SL placement: {sl_order}")
                        continue
                    logger.info(f"Stop Loss (Algo Order) placed at {stop_price_str} for quantity {qty_str}, algoId: {order_id}")
                    
                    # BULLETPROOFING: Verify order is active
                    verified = await self.verify_order_active(symbol, order_id)
                    if verified:
                        self.sl_orders[symbol] = order_id
                        logger.info(f"✓ SL order verified active for {symbol}")
                        try:
                            await self._save_state()
                        except Exception:
                            logger.exception("Failed to save state after placing SL")
                        return order_id
                    else:
                        logger.warning(f"SL order placed but not verified. Retrying...")
                        # Cancel the unverified order
                        await self.cancel_algo_order_safe(symbol, order_id)
                        if attempt < 4:
                            await asyncio.sleep(2 ** attempt)  # Exponential backoff
                        
                except BinanceAPIException as e:
                    if e.code == -4120:  # STOP_ORDER_SWITCH_ALGO - Must use Algo Service
                        logger.error(
                            f"Error -4120: Conditional orders must use Algo Service. "
                            f"Ensure python-binance >= 1.0.33 and using futures_create_algo_order(). "
                            f"This error indicates the order was routed to the wrong endpoint."
                        )
                        raise  # Don't retry, this is a configuration issue
                    elif e.code == -2022:  # ReduceOnly rejected
                        logger.warning(f"Attempt {attempt+1}/5: SL rejected (code -2022). Retrying...")
                        await asyncio.sleep(2 ** attempt)  # Exponential backoff: 1s, 2s, 4s, 8s, 16s
                    elif e.code == -1021:  # Timestamp error
                        logger.warning(f"Timestamp error. Syncing time and retrying...")
                        await asyncio.sleep(1)
                    else:
                        logger.error(f"SL placement error (code {e.code}): {e}")
                        if attempt >= 4:
                            raise
                        await asyncio.sleep(2)
                except asyncio.TimeoutError:
                    logger.error(f"Attempt {attempt+1}/5: Timeout placing SL")
                    if attempt < 4:
                        await asyncio.sleep(2)
                        
            logger.error(f"CRITICAL: Failed to place SL at {stop_price_str} after 5 attempts.")
            return None
        except asyncio.TimeoutError:
            logger.error(f"Timeout placing stop loss for {symbol} after 15 seconds.")
            return None
        except Exception as e:
            logger.error(f"Error placing stop loss for {symbol}: {e}")
            return None

    # --- IMPROVEMENT 8: Simplified TP strategy for market orders (2x 50% TP) ---
    async def place_tp_orders(self, symbol: str, opposite_side: str, quantity: float, targets: list[float], qty_precision: int, price_precision: int, current_price: float, tick_size: float) -> bool:
        """
        Places TP1 and TP2 orders (50% each).
        Note: The monitoring logic will handle the trailing SL after TP1 fill.
        BULLETPROOFING: Now returns True if successful, False otherwise.
        """
        if symbol not in self.tp_orders:
            self.tp_orders[symbol] = {}

        if quantity <= 0:
            logger.warning(f"Quantity is zero or negative ({quantity}). Skipping TP orders for {symbol}.")
            return False

        half_qty = quantity / 2
        
        # Ensure quantities are valid for the exchange
        # Note: If the symbol lot size is not divisible by 2, this might fail slightly. 
        # We'll rely on round_to_lot_size to handle it gracefully.
        symbol_info = await asyncio.wait_for(
            asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_exchange_info),
            timeout=15
        )
        symbol_filter = next((s for s in symbol_info['symbols'] if s['symbol'] == symbol), None)
        step_size = float(next(f['stepSize'] for f in symbol_filter['filters'] if f['filterType'] == 'LOT_SIZE'))

        tp1_qty = self.round_to_lot_size(half_qty, step_size, qty_precision)
        tp2_qty = self.round_to_lot_size(half_qty, step_size, qty_precision)
        
        # Adjust total quantity if rounding made it too large (shouldn't happen with floor/truncate, but safe)
        if tp1_qty + tp2_qty > quantity + step_size * 2: # Allow small tolerance for rounding
             # Simple adjustment: just use half the initial quantity, relying on round_to_lot_size
             tp1_qty = self.round_to_lot_size(quantity / 2, step_size, qty_precision)
             tp2_qty = self.round_to_lot_size(quantity - tp1_qty, step_size, qty_precision)
             logger.warning(f"Adjusting TP quantities due to rounding: TP1={tp1_qty}, TP2={tp2_qty}")


        # BULLETPROOFING: Track success for each TP
        tp1_success = False
        tp2_success = False

        # Place TP1 (First half)
        if len(targets) >= 1 and tp1_qty > 0:
            target = targets[0] # TP1 price
            target = self.round_to_tick_size(target, tick_size)
            target_str = "{:.{}f}".format(target, price_precision)
            
            # BULLETPROOFING: Better price validation with adjustment
            if (opposite_side == SIDE_SELL and target <= current_price) or \
               (opposite_side == SIDE_BUY and target >= current_price):
                logger.warning(f"TP1 price {target_str} invalid vs current {current_price}. Adjusting...")
                # Adjust to valid price
                if opposite_side == SIDE_SELL:
                    target = current_price * 0.995  # 0.5% below current
                else:
                    target = current_price * 1.005  # 0.5% above current
                target = self.round_to_tick_size(target, tick_size)
                target_str = "{:.{}f}".format(target, price_precision)
                logger.info(f"Adjusted TP1 to {target_str}")
            
            # BULLETPROOFING: Increased retries to 5 with exponential backoff
            for attempt in range(5):
                    try:
                        # FIX: Use Algo Order API for TAKE_PROFIT_MARKET orders (Binance API update Dec 2025)
                        # API error -4120 requires using futures_create_algo_order instead of futures_create_order
                        tp_order = await asyncio.wait_for(
                            asyncio.get_running_loop().run_in_executor(self.executor, partial(
                                self.trading_client.futures_create_algo_order,  # Changed from futures_create_order
                                symbol=symbol,
                                side=opposite_side,
                                type='TAKE_PROFIT_MARKET',  # Changed from 'TAKE_PROFIT' to 'TAKE_PROFIT_MARKET'
                                algoType='CONDITIONAL',  # Required for algo orders
                                triggerPrice=target_str,  # Changed from stopPrice
                                quantity=float(tp1_qty),
                                reduceOnly='true',  # String not boolean
                                workingType='MARK_PRICE'
                                # No 'price' parameter - not needed for MARKET orders
                                # No 'timeInForce' - not applicable for algo orders
                            )),
                            timeout=15
                        )
                        # FIX: Algo orders return 'algoId' instead of 'orderId'
                        order_id = tp_order.get('algoId') or tp_order.get('orderId')
                        if not order_id:
                            logger.error(f"No order ID returned from TP1 placement: {tp_order}")
                            if attempt < 4:
                                await asyncio.sleep(2 ** attempt)
                            continue
                        logger.info(f"Take Profit 1 ({tp1_qty}, {self.tp1_roi*100}% ROI) placed at: {target_str}, algoId: {order_id}")
                        
                        # BULLETPROOFING: Verify order is active
                        verified = await self.verify_order_active(symbol, order_id)
                        if verified:
                            self.tp_orders[symbol]['TP1'] = order_id
                            logger.info(f"✓ TP1 order verified active for {symbol}")
                            tp1_success = True
                            try:
                                await self._save_state()
                            except Exception:
                                logger.exception("Failed to save state after placing TP1")
                            break
                        else:
                            logger.warning(f"TP1 order placed but not verified. Retrying...")
                            await self.cancel_order_safe(symbol, order_id)
                            if attempt < 4:
                                await asyncio.sleep(2 ** attempt)
                                
                    except BinanceAPIException as e:
                        if e.code == -4120:  # STOP_ORDER_SWITCH_ALGO - Must use Algo Service
                            logger.error(
                                f"Error -4120: Conditional orders must use Algo Service. "
                                f"Ensure python-binance >= 1.0.33 and using futures_create_algo_order(). "
                                f"This error indicates the order was routed to the wrong endpoint."
                            )
                            break  # Don't retry, this is a configuration issue
                        elif e.code == -2022:
                            logger.warning(f"Attempt {attempt+1}/5: TP1 rejected (code -2022). Retrying...")
                            await asyncio.sleep(2 ** attempt)  # Exponential backoff
                        elif e.code == -1021:
                            logger.warning(f"Timestamp error. Syncing time and retrying...")
                            await asyncio.sleep(1)
                        else:
                            logger.error(f"TP1 placement error (code {e.code}): {e}")
                            if attempt >= 4:
                                break
                            await asyncio.sleep(2)
                    except asyncio.TimeoutError:
                        logger.error(f"Attempt {attempt+1}/5: Timeout placing TP1")
                        if attempt < 4:
                            await asyncio.sleep(2)

        # Place TP2 (Second half)
        if len(targets) >= 2 and tp2_qty > 0:
            target = targets[1] # TP2 price
            target = self.round_to_tick_size(target, tick_size)
            target_str = "{:.{}f}".format(target, price_precision)
            
            # BULLETPROOFING: Better price validation with adjustment
            if (opposite_side == SIDE_SELL and target <= current_price) or \
               (opposite_side == SIDE_BUY and target >= current_price):
                logger.warning(f"TP2 price {target_str} invalid vs current {current_price}. Adjusting...")
                # Adjust to valid price
                if opposite_side == SIDE_SELL:
                    target = current_price * 0.99  # 1% below current
                else:
                    target = current_price * 1.01  # 1% above current
                target = self.round_to_tick_size(target, tick_size)
                target_str = "{:.{}f}".format(target, price_precision)
                logger.info(f"Adjusted TP2 to {target_str}")
            
            # BULLETPROOFING: Increased retries to 5 with exponential backoff
            for attempt in range(5):
                    try:
                        # FIX: Use Algo Order API for TAKE_PROFIT_MARKET orders (Binance API update Dec 2025)
                        # API error -4120 requires using futures_create_algo_order instead of futures_create_order
                        tp_order = await asyncio.wait_for(
                            asyncio.get_running_loop().run_in_executor(self.executor, partial(
                                self.trading_client.futures_create_algo_order,  # Changed from futures_create_order
                                symbol=symbol,
                                side=opposite_side,
                                type='TAKE_PROFIT_MARKET',  # Changed from 'TAKE_PROFIT' to 'TAKE_PROFIT_MARKET'
                                algoType='CONDITIONAL',  # Required for algo orders
                                triggerPrice=target_str,  # Changed from stopPrice
                                quantity=float(tp2_qty),
                                reduceOnly='true',  # String not boolean
                                workingType='MARK_PRICE'
                                # No 'price' parameter - not needed for MARKET orders
                                # No 'timeInForce' - not applicable for algo orders
                            )),
                            timeout=15
                        )
                        # FIX: Algo orders return 'algoId' instead of 'orderId'
                        order_id = tp_order.get('algoId') or tp_order.get('orderId')
                        if not order_id:
                            logger.error(f"No order ID returned from TP2 placement: {tp_order}")
                            if attempt < 4:
                                await asyncio.sleep(2 ** attempt)
                            continue
                        logger.info(f"Take Profit 2 ({tp2_qty}, {self.tp2_roi*100}% ROI) placed at: {target_str}, algoId: {order_id}")
                        
                        # BULLETPROOFING: Verify order is active
                        verified = await self.verify_order_active(symbol, order_id)
                        if verified:
                            self.tp_orders[symbol]['TP2'] = order_id
                            logger.info(f"✓ TP2 order verified active for {symbol}")
                            tp2_success = True
                            try:
                                await self._save_state()
                            except Exception:
                                logger.exception("Failed to save state after placing TP2")
                            break
                        else:
                            logger.warning(f"TP2 order placed but not verified. Retrying...")
                            await self.cancel_order_safe(symbol, order_id)
                            if attempt < 4:
                                await asyncio.sleep(2 ** attempt)
                                
                    except BinanceAPIException as e:
                        if e.code == -4120:  # STOP_ORDER_SWITCH_ALGO - Must use Algo Service
                            logger.error(
                                f"Error -4120: Conditional orders must use Algo Service. "
                                f"Ensure python-binance >= 1.0.33 and using futures_create_algo_order(). "
                                f"This error indicates the order was routed to the wrong endpoint."
                            )
                            break  # Don't retry, this is a configuration issue
                        elif e.code == -2022:
                            logger.warning(f"Attempt {attempt+1}/5: TP2 rejected (code -2022). Retrying...")
                            await asyncio.sleep(2 ** attempt)  # Exponential backoff
                        elif e.code == -1021:
                            logger.warning(f"Timestamp error. Syncing time and retrying...")
                            await asyncio.sleep(1)
                        else:
                            logger.error(f"TP2 placement error (code {e.code}): {e}")
                            if attempt >= 4:
                                break
                            await asyncio.sleep(2)
                    except asyncio.TimeoutError:
                        logger.error(f"Attempt {attempt+1}/5: Timeout placing TP2")
                        if attempt < 4:
                            await asyncio.sleep(2)
        
        # BULLETPROOFING: Return success status
        if not tp1_success and not tp2_success:
            logger.error(f"CRITICAL: Failed to place ANY TP orders for {symbol}")
            return False
        elif not tp1_success or not tp2_success:
            logger.warning(f"WARNING: Only partial TP coverage for {symbol}")
            return True
        else:
            logger.info(f"✓ All TP orders placed successfully for {symbol}")
            return True

    async def cancel_tp_orders(self, symbol: str):
        if symbol in self.tp_orders:
            for tp_type, order_id in list(self.tp_orders[symbol].items()): # Iterate over a copy
                try:
                    # FIX: Use futures_get_algo_order for algo orders
                    order = await asyncio.wait_for(
                        asyncio.get_running_loop().run_in_executor(self.executor, partial(
                            self.trading_client.futures_get_algo_order,
                            algoId=order_id  # Use algoId for algo orders
                        )),
                        timeout=15
                    )
                    # FIX: Check algoStatus instead of status
                    if order.get('algoStatus') in ['WORKING', 'NEW']:
                        # FIX: Use futures_cancel_algo_order for algo orders
                        await asyncio.wait_for(
                            asyncio.get_running_loop().run_in_executor(self.executor, partial(
                                self.trading_client.futures_cancel_algo_order,
                                algoId=order_id  # Use algoId for algo orders
                            )),
                            timeout=15
                        )
                        logger.info(f"Canceled {tp_type} order {order_id} for {symbol}")
                    else:
                        logger.info(f"{tp_type} order {order_id} for {symbol} already processed: {order.get('algoStatus')}")
                except BinanceAPIException as e:
                    if e.code == -2011:
                        logger.warning(f"{tp_type} order {order_id} for {symbol} not found. Likely already processed.")
                    elif e.code == -2013:  # Order does not exist (algo order specific)
                        logger.warning(f"{tp_type} algo order {order_id} for {symbol} not found. Likely already filled/canceled.")
                    else:
                        logger.warning(f"Failed to cancel {tp_type} order {order_id}: {e}")
                except asyncio.TimeoutError:
                    logger.warning(f"Timeout canceling {tp_type} order {order_id} for {symbol} after 15 seconds.")
                finally:
                    # Remove from in-memory state after checking/attempting cancel
                    self.tp_orders[symbol].pop(tp_type, None)
            
            if not self.tp_orders.get(symbol):
                self.tp_orders.pop(symbol, None)
            
            try:
                await self._save_state()
            except Exception:
                logger.exception("Failed to save state after canceling TP orders")

    async def cancel_sl_order(self, symbol: str):
        if symbol in self.sl_orders:
            order_id = self.sl_orders.get(symbol)
            try:
                # FIX: Use futures_get_algo_order for algo orders
                order = await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(self.executor, partial(
                        self.trading_client.futures_get_algo_order,
                        algoId=order_id  # Use algoId for algo orders
                    )),
                    timeout=15
                )
                # FIX: Check algoStatus instead of status
                status = order.get('algoStatus')
                if status in ['WORKING', 'NEW']:
                    # FIX: Use futures_cancel_algo_order for algo orders
                    await asyncio.wait_for(
                        asyncio.get_running_loop().run_in_executor(self.executor, partial(
                            self.trading_client.futures_cancel_algo_order,
                            algoId=order_id  # Use algoId for algo orders
                        )),
                        timeout=15
                    )
                    logger.info(f"Canceled SL order {order_id} for {symbol}")
                    result = 'canceled'
                else:
                    logger.info(f"SL order {order_id} for {symbol} already processed: {status}")
                    result = 'filled' if status == 'FILLED' else 'processed'

                # Remove record of SL order from runtime state
                self.sl_orders.pop(symbol, None)
                try:
                    await self._save_state()
                except Exception:
                    logger.exception("Failed to save state after canceling SL order")

                return result
            except BinanceAPIException as e:
                if getattr(e, 'code', None) == -2011:
                    logger.warning(f"SL order {order_id} for {symbol} not found. Likely already processed.")
                    # Remove local record if Binance reports not found
                    self.sl_orders.pop(symbol, None)
                    try:
                        await self._save_state()
                    except Exception:
                        logger.exception("Failed to save state after removing missing SL order")
                    return 'not_found'
                elif getattr(e, 'code', None) == -2013:  # Order does not exist (algo order specific)
                    logger.warning(f"SL algo order {order_id} for {symbol} not found. Likely already filled/canceled.")
                    # Remove local record if Binance reports not found
                    self.sl_orders.pop(symbol, None)
                    try:
                        await self._save_state()
                    except Exception:
                        logger.exception("Failed to save state after removing missing SL order")
                    return 'not_found'
                else:
                    logger.warning(f"Failed to cancel SL order for {symbol}: {e}")
                    return f'error:{e}'
            except asyncio.TimeoutError:
                logger.warning(f"Timeout canceling SL order for {symbol} after 15 seconds.")
                # We can't be sure if it was filled or not canceled, safer to remove local record and let monitor catch it
                self.sl_orders.pop(symbol, None)
                try:
                    await self._save_state()
                except Exception:
                    logger.exception("Failed to save state after timeout removing SL order")
                return 'timeout'
        else:
            return 'no_sl'

    async def check_order_status(self, symbol: str, order_id: int) -> bool:
        for attempt in range(3):
            try:
                order_status = await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(self.executor, partial(
                        self.monitoring_client.futures_get_order,
                        symbol=symbol,
                        orderId=order_id
                    )),
                    timeout=15
                )
                # For TAKE_PROFIT_MARKET/STOP_MARKET, status will be FILLED upon trigger
                # We must also check for CANCELED/EXPIRED/REJECTED as the order is gone
                if order_status['status'] in ['FILLED', 'EXPIRED', 'CANCELED', 'REJECTED']:
                    # For market take profit orders, FILLED means it was triggered and closed.
                    return order_status['status'] == 'FILLED'
                return False
            except asyncio.TimeoutError:
                logger.warning(f"Attempt {attempt+1}: Timeout checking order status for {symbol} order {order_id} after 15 seconds.")
                if attempt < 2:
                    await asyncio.sleep(1)
                    continue
                return False
            except BinanceAPIException as e:
                if getattr(e, 'code', None) == -2011: # Order not found
                    logger.warning(f"Order {order_id} not found on exchange. Assuming filled/canceled.")
                    return True # Assume filled if order is not found (Binance quirk with some stop orders)
                logger.warning(f"Attempt {attempt+1}: Failed to check order status for {symbol} order {order_id}: {e}")
                if attempt < 2:
                    await asyncio.sleep(1)
                    continue
                return False

    async def is_algo_order_filled(self, symbol: str, algo_id: int) -> Tuple[bool, str]:
        """
        Check if algo order is filled. Returns (is_filled, reason).
        Handles -2013 errors intelligently by treating them as 'filled'.
        """
        try:
            order = await asyncio.get_running_loop().run_in_executor(
                self.executor,
                partial(self.monitoring_client.futures_get_algo_order, algoId=algo_id)
            )
            status = order.get('algoStatus', '').upper()
            if status == 'FILLED':
                return True, 'filled'
            elif status in ['CANCELED', 'REJECTED']:
                return True, 'canceled'
            return False, 'open'
        except BinanceAPIException as e:
            if e.code == -2013:  # Order does not exist -> likely filled or already handled
                logger.info(f"Algo order {algo_id} for {symbol} not found (-2013). Assuming filled/completed.")
                return True, 'not_found (assumed filled)'
            raise
        except Exception as e:
            logger.error(f"Error checking algo order {algo_id} for {symbol}: {e}")
            raise

    async def adjust_sl_to_breakeven(self, symbol: str, entry_price: float, remaining_qty: float, 
                                       side: str, price_precision: int, tick_size: float, qty_precision: int):
        """Adjust SL to breakeven (entry price) after TP1 fill"""
        try:
            # Cancel old SL
            old_sl_id = self.sl_orders.get(symbol)
            if old_sl_id:
                logger.info(f"Canceling old SL order {old_sl_id} for {symbol}")
                await self.cancel_sl_order(symbol)
            
            # Place new SL at entry price (breakeven)
            rounded_entry_price = self.round_to_tick_size(entry_price, tick_size)
            logger.info(f"Placing new SL at breakeven price {rounded_entry_price} for {symbol} with qty {remaining_qty}")
            
            await self.place_sl_order(
                symbol=symbol,
                side=side,
                stop_price=rounded_entry_price,
                qty_precision=qty_precision,
                price_precision=price_precision,
                tick_size=tick_size,
                quantity=remaining_qty
            )
            logger.info(f"Successfully adjusted SL to breakeven for {symbol} at {rounded_entry_price}")
        except Exception as e:
            logger.error(f"Error adjusting SL to breakeven for {symbol}: {e}")
            raise

    async def detect_close_reason(self, symbol: str) -> Tuple[str, float]:
        """
        Analyse recent trades to determine why the position was closed.
        Returns (reason: str, realized_pnl: float)
        """
        try:
            trades = await asyncio.get_running_loop().run_in_executor(
                self.executor,
                partial(self.monitoring_client.futures_account_trades, symbol=symbol, limit=10)
            )
            if not trades:
                return "Unknown (no recent trades)", 0.0

            # Trades are returned newest first
            total_pnl = 0.0
            closing_trades = []
            for trade in trades:
                pnl = float(trade.get('realizedPnl', 0.0))
                total_pnl += pnl
                if pnl != 0:  # Only trades that realized PNL are relevant for close
                    closing_trades.append(trade)

            if not closing_trades:
                return "Breakeven / Manual Close", 0.0

            # If any trade has negative PNL → likely SL hit
            if any(float(t.get('realizedPnl', 0.0)) < 0 for t in closing_trades):
                return "Stop Loss Hit", total_pnl

            # Otherwise positive PNL → TP hit
            return "Take Profit Hit", total_pnl

        except BinanceAPIException as e:
            action = self._handle_api_exception(e, f"detecting close reason for {symbol}", symbol)
            if action == 'abort':
                logger.error(f"API permission issue while fetching trades for {symbol}")
            return "Unknown (API error)", 0.0
        except Exception as e:
            logger.error(f"Unexpected error detecting close reason for {symbol}: {e}", exc_info=True)
            return "Unknown (error)", 0.0

    def record_trade_win(self):
        """Record a winning trade and reset loss streak"""
        self.consecutive_losses_count = 0
        self.last_loss_time = None
        logger.info("Trade recorded as WIN. Loss streak reset to 0.")

    def record_trade_loss(self):
        """Record a losing trade and check if pause is needed"""
        self.consecutive_losses_count += 1
        self.last_loss_time = time.time()
        logger.warning(f"Trade recorded as LOSS. Consecutive losses: {self.consecutive_losses_count}/{self.max_consecutive_losses}")
        
        if self.consecutive_losses_count >= self.max_consecutive_losses:
            pause_duration_seconds = self.consecutive_loss_pause_hours * 3600
            self.trading_paused_until = time.time() + pause_duration_seconds
            logger.critical(f"Max consecutive losses ({self.max_consecutive_losses}) reached. Pausing trading for {self.consecutive_loss_pause_hours} hours until {datetime.fromtimestamp(self.trading_paused_until).strftime('%Y-%m-%d %H:%M:%S')}")

    async def validate_signal_direction(self, symbol: str, direction: str) -> bool:
        """
        Validate if signal direction aligns with current market trend.
        Returns True if aligned (or validation disabled), False if misaligned.
        """
        if not self.enable_direction_validation:
            return True
        
        try:
            # Apply rate limiting before API call
            await self._throttle_api_request()
            
            # Fetch recent 1m klines
            klines = await asyncio.get_running_loop().run_in_executor(
                self.executor,
                partial(self.monitoring_client.futures_klines, 
                       symbol=symbol, interval='1m', limit=self.direction_klines_limit)
            )
            
            if not klines or len(klines) < self.direction_klines_limit:
                logger.warning(f"Insufficient klines data for {symbol} ({len(klines) if klines else 0}/{self.direction_klines_limit}). Skipping direction validation.")
                return True  # Allow trade if data unavailable
            
            # Extract close prices
            closes = [float(k[4]) for k in klines]
            current_price = closes[-1]
            
            # Compute indicators
            rsi = calculate_rsi(closes, self.direction_rsi_period)
            ema = calculate_ema(closes, self.direction_ema_period)
            
            # FIXED: Properly unpack MACD tuple to avoid TypeError
            macd_result = calculate_macd_histogram(
                closes, 
                self.direction_macd_fast, 
                self.direction_macd_slow, 
                self.direction_macd_signal
            )
            
            if rsi is None or ema is None or macd_result is None:
                logger.warning(f"Could not compute indicators for {symbol}. Skipping direction validation.")
                return True  # Allow trade if indicators fail
            
            # Unpack MACD result (status, value)
            macd_status, macd_value = macd_result
            
            # Enhanced logging for debugging
            logger.debug(
                f"Direction validation indicators for {symbol}: "
                f"RSI={rsi:.2f}, EMA={ema:.2f}, Price={current_price:.4f}, "
                f"MACD Status={macd_status}, MACD Value={macd_value:.4f}"
            )
            
            # Check individual indicators (using majority voting instead of strict AND)
            signal_direction = direction.upper()
            
            # For LONG signals
            if signal_direction in ['LONG', 'BUY']:
                rsi_bullish = rsi > self.direction_rsi_bullish_threshold
                price_above_ema = current_price > ema
                macd_bullish = macd_value > 0
                
                # Count how many indicators are bullish (majority voting: 2 out of 3)
                bullish_count = sum([rsi_bullish, price_above_ema, macd_bullish])
                
                logger.info(
                    f"LONG signal validation for {symbol}: "
                    f"RSI {'✓' if rsi_bullish else '✗'} ({rsi:.2f} {'>' if rsi_bullish else '<='} {self.direction_rsi_bullish_threshold}), "
                    f"Price vs EMA {'✓' if price_above_ema else '✗'} ({current_price:.4f} {'>' if price_above_ema else '<='} {ema:.4f}), "
                    f"MACD {'✓' if macd_bullish else '✗'} ({macd_status}, {macd_value:.4f} {'>' if macd_bullish else '<='} 0) "
                    f"→ {bullish_count}/3 bullish indicators"
                )
                
                # Require at least 2 out of 3 indicators to agree (majority voting)
                if bullish_count < 2:
                    if self.direction_alert_on_skip:
                        await self.alerts.send_alert(
                            "INFO",
                            f"🚫 Signal Skipped: Wrong Direction",
                            f"<b>LONG signal misaligned with market trend</b>\n\n"
                            f"• Symbol: {symbol}\n"
                            f"• Signal Direction: LONG\n"
                            f"• Bullish Indicators: {bullish_count}/3\n\n"
                            f"<b>Indicator Details:</b>\n"
                            f"• RSI: {rsi:.2f} {'✓ Bullish' if rsi_bullish else f'✗ Not bullish (need >{self.direction_rsi_bullish_threshold})'}\n"
                            f"• Price vs EMA: {current_price:.4f} vs {ema:.4f} {'✓ Above' if price_above_ema else '✗ Below'}\n"
                            f"• MACD: {macd_status}, Value={macd_value:.4f} {'✓ Positive' if macd_bullish else '✗ Negative'}\n\n"
                            f"⚠️ Need at least 2/3 indicators bullish. Skipping to avoid counter-trend loss."
                        )
                    logger.info(f"Skipping {symbol} LONG signal: Only {bullish_count}/3 indicators bullish (need 2/3)")
                    return False
            
            # For SHORT signals
            elif signal_direction in ['SHORT', 'SELL']:
                rsi_bearish = rsi < self.direction_rsi_bearish_threshold
                price_below_ema = current_price < ema
                macd_bearish = macd_value < 0
                
                # Count how many indicators are bearish (majority voting: 2 out of 3)
                bearish_count = sum([rsi_bearish, price_below_ema, macd_bearish])
                
                logger.info(
                    f"SHORT signal validation for {symbol}: "
                    f"RSI {'✓' if rsi_bearish else '✗'} ({rsi:.2f} {'<' if rsi_bearish else '>='} {self.direction_rsi_bearish_threshold}), "
                    f"Price vs EMA {'✓' if price_below_ema else '✗'} ({current_price:.4f} {'<' if price_below_ema else '>='} {ema:.4f}), "
                    f"MACD {'✓' if macd_bearish else '✗'} ({macd_status}, {macd_value:.4f} {'<' if macd_bearish else '>='} 0) "
                    f"→ {bearish_count}/3 bearish indicators"
                )
                
                # Require at least 2 out of 3 indicators to agree (majority voting)
                if bearish_count < 2:
                    if self.direction_alert_on_skip:
                        await self.alerts.send_alert(
                            "INFO",
                            f"🚫 Signal Skipped: Wrong Direction",
                            f"<b>SHORT signal misaligned with market trend</b>\n\n"
                            f"• Symbol: {symbol}\n"
                            f"• Signal Direction: SHORT\n"
                            f"• Bearish Indicators: {bearish_count}/3\n\n"
                            f"<b>Indicator Details:</b>\n"
                            f"• RSI: {rsi:.2f} {'✓ Bearish' if rsi_bearish else f'✗ Not bearish (need <{self.direction_rsi_bearish_threshold})'}\n"
                            f"• Price vs EMA: {current_price:.4f} vs {ema:.4f} {'✓ Below' if price_below_ema else '✗ Above'}\n"
                            f"• MACD: {macd_status}, Value={macd_value:.4f} {'✓ Negative' if macd_bearish else '✗ Positive'}\n\n"
                            f"⚠️ Need at least 2/3 indicators bearish. Skipping to avoid counter-trend loss."
                        )
                    logger.info(f"Skipping {symbol} SHORT signal: Only {bearish_count}/3 indicators bearish (need 2/3)")
                    return False
            
            logger.info(
                f"✅ Direction validation passed for {symbol} {signal_direction} signal "
                f"(RSI={rsi:.2f}, Price vs EMA: {current_price:.4f} vs {ema:.4f}, "
                f"MACD Status={macd_status}, MACD Value={macd_value:.4f})"
            )
            return True
            
        except Exception as e:
            # FIXED: Return False to skip trade on error (safer than allowing)
            logger.error(f"Error during direction validation for {symbol}: {e}", exc_info=True)
            if self.direction_alert_on_skip:
                await self.alerts.send_alert(
                    "WARNING",
                    f"⚠️ Direction Validation Error",
                    f"<b>Failed to validate signal direction</b>\n\n"
                    f"• Symbol: {symbol}\n"
                    f"• Direction: {direction}\n"
                    f"• Error: {str(e)}\n\n"
                    f"⚠️ Skipping trade for safety"
                )
            return False  # Skip trade on error to avoid risky trades

    async def handle_full_close(self, symbol: str, direction: str, avg_entry_price: float, 
                                 tp1_filled: bool, tp2_filled: bool) -> str:
        """
        Handle full position close with proper detection of close reason.
        Returns: 'tp', 'sl', 'manual', or 'unknown'
        """
        try:
            # Fetch latest position info to calculate actual PNL
            try:
                positions = await asyncio.get_running_loop().run_in_executor(
                    self.executor, 
                    partial(self.monitoring_client.futures_position_information, symbol=symbol)
                )
                pos = next((p for p in positions if p['symbol'] == symbol), None)
                if pos:
                    exit_price = float(pos.get('markPrice', 0.0))  # Use mark price if available
                    pnl = float(pos.get('unRealizedProfit', 0.0))  # Actual unrealized PNL at close
                else:
                    exit_price = 0.0
                    pnl = 0.0
            except Exception:
                logger.exception(f"Failed to fetch position for PNL calculation on close for {symbol}")
                exit_price = 0.0
                pnl = 0.0

            # Check recent trades to determine close reason and get actual values
            trades = await asyncio.get_running_loop().run_in_executor(
                self.executor, 
                partial(self.monitoring_client.futures_account_trades, symbol=symbol, limit=10)
            )
            
            if not trades:
                logger.warning(f"No recent trades found for {symbol}. Cannot determine close reason.")
                await self.alerts.send_alert(
                    "WARNING", 
                    "⚠️ Position Closed", 
                    f"<b>Position closed</b>\n\n"
                    f"• Symbol: {symbol}\n"
                    f"• Reason: Unknown\n\n"
                    f"ℹ️ No recent trades found",
                    symbol=symbol
                )
                return 'unknown'
            
            # Get the most recent closing trade
            last_trade = trades[-1]
            close_price = float(last_trade['price'])
            is_reduce_only = last_trade.get('reduceOnly', False)
            realized_pnl = float(last_trade.get('realizedPnl', 0.0))  # Extract actual PNL
            
            # Use exit price from trade if position data wasn't available
            if exit_price == 0.0:
                exit_price = close_price
            
            # Calculate PNL if not available from position
            if pnl == 0.0 and realized_pnl != 0.0:
                pnl = realized_pnl
            elif pnl == 0.0:
                # Fallback calculation
                qty = self.position_info.get(symbol, {}).get('total_qty', 0.0)
                if direction == SIDE_BUY:
                    pnl = qty * (exit_price - avg_entry_price)
                else:
                    pnl = qty * (avg_entry_price - exit_price)
            
            # Determine close type
            close_reason = 'unknown'
            if is_reduce_only:
                if pnl > 0 or tp1_filled or tp2_filled:
                    close_reason = 'tp'
                else:
                    close_reason = 'sl'
            else:
                close_reason = 'manual'
            
            # Log the close details explicitly
            logger.info(f"Position closed: {symbol}, Reason: {close_reason}, Entry: {avg_entry_price:.4f}, Exit: {exit_price:.4f}, PNL: {pnl:.4f} USDT")
            
            # Handle based on reason
            is_loss = pnl < 0
            count_as_loss = False
            
            if close_reason == 'tp':
                # TP hit (or closed profitably via protective order)
                await self.alerts.send_alert(
                    "INFO",
                    f"✅ TP Hit: {symbol}",
                    f"<b>Take Profit triggered</b>\n\n"
                    f"• Symbol: {symbol}\n"
                    f"• Direction: {direction}\n"
                    f"• Entry: {avg_entry_price:.4f}\n"
                    f"• Exit: {exit_price:.4f}\n"
                    f"• PNL: {pnl:.4f} USDT\n\n"
                    f"🎉 Profitable trade!",
                    symbol=symbol
                )
                self.record_trade_win()
                
                # Update journal with actual PNL
                await self.update_journal({
                    'Symbol': symbol,
                    'Close Price': exit_price,
                    'PNL': pnl,
                    'Close Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': 'Closed - TP'
                })
                
                # Update symbol performance for blacklisting
                if self.enable_symbol_blacklist:
                    if symbol not in self.symbol_performance:
                        self.symbol_performance[symbol] = {'wins': 0, 'losses': 0, 'total_pnl': 0.0, 'last_updated': time.time()}
                    self.symbol_performance[symbol]['wins'] += 1
                    self.symbol_performance[symbol]['total_pnl'] += pnl
                    self.symbol_performance[symbol]['last_updated'] = time.time()
                    logger.info(f"Updated performance for {symbol}: {self.symbol_performance[symbol]['wins']}W-{self.symbol_performance[symbol]['losses']}L")
                
            elif close_reason == 'sl':
                # SL hit
                await self.alerts.send_alert(
                    "WARNING",
                    f"⚠️ SL Hit: {symbol}",
                    f"<b>Stop Loss triggered</b>\n\n"
                    f"• Symbol: {symbol}\n"
                    f"• Direction: {direction}\n"
                    f"• Entry: {avg_entry_price:.4f}\n"
                    f"• Exit: {exit_price:.4f}\n"
                    f"• PNL: {pnl:.4f} USDT\n\n"
                    f"📊 Trade recorded as loss (counted in streak)",
                    symbol=symbol
                )
                self.record_trade_loss()
                count_as_loss = True
                
                # Update journal with actual PNL
                await self.update_journal({
                    'Symbol': symbol,
                    'Close Price': exit_price,
                    'PNL': pnl,
                    'Close Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': 'Closed - SL'
                })
                
                # Update symbol performance for blacklisting
                if self.enable_symbol_blacklist:
                    if symbol not in self.symbol_performance:
                        self.symbol_performance[symbol] = {'wins': 0, 'losses': 0, 'total_pnl': 0.0, 'last_updated': time.time()}
                    self.symbol_performance[symbol]['losses'] += 1
                    self.symbol_performance[symbol]['total_pnl'] += pnl
                    self.symbol_performance[symbol]['last_updated'] = time.time()
                    logger.info(f"Updated performance for {symbol}: {self.symbol_performance[symbol]['wins']}W-{self.symbol_performance[symbol]['losses']}L")
                
            elif close_reason == 'manual':
                # Manual close (not reduce-only means it wasn't a protective order)
                result_text = "Loss" if is_loss else ("Win" if pnl > 0 else "Breakeven")
                streak_text = ""
                
                # Check if we should count manual losses
                if is_loss and self.count_losses_on_manual_close:
                    streak_text = " (counted in streak)"
                    self.record_trade_loss()
                    count_as_loss = True
                    logger.info(f"Manual loss counted for streak: {symbol}")
                elif is_loss:
                    streak_text = " Not counted in loss streak"
                elif pnl > 0:
                    # Manual win - reset streak
                    self.record_trade_win()
                
                await self.alerts.send_alert(
                    "WARNING",
                    f"🛑 Manual Close: {symbol}",
                    f"<b>Position manually closed</b>\n\n"
                    f"• Symbol: {symbol}\n"
                    f"• Direction: {direction}\n"
                    f"• Entry: {avg_entry_price:.4f}\n"
                    f"• Exit: {exit_price:.4f}\n"
                    f"• PNL: {pnl:.4f} USDT\n"
                    f"• Result: {result_text}{streak_text}\n\n"
                    f"<i>Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>",
                    symbol=symbol
                )
                
                # Update journal with actual PNL
                await self.update_journal({
                    'Symbol': symbol,
                    'Close Price': exit_price,
                    'PNL': pnl,
                    'Close Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': f'Closed - Manual ({result_text})'
                })
                
                # Update symbol performance for blacklisting (manual closes count too)
                if self.enable_symbol_blacklist:
                    if symbol not in self.symbol_performance:
                        self.symbol_performance[symbol] = {'wins': 0, 'losses': 0, 'total_pnl': 0.0, 'last_updated': time.time()}
                    if pnl > 0:
                        self.symbol_performance[symbol]['wins'] += 1
                    elif pnl < 0:
                        self.symbol_performance[symbol]['losses'] += 1
                    self.symbol_performance[symbol]['total_pnl'] += pnl
                    self.symbol_performance[symbol]['last_updated'] = time.time()
                    logger.info(f"Updated performance for {symbol}: {self.symbol_performance[symbol]['wins']}W-{self.symbol_performance[symbol]['losses']}L")
            else:
                # Unknown close reason
                await self.alerts.send_alert(
                    "WARNING",
                    f"❓ Unknown Close: {symbol}",
                    f"<b>Position closed (reason unknown)</b>\n\n"
                    f"• Symbol: {symbol}\n"
                    f"• Direction: {direction}\n"
                    f"• Entry: {avg_entry_price:.4f}\n"
                    f"• Exit: {exit_price:.4f}\n"
                    f"• PNL: {pnl:.4f} USDT",
                    symbol=symbol
                )
            
            # Record last trade closed time
            self.last_trade_closed_time = time.time()
            logger.info(f"Recorded last trade closed time: {self.last_trade_closed_time}")
            
            # Save state after close
            await self._save_state()
            
            return close_reason
                
        except Exception as e:
            logger.error(f"Error determining close reason for {symbol}: {e}")
            await self.alerts.send_alert(
                "WARNING",
                "⚠️ Error Detecting Close",
                f"Could not determine close reason for {symbol}: {e}",
                symbol=symbol
            )
            return 'error'

    async def calculate_position_info(self, symbol: str, filled_entries: List[Tuple[float, float]]) -> Tuple[float, float]:
        total_qty = 0.0
        total_cost = 0.0
        for qty, price in filled_entries:
            total_qty += qty
            total_cost += qty * price
        avg_entry_price = total_cost / total_qty if total_qty > 0 else 0.0
        return total_qty, avg_entry_price

    async def verify_position_closed(self, symbol: str) -> bool:
        """Verify if position is actually closed through multiple checks"""
        try:
            # Apply rate limiting before API call
            await self._throttle_api_request()
            
            # Check 1: Get all positions and look for our symbol
            try:
                positions = await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_position_information),
                    timeout=15
                )
            except BinanceAPIException as e:
                action = await self._handle_api_exception(e, "verifying position closed (fetching positions)", symbol)
                if action == 'abort':
                    logger.error(f"Cannot verify position closure for {symbol} due to API permission error")
                    return False  # Assume position still open to be safe
                # For retry/continue, we'll return False to be safe
                return False
            
            # Find our position
            position = None
            for pos in positions:
                if pos['symbol'] == symbol:
                    position = pos
                    break
            
            if not position:
                logger.debug(f"No position data found for {symbol} in position information")
                return True  # No position data means position is closed
            
            position_amt = float(position['positionAmt'])
            logger.debug(f"Position amount for {symbol}: {position_amt}")
            
            if position_amt == 0:
                # Check 2: Verify no open TP/SL orders exist
                try:
                    open_orders = await asyncio.wait_for(
                        asyncio.get_running_loop().run_in_executor(self.executor, partial(self.monitoring_client.futures_get_open_orders, symbol=symbol)),
                        timeout=15
                    )
                except BinanceAPIException as e:
                    action = await self._handle_api_exception(e, "verifying position closed (fetching orders)", symbol)
                    if action == 'abort':
                        logger.error(f"Cannot fetch orders for {symbol} due to API permission error")
                    # If we can't fetch orders but position is 0, assume closed
                    logger.warning(f"Position is 0 but cannot verify orders for {symbol}. Assuming closed.")
                    return True
                
                # Filter for our TP and SL orders
                protective_orders = []
                if symbol in self.tp_orders:
                    for tp_type, order_id in self.tp_orders[symbol].items():
                        protective_orders.append(order_id)
                if symbol in self.sl_orders:
                    protective_orders.append(self.sl_orders[symbol])
                
                # Check if any of our protective orders are still open
                has_open_protective_orders = False
                for order in open_orders:
                    if order['orderId'] in protective_orders and order['status'] in ['NEW', 'PARTIALLY_FILLED']:
                        has_open_protective_orders = True
                        logger.debug(f"Found open protective order {order['orderId']} for {symbol}")
                        break
                
                if not has_open_protective_orders:
                    logger.info(f"Position for {symbol} confirmed closed (quantity=0, no protective orders)")
                    return True
                else:
                    logger.warning(f"Position quantity is 0 but protective orders still exist for {symbol}")
                    # If this happens, try to cancel them
                    await self.cancel_all_pending_orders(symbol)
                    return True # Assume closed after forced cancel attempt
            else:
                logger.debug(f"Position still open for {symbol} with quantity {position_amt}")
                return False
                
        except Exception as e:
            logger.error(f"Error verifying position closure for {symbol}: {e}")
            # In case of error, assume position is still open to be safe
            return False

    async def monitor_market_position(self, symbol: str, direction: str, opposite_side: str, avg_entry_price: float, total_qty: float, qty_precision: int, price_precision: int, tick_size: float, leverage: int):
        """Monitor market position for TP1/TP2 fills and SL adjustments."""
        try:
            check_interval = 10  # Increased from 5 to 10 seconds to reduce API calls
            tp1_filled = False
            tp2_filled = False
            
            # Trailing trigger logic is handled internally now, not just on breakeven threshold
            # trailing_triggered = False # Not needed anymore, SL is adjusted only after TP1
            
            consecutive_errors = 0
            max_consecutive_errors = 3
            verification_attempts = 0
            max_verification_attempts = 10 

            logger.info(f"Starting market position monitoring for {symbol}")
            self.monitor_tasks[symbol] = asyncio.current_task()

            while True:
                try:
                    position_qty = await self.get_position_quantity(symbol, direction)
                    consecutive_errors = 0  # Reset error counter on success
                    
                    if position_qty <= 0:
                        logger.info(f"Position quantity check returned 0 for {symbol}. Verifying position status...")
                        is_really_closed = await self.verify_position_closed(symbol)
                        
                        if is_really_closed:
                            logger.info(f"Position for {symbol} verified as fully closed. Determining close reason...")
                            
                            # Use new helper method to determine close reason and send appropriate alerts
                            close_reason = await self.handle_full_close(symbol, direction, avg_entry_price, tp1_filled, tp2_filled)
                            
                            # Record win/loss based on close reason
                            if close_reason == 'sl':
                                # Already handled in handle_full_close (incremented consecutive_losses_count)
                                try:
                                    self.record_trade_loss()
                                except Exception:
                                    logger.exception("Failed to record global trade loss")
                            elif close_reason == 'tp':
                                # Already handled in handle_full_close (reset consecutive_losses_count)
                                pass
                            elif tp1_filled and not tp2_filled:
                                # Partial win (closed at breakeven after TP1)
                                logger.info(f"Position closed at breakeven after TP1 fill. Recording as win for {symbol}.")
                                self.record_trade_win()

                            # --- Update Trading Journal ---
                            close_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            status = 'Partial Win' if tp1_filled and not tp2_filled else 'Loss' if close_reason == 'sl' else 'Full Win' if close_reason == 'tp' else 'Manual Close'
                            await self.update_journal({'Symbol': symbol, 'Close Time': close_time, 'Status': status})

                            await self.cancel_all_pending_orders(symbol)
                            self.cleanup_symbol(symbol)
                            # --- IMPROVEMENT 6: Record trade closure time ---
                            self.last_trade_closed_time = time.time()
                            logger.info(f"Recorded last trade closed time: {self.last_trade_closed_time}")
                            return
                        else:
                            verification_attempts += 1
                            logger.warning(f"Position check returned 0 but verification shows position still open for {symbol} (attempt {verification_attempts}/{max_verification_attempts}).")
                            
                            if verification_attempts >= max_verification_attempts:
                                logger.error(f"Max verification attempts ({max_verification_attempts}) reached for {symbol}. Forcing cleanup to prevent infinite loop.")
                                await self.cancel_all_pending_orders(symbol)
                                self.cleanup_symbol(symbol)
                                self.last_trade_closed_time = time.time()
                                return
                            
                            await asyncio.sleep(check_interval) # Wait and re-check

                    # Fetch current price with error handling
                    try:
                        ticker = await asyncio.wait_for(
                            asyncio.get_running_loop().run_in_executor(self.executor, partial(self.monitoring_client.futures_symbol_ticker, symbol=symbol)),
                            timeout=15
                        )
                        current_price = float(ticker['price'])
                    except BinanceAPIException as e:
                        action = await self._handle_api_exception(e, "fetching ticker price", symbol)
                        if action == 'abort':
                            logger.error(f"Cannot fetch ticker for {symbol} due to critical API error. Continuing monitoring with degraded functionality.")
                        # Continue monitoring even if we can't get current price
                        consecutive_errors += 1
                        await asyncio.sleep(check_interval)
                        continue
                    except asyncio.TimeoutError:
                        logger.warning(f"Timeout fetching ticker for {symbol}. Continuing monitoring.")
                        consecutive_errors += 1
                        await asyncio.sleep(check_interval)
                        continue
                    
                    # --- IMPROVEMENT 8: ADVANCED TP LOGIC - Trail SL after TP1 ---
                    if tp1_filled and not tp2_filled:
                        # Get the latest 1-minute kline for wick check
                        klines_1m = await self.get_klines_data(symbol, '1m', limit=2)
                        if klines_1m is not None and len(klines_1m['close']) > 0:
                            last_high = klines_1m['high'][-1]
                            last_low = klines_1m['low'][-1]
                            
                            # Determine new SL price based on direction
                            if direction == SIDE_BUY:
                                # Trail SL by the previous candle's LOW
                                new_sl_price = last_low
                                # If the current SL is lower than the new trail SL, update it
                                current_sl_order_id = self.sl_orders.get(symbol)
                                if current_sl_order_id:
                                    # Fetch current SL price using algo order API
                                    try:
                                        # FIX: Use futures_get_algo_order for algo orders
                                        sl_order = await asyncio.wait_for(
                                            asyncio.get_running_loop().run_in_executor(self.executor, partial(
                                                self.monitoring_client.futures_get_algo_order,
                                                algoId=current_sl_order_id  # Use algoId for algo orders
                                            )),
                                            timeout=5
                                        )
                                        # FIX: Use triggerPrice instead of stopPrice for algo orders
                                        current_sl_price = float(sl_order.get('triggerPrice', 0.0))
                                    except Exception:
                                        current_sl_price = 0.0 # Assume 0 if error fetching
                                else:
                                    # If no SL order exists, its likely a bug or initial breakeven SL was missed.
                                    # We should always have an SL after TP1 fill.
                                    current_sl_price = avg_entry_price # Default to entry price if SL is missing
                                
                                if new_sl_price > current_sl_price:
                                    logger.info(f"Trailing SL LONG for {symbol}: New SL {new_sl_price:.4f} > Current SL {current_sl_price:.4f}. Canceling old SL and placing new one.")
                                    await self.cancel_sl_order(symbol)
                                    remaining_qty = position_qty
                                    await self.place_sl_order(
                                        symbol=symbol,
                                        side=opposite_side,
                                        stop_price=new_sl_price,
                                        qty_precision=qty_precision,
                                        price_precision=price_precision,
                                        tick_size=tick_size,
                                        quantity=remaining_qty
                                    )
                                else:
                                     logger.debug(f"Trailing SL LONG for {symbol}: Current SL {current_sl_price:.4f} is better than new trail {new_sl_price:.4f}. No change.")

                            elif direction == SIDE_SELL:
                                # Trail SL by the previous candle's HIGH
                                new_sl_price = last_high
                                
                                current_sl_order_id = self.sl_orders.get(symbol)
                                if current_sl_order_id:
                                    # Fetch current SL price using algo order API
                                    try:
                                        # FIX: Use futures_get_algo_order for algo orders
                                        sl_order = await asyncio.wait_for(
                                            asyncio.get_running_loop().run_in_executor(self.executor, partial(
                                                self.monitoring_client.futures_get_algo_order,
                                                algoId=current_sl_order_id  # Use algoId for algo orders
                                            )),
                                            timeout=5
                                        )
                                        # FIX: Use triggerPrice instead of stopPrice for algo orders
                                        current_sl_price = float(sl_order.get('triggerPrice', 0.0))
                                    except Exception:
                                        current_sl_price = float('inf') # Assume inf if error fetching
                                else:
                                    current_sl_price = avg_entry_price # Default to entry price if SL is missing

                                if new_sl_price < current_sl_price:
                                    logger.info(f"Trailing SL SHORT for {symbol}: New SL {new_sl_price:.4f} < Current SL {current_sl_price:.4f}. Canceling old SL and placing new one.")
                                    await self.cancel_sl_order(symbol)
                                    remaining_qty = position_qty
                                    await self.place_sl_order(
                                        symbol=symbol,
                                        side=opposite_side,
                                        stop_price=new_sl_price,
                                        qty_precision=qty_precision,
                                        price_precision=price_precision,
                                        tick_size=tick_size,
                                        quantity=remaining_qty
                                    )
                                else:
                                    logger.debug(f"Trailing SL SHORT for {symbol}: Current SL {current_sl_price:.4f} is better than new trail {new_sl_price:.4f}. No change.")
                        
                        else:
                            logger.warning(f"Could not get 1m klines for trailing SL adjustment on {symbol}.")


                    # Check TP1 fill - Use position-based detection
                    if not tp1_filled and position_qty > 0:
                        # Check if position quantity has reduced (indicating TP1 hit)
                        # TP1 should close 50% of position, so remaining should be ~half of total_qty
                        expected_remaining_after_tp1 = total_qty * 0.5
                        tolerance = total_qty * 0.05  # 5% tolerance for rounding
                        
                        if abs(position_qty - expected_remaining_after_tp1) < tolerance and position_qty < total_qty * 0.9:
                            tp1_filled = True
                            logger.info(f"TP1 detected for {symbol} via position quantity reduction: {total_qty} -> {position_qty}")
                            
                            # Record first win (TP1 hit means at least partial profit)
                            self.record_trade_win()

                            # --- Update Trading Journal ---
                            await self.update_journal({'Symbol': symbol, 'TP1 Filled Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})

                            # Cancel TP1 order record from memory as it is filled
                            if 'TP1' in self.tp_orders.get(symbol, {}):
                                self.tp_orders[symbol].pop('TP1', None)
                            
                            # Use new helper method to adjust SL to breakeven
                            try:
                                await self.adjust_sl_to_breakeven(
                                    symbol=symbol,
                                    entry_price=avg_entry_price,
                                    remaining_qty=position_qty,
                                    side=opposite_side,
                                    price_precision=price_precision,
                                    tick_size=tick_size,
                                    qty_precision=qty_precision
                                )
                            except Exception as e:
                                logger.error(f"Failed to adjust SL to breakeven for {symbol}: {e}")
                            
                            # ALERT: TP1 Filled + SL Moved to Breakeven
                            try:
                                profit_pct = (self.tp1_roi / leverage) * 100
                                await self.alerts.send_alert(
                                    "INFO",
                                    f"🎯 TP1 Hit: {symbol}",
                                    f"<b>First target reached!</b>\n\n"
                                    f"• Symbol: {symbol}\n"
                                    f"• Direction: {direction}\n"
                                    f"• Profit Locked: ~{profit_pct:.1f}% on 50% position\n"
                                    f"• Remaining Qty: {position_qty:.4f}\n\n"
                                    f"✅ SL moved to BREAKEVEN ({avg_entry_price:.4f})\n"
                                    f"💰 Position now RISK-FREE!",
                                    symbol=symbol
                                )
                            except Exception as e:
                                logger.error(f"Failed to send TP1 alert for {symbol}: {e}")
                        
                        # Fallback: Also check via algo order status if position-based detection didn't trigger
                        elif 'TP1' in self.tp_orders.get(symbol, {}):
                            try:
                                is_filled, reason = await self.is_algo_order_filled(symbol, self.tp_orders[symbol]['TP1'])
                                if is_filled:
                                    tp1_filled = True
                                    logger.info(f"TP1 filled for {symbol} (detected via algo order status: {reason})")
                                    
                                    # Record first win
                                    self.record_trade_win()

                                    # Update journal
                                    await self.update_journal({'Symbol': symbol, 'TP1 Filled Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})

                                    # Remove from tracking
                                    self.tp_orders[symbol].pop('TP1', None)
                                    
                                    # Adjust SL to breakeven
                                    try:
                                        await self.adjust_sl_to_breakeven(
                                            symbol=symbol,
                                            entry_price=avg_entry_price,
                                            remaining_qty=position_qty,
                                            side=opposite_side,
                                            price_precision=price_precision,
                                            tick_size=tick_size,
                                            qty_precision=qty_precision
                                        )
                                    except Exception as e:
                                        logger.error(f"Failed to adjust SL to breakeven for {symbol}: {e}")
                                    
                                    # Send alert
                                    try:
                                        profit_pct = (self.tp1_roi / leverage) * 100
                                        await self.alerts.send_alert(
                                            "INFO",
                                            f"🎯 TP1 Hit: {symbol}",
                                            f"<b>First target reached!</b>\n\n"
                                            f"• Symbol: {symbol}\n"
                                            f"• Direction: {direction}\n"
                                            f"• Profit Locked: ~{profit_pct:.1f}% on 50% position\n"
                                            f"• Remaining Qty: {position_qty:.4f}\n\n"
                                            f"✅ SL moved to BREAKEVEN ({avg_entry_price:.4f})\n"
                                            f"💰 Position now RISK-FREE!",
                                            symbol=symbol
                                        )
                                    except Exception as e:
                                        logger.error(f"Failed to send TP1 alert for {symbol}: {e}")
                            except Exception as e:
                                logger.warning(f"Error checking TP1 algo order status for {symbol}: {e}")


                    # Check TP2 fill - Use improved algo order checking
                    if 'TP2' in self.tp_orders.get(symbol, {}) and not tp2_filled:
                        try:
                            is_filled, reason = await self.is_algo_order_filled(symbol, self.tp_orders[symbol]['TP2'])
                            if is_filled:
                                tp2_filled = True
                                logger.info(f"TP2 filled for {symbol} (detected via algo order status: {reason}). Position fully closed.")
                                # Record win for TP2 fill
                                self.record_trade_win()
                                
                                # Cancel remaining orders and cleanup
                                await self.cancel_all_pending_orders(symbol)
                                self.cleanup_symbol(symbol)
                                self.last_trade_closed_time = time.time()
                                logger.info(f"Position for {symbol} fully closed after TP2. Cleanup complete.")
                                
                                # ALERT: TP2 Filled + Position Closed
                                try:
                                    total_profit_pct = ((self.tp1_roi + self.tp2_roi) / 2 / leverage) * 100
                                    await self.alerts.send_alert(
                                        "INFO",
                                        f"🎉 TP2 Hit: {symbol}",
                                        f"<b>Full target reached! Trade complete.</b>\n\n"
                                        f"• Symbol: {symbol}\n"
                                        f"• Direction: {direction}\n"
                                        f"• Total Profit: ~{total_profit_pct:.1f}%\n"
                                        f"• Status: Position fully closed\n\n"
                                        f"✅ All orders cleaned up\n"
                                        f"🏆 Successful trade!",
                                        symbol=symbol
                                    )
                                except Exception as e:
                                    logger.error(f"Failed to send TP2 alert for {symbol}: {e}")
                                
                                return
                        except Exception as e:
                            logger.warning(f"Error checking TP2 algo order status for {symbol}: {e}")


                    await asyncio.sleep(check_interval)

                except asyncio.CancelledError:
                    logger.info(f"Market position monitoring for {symbol} cancelled. Preserving pending orders.")
                    self.cleanup_symbol(symbol)
                    return
                except Exception as e:
                    consecutive_errors += 1
                    logger.error(f"Error in market position monitoring for {symbol} (attempt {consecutive_errors}/{max_consecutive_errors}): {e}")
                    
                    if consecutive_errors >= max_consecutive_errors:
                        logger.error(f"Max consecutive errors ({max_consecutive_errors}) reached for {symbol}. Keeping monitoring active to preserve orders.")
                        # Exponential backoff with cap at 60 seconds
                        backoff_delay = min(check_interval * (2 ** consecutive_errors), 60)
                        await asyncio.sleep(backoff_delay)
                        consecutive_errors = 0
                    else:
                        # Exponential backoff: 10s, 20s, 40s
                        backoff_delay = check_interval * (2 ** (consecutive_errors - 1))
                        await asyncio.sleep(backoff_delay)

        except asyncio.CancelledError:
            logger.info(f"Market position monitoring for {symbol} cancelled. Preserving pending orders.")
            self.cleanup_symbol(symbol)
        except Exception as e:
            logger.error(f"Error in monitor_market_position for {symbol}: {e}")
            self.cleanup_symbol(symbol)

    def cleanup_symbol(self, symbol: str):
        """Clean up symbol-specific data without canceling orders."""
        self.tp_orders.pop(symbol, None)
        self.sl_orders.pop(symbol, None)
        self.signal_count_per_symbol.pop(symbol, None)
        self.position_info.pop(symbol, None)
        self.placed_order_ids.pop(symbol, None)
        if symbol in self.monitor_tasks and not self.monitor_tasks[symbol].done():
            self.monitor_tasks[symbol].cancel()
            self.monitor_tasks.pop(symbol, None)
        
        # If this symbol was holding an open-position slot, release it now
        # IMPROVED: Better guard to prevent semaphore over-release
        if hasattr(self, 'symbols_holding_slots') and symbol in self.symbols_holding_slots:
            try:
                self.symbols_holding_slots.remove(symbol)
                self.open_positions_semaphore.release()
                logger.info(f"Released position slot for {symbol}")
            except ValueError as e:
                # Semaphore wasn't acquired or already released
                logger.warning(f"Could not release semaphore for {symbol}: {e}")
            except Exception as e:
                logger.error(f"Unexpected error releasing semaphore for {symbol}: {e}")
        else:
            logger.debug(f"No semaphore to release for {symbol} (not in holding slots)")

        logger.info(f"Cleaned up tracking data for {symbol}")

        
    # --- NEW: Check Minimum Risk:Reward (RR) ---
    async def check_min_risk_reward(self, symbol: str, direction: str, leverage: int, entry_price: float, sl_percentage: float) -> Tuple[bool, List[float]]:
        """
        Calculates the minimum required TP to meet the MIN_RISK_REWARD (1:1.5)
        for the given entry/SL and returns the required TP prices.
        
        This assumes a market entry and calculates SL based on leverage-adjusted percentage.
        """
        
        # SL price calculated directly from SL percentage (NOT divided by leverage)
        # For LONG: entry_price * (1 - SL_PERCENTAGE/100)
        # For SHORT: entry_price * (1 + SL_PERCENTAGE/100)
        sl_pct = sl_percentage / 100.0  # Convert percentage to decimal (e.g., 2.0 -> 0.02)
        
        if direction == SIDE_BUY:
            initial_stop_price = entry_price * (1 - sl_pct)
        else: # SIDE_SELL
            initial_stop_price = entry_price * (1 + sl_pct)
            
        # Calculate max loss (Risk) in price units
        max_loss_price_units = abs(entry_price - initial_stop_price)
        
        # Calculate minimum required profit (Reward) in price units
        # Reward >= Risk * MIN_RISK_REWARD
        min_profit_price_units = max_loss_price_units * self.min_risk_reward
        
        # Calculate minimum required TP price
        if direction == SIDE_BUY:
            min_tp_price = entry_price + min_profit_price_units
        else: # SIDE_SELL
            min_tp_price = entry_price - min_profit_price_units
            
        # Check if the calculated TP2 price meets the min RR
        # TP2 is our target for the remaining 50%
        # TP ROI is already a percentage (e.g., 1.0 = 100% ROI), convert to price change
        tp2_price_change_pct = self.tp2_roi / 100.0  # Convert to decimal
        if direction == SIDE_BUY:
            target_tp2_price = entry_price * (1 + tp2_price_change_pct)
            if target_tp2_price < min_tp_price:
                logger.warning(f"{symbol} RR FAIL: Target TP2 ({target_tp2_price:.4f}) is lower than minimum required TP ({min_tp_price:.4f}) for 1:{self.min_risk_reward} RR.")
                return False, []
        else: # SIDE_SELL
            target_tp2_price = entry_price * (1 - tp2_price_change_pct)
            if target_tp2_price > min_tp_price:
                logger.warning(f"{symbol} RR FAIL: Target TP2 ({target_tp2_price:.4f}) is higher than minimum required TP ({min_tp_price:.4f}) for 1:{self.min_risk_reward} RR.")
                return False, []
                
        # Calculate TP1 price (using the ROI config)
        tp1_price_change_pct = self.tp1_roi / 100.0  # Convert to decimal
        tp1_price = entry_price * (1 + (1 if direction == SIDE_BUY else -1) * tp1_price_change_pct)
        
        logger.info(f"{symbol} passed RR Check. Risk: {max_loss_price_units:.4f}. Min Reward: {min_profit_price_units:.4f}. Target TP2: {target_tp2_price:.4f}")
        return True, [tp1_price, target_tp2_price]


    async def handle_signal(self, signal: TradingSignal) -> bool:
        symbol = signal.coin
        
        # --- IMPROVEMENT 5 & 6: Pre-Checks for Pause/Cooldown/Blocking ---
        if await self.check_trading_paused():
            logger.info(f"Trading is globally paused due to consecutive losses. Skipping signal for {symbol}.")
            return True
        
        if not self.check_trade_cooldown():
            logger.info(f"Trade cooldown active. Skipping signal for {symbol}.")
            return False

        # Check for temporary symbol block (STRATEGY 3: Asset Blacklisting)
        if not await self.check_symbol_blacklist(symbol, signal.direction):
            logger.info(f"Symbol blacklist check failed for {symbol}. Skipping signal.")
            return False

        # Check for temporary symbol block (existing logic)
        if symbol in self.blocked:
            if time.time() < self.blocked[symbol]['until']:
                logger.info(f"Symbol {symbol} is temporarily blocked until {datetime.fromtimestamp(self.blocked[symbol]['until']).strftime('%H:%M:%S')} due to {self.blocked[symbol]['reason']}. Skipping signal.")
                return False
            else:
                self.blocked.pop(symbol) # Block expired
        
        acquired_slot = False
        try:
            async with self.global_lock:
                if symbol not in self.symbol_locks:
                    self.symbol_locks[symbol] = asyncio.Lock()
                if symbol not in self.placed_order_ids:
                    self.placed_order_ids[symbol] = set()

            # ... (Existing opposite position check and closure logic - No major changes needed here) ...
            try:
                # Ensure we serialize symbol-specific operations while closing
                async with self.symbol_locks[symbol]:
                    direction = SIDE_BUY if signal.direction.upper() == 'LONG' else SIDE_SELL
                    opposite_side = SIDE_SELL if direction == SIDE_BUY else SIDE_BUY

                    # Quantity of existing position in the opposite direction
                    opposite_qty = await self.get_position_quantity(symbol, opposite_side)
                    if opposite_qty > 0:
                        logger.info(f"Received opposite signal for {symbol}. Closing existing position of {opposite_qty} at market and NOT opening a new one.")

                        closed = await self.close_existing_position_market(symbol, direction, opposite_qty)
                        if not closed:
                            logger.error(f"Failed to close existing opposite position for {symbol}. Aborting handling.")
                            return False

                        # Note: cancel_all_pending_orders is now called inside close_existing_position_market
                        # before placing the market close order to prevent race conditions

                        # Verify closure with some retries
                        verify_wait_secs = 15
                        verify_interval = 5  # Increased from 2 to 5 seconds to reduce API calls
                        elapsed = 0
                        verified = False
                        while elapsed < verify_wait_secs:
                            if await self.verify_position_closed(symbol):
                                verified = True
                                break
                            await asyncio.sleep(verify_interval)
                            elapsed += verify_interval

                        if not verified:
                            logger.warning(f"Position for {symbol} not verified closed after {verify_wait_secs}s. Manual check recommended.")
                        else:
                            logger.info(f"Existing opposite position for {symbol} confirmed closed.")

                        # Record this as a symbol-level loss event (direction change closure)
                        try:
                            self.record_trade_loss()
                        except Exception:
                            logger.exception("Failed to record global trade loss after opposite-close")

                        # Perform local cleanup for the symbol and return True to indicate
                        # the signal was handled by closing the position (no new entry).
                        try:
                            self.cleanup_symbol(symbol)
                        except Exception:
                            logger.exception("Failed during symbol cleanup after opposite-close")

                        return True
            except Exception as e:
                logger.exception(f"Error while attempting to close existing opposite position for {symbol}: {e}")
                return False
            # ... (End existing opposite position check and closure logic) ...


            # Try to acquire a slot permit quickly. If none available, queue the signal.
            try:
                await asyncio.wait_for(self.open_positions_semaphore.acquire(), timeout=0.1)
                acquired_slot = True
                logger.debug(f"Acquired slot permit for {symbol}")
            except asyncio.TimeoutError:
                logger.info(f"Max open positions ({self.max_open_positions}) reached. Queuing signal for {symbol}.")
                return False

            async with self.symbol_locks[symbol]:
                if not await self.validate_symbol(symbol):
                    logger.warning(f"Invalid or non-tradable symbol {symbol}. Skipping signal.")
                    return False

                # Re-check core limits/filters before proceeding
                open_positions = await self.get_open_positions_count()
                open_orders = await self.get_open_entry_orders_count()
                balance_ok, balance = await self.check_balance()

                if open_positions >= self.max_open_positions:
                    logger.info(f"Max open positions ({self.max_open_positions}) reached. Queuing signal for {symbol}.")
                    return False
                if open_orders + len(signal.entry_prices) > self.max_open_entry_orders:
                    logger.info(f"Max open entry orders ({self.max_open_entry_orders}) reached. Queuing signal for {symbol}.")
                    return False
                if not balance_ok:
                    logger.warning(f"Insufficient balance: {balance:.2f} USDT. Minimum required: {self.min_balance:.2f} USDT.")
                    return False

                if await self.check_trading_paused():
                    logger.info(f"Trading paused. Skipping signal for {symbol}.")
                    return False
                if not self.check_trade_frequency():
                    logger.info(f"Trade frequency limit exceeded. Skipping signal for {symbol}.")
                    return False
                if not self.check_trade_cooldown():
                    logger.info(f"Trade cooldown active. Skipping signal for {symbol}.")
                    return False

                # --- IMPROVEMENT 5: Time-Based Filter (Optional) ---
                if self.enable_time_filter:
                    if not await self.check_time_filters(symbol):
                        logger.info(f"Time-based filter failed for {symbol}. Skipping signal.")
                        return False
                else:
                    logger.debug(f"Time-based filter is disabled. Skipping time check for {symbol}.")

                # --- IMPROVEMENT 4: Candle-Wick Filter (Optional) ---
                if self.enable_candle_wick_filter:
                    if not await self.check_candle_wick_filter(symbol):
                        # Check_candle_wick_filter handles the blocking logic internally
                        return False
                else:
                    logger.debug(f"Candle wick filter is disabled. Skipping wick check for {symbol}.")

                # --- IMPROVEMENT 3: Volatility Filter (Optional) ---
                if self.enable_volatility_filter:
                    if not await self.check_volatility_filter(symbol):
                        logger.info(f"Volatility filter failed for {symbol}. Skipping signal.")
                        return False
                else:
                    logger.debug(f"Volatility filter is disabled. Skipping volatility check for {symbol}.")
                
                # --- IMPROVEMENT 7: Volume Confirmation Filter (Optional) ---
                if self.enable_volume_filter:
                    if not await self.check_volume_confirmation(symbol):
                        logger.info(f"Volume confirmation filter failed for {symbol}. Skipping signal.")
                        return False
                else:
                    logger.debug(f"Volume confirmation filter is disabled. Skipping volume check for {symbol}.")
                
                direction = SIDE_BUY if signal.direction.upper() == 'LONG' else SIDE_SELL
                opposite_side = SIDE_SELL if direction == SIDE_BUY else SIDE_BUY

                # --- STRATEGY 1: Technical Indicator Validation ---
                indicator_data = {}  # Store for later use in confidence scoring
                if self.enable_technical_indicator_filter:
                    passes_filter, indicator_data = await self.check_technical_indicators_filter(symbol, direction)
                    if not passes_filter:
                        logger.info(f"Technical indicator filter failed for {symbol}. Skipping signal.")
                        return False
                else:
                    logger.debug(f"Technical indicator filter is disabled for {symbol}.")
                
                # --- User's Simplified RSI Filter ---
                if self.enable_rsi_filter:
                    if not await self.validate_signal(signal):
                        logger.info(f"RSI filter validation failed for {symbol}. Skipping signal.")
                        return False

                # --- IMPROVEMENT 2: Trend Filter (Optional) ---
                if self.enable_trend_filter:
                    if not await self.check_trend_filter(symbol, direction):
                        logger.info(f"Trend filter failed for {symbol}. Skipping signal.")
                        return False
                else:
                    logger.debug(f"Trend filter is disabled. Skipping trend check for {symbol}.")

                # --- Direction Validation Filter (Prevent Wrong-Direction Trades) ---
                direction_validation_passed = await self.validate_signal_direction(symbol, signal.direction)
                if not direction_validation_passed:
                    logger.info(f"Direction validation failed for {symbol}. Signal misaligned with market trend. Skipping signal.")
                    return False



                # If an opposite position exists for this symbol, close it immediately at market
                # This is a redundant check after the initial closure attempt, but kept for robustness
                try:
                    opposite_qty = await self.get_position_quantity(symbol, opposite_side)
                    if opposite_qty > 0:
                        logger.info(f"Detected existing opposite position for {symbol} (Post-Filter Check). Closing {opposite_qty} at market before opening new position.")
                        closed = await self.close_existing_position_market(symbol, direction, opposite_qty)
                        if not closed:
                            logger.error(f"Failed to close existing opposite position for {symbol}. Aborting new signal handling.")
                            return False

                        # Wait briefly and verify closure
                        verify_wait_secs = 15
                        verify_interval = 5  # Increased from 2 to 5 seconds to reduce API calls
                        elapsed = 0
                        verified = False
                        while elapsed < verify_wait_secs:
                            if await self.verify_position_closed(symbol):
                                verified = True
                                break
                            await asyncio.sleep(verify_interval)
                            elapsed += verify_interval

                        if not verified:
                            logger.warning(f"Position for {symbol} not verified closed after {verify_wait_secs}s. Proceeding cautiously.")
                        else:
                            logger.info(f"Existing opposite position for {symbol} confirmed closed.")
                            try:
                                self.record_trade_loss()
                            except Exception:
                                logger.exception("Failed to record global trade loss after opposite-close (pre-open)")
                except Exception as e:
                    logger.exception(f"Error while attempting to close existing opposite position for {symbol}: {e}")
                    return False

                leverage = min(
                    int(float(signal.leverage.replace('X', ''))) if signal.leverage else self.default_leverage,
                    self.max_leverage
                )
                if symbol not in self.leverage_set or self.leverage_set[symbol] != leverage:
                    try:
                        # Apply rate limiting before leverage change
                        await self._throttle_api_request()
                        
                        await asyncio.wait_for(
                            asyncio.get_running_loop().run_in_executor(self.executor, partial(
                                self.trading_client.futures_change_leverage,
                                symbol=symbol,
                                leverage=leverage
                            )),
                            timeout=15
                        )
                        self.leverage_set[symbol] = leverage
                        logger.info(f"Set leverage for {symbol} to {leverage}X")
                    except BinanceAPIException as e:
                        logger.error(f"Failed to set leverage for {symbol}: {e}")
                        return False
                    except Exception as e:
                        logger.error(f"Failed to set leverage for {symbol}: {e}")
                        return False
                
                # Set margin type based on environment variable (ISOLATED or CROSS)
                margin_type = os.getenv('MARGIN_TYPE', 'ISOLATED').upper()
                
                # Check current margin type first to avoid unnecessary API calls
                try:
                    positions = await asyncio.get_running_loop().run_in_executor(
                        self.executor, 
                        self.monitoring_client.futures_position_information
                    )
                    current_pos = next((p for p in positions if p['symbol'] == symbol), None)
                    if current_pos and current_pos.get('marginType', '').upper() == margin_type:
                        logger.info(f"Margin already {margin_type} for {symbol}, skipping change")
                    else:
                        # Attempt to change margin type
                        try:
                            await asyncio.get_running_loop().run_in_executor(
                                self.executor,
                                partial(self.trading_client.futures_change_margin_type, symbol=symbol, marginType=margin_type)
                            )
                            logger.info(f"Successfully set margin type to {margin_type} for {symbol}")
                        except BinanceAPIException as e:
                            if e.code in [-4046, -1102]:  # -4046: already set, -1102: likely already CROSS or param issue
                                logger.info(f"Margin type change skipped for {symbol} (code {e.code}): {e.message}")
                                # Proceed with order placement - no need to abort
                            else:
                                logger.error(f"Critical error setting margin type for {symbol}: {e}")
                                await self.alerts.send_alert(
                                    "WARNING",
                                    "Margin Type Set Failed",
                                    f"Failed to set {margin_type} margin for {symbol}: {e.message}\nTrade skipped."
                                )
                                return False
                        except Exception as e:
                            logger.error(f"Failed to set margin type for {symbol}: {e}")
                            return False
                except Exception as e:
                    logger.warning(f"Could not check current margin type for {symbol}: {e}")
                    # If we can't check, try to set it anyway
                    try:
                        await asyncio.get_running_loop().run_in_executor(
                            self.executor,
                            partial(self.trading_client.futures_change_margin_type, symbol=symbol, marginType=margin_type)
                        )
                        logger.info(f"Successfully set margin type to {margin_type} for {symbol}")
                    except BinanceAPIException as e:
                        if e.code in [-4046, -1102]:  # -4046: already set, -1102: likely already CROSS or param issue
                            logger.info(f"Margin type change skipped for {symbol} (code {e.code}): {e.message}")
                            # Proceed with order placement - no need to abort
                        else:
                            logger.error(f"Critical error setting margin type for {symbol}: {e}")
                            await self.alerts.send_alert(
                                "WARNING",
                                "Margin Type Set Failed",
                                f"Failed to set {margin_type} margin for {symbol}: {e.message}\nTrade skipped."
                            )
                            return False
                    except Exception as e:
                        logger.error(f"Failed to set margin type for {symbol}: {e}")
                        return False

                exchange_info = await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_exchange_info),
                    timeout=15
                )
                symbol_info = next((s for s in exchange_info['symbols'] if s['symbol'] == symbol), None)
                if not symbol_info:
                    logger.error(f"Symbol {symbol} not found in exchange info.")
                    return False

                qty_precision = symbol_info['quantityPrecision']
                price_precision = symbol_info['pricePrecision']
                tick_size = float(next(f['tickSize'] for f in symbol_info['filters'] if f['filterType'] == 'PRICE_FILTER'))
                step_size = float(next(f['stepSize'] for f in symbol_info['filters'] if f['filterType'] == 'LOT_SIZE'))
                min_notional = float(next(f['notional'] for f in symbol_info['filters'] if f['filterType'] == 'MIN_NOTIONAL') if any(f['filterType'] == 'MIN_NOTIONAL' for f in symbol_info['filters']) else 5.0)

                ticker = await asyncio.wait_for(
                    asyncio.get_running_loop().run_in_executor(self.executor, partial(self.monitoring_client.futures_symbol_ticker, symbol=symbol)),
                    timeout=15
                )
                current_price = float(ticker['price'])

                # Handle market entry for new signal format
                if not signal.entry_prices:  # Market entry
                    logger.info(f"Processing MARKET PRICE entry for {symbol}")
                    
                    # Calculate target notional based on desired margin and leverage
                    target_notional = self.target_margin_per_trade * leverage
                    quantity = target_notional / current_price
                    notional = quantity * current_price

                    # Ensure notional meets minimum requirements
                    if notional < min_notional:
                        logger.warning(f"Notional {notional:.2f} for {symbol} below minimum {min_notional}. Adjusting quantity.")
                        # Adjusted logic to ensure the adjusted quantity respects the lot size
                        quantity = math.ceil((min_notional * 1.01) / current_price / step_size) * step_size
                        notional = quantity * current_price

                    quantity = self.round_to_lot_size(quantity, step_size, qty_precision)
                    notional = quantity * current_price
                    logger.info(f"Market entry for {symbol}: quantity={quantity}, notional={notional:.2f} USDT, margin={notional/leverage:.2f} USDT")

                    if notional > self.max_total_notional:
                        logger.warning(f"Entry for {symbol} exceeds total notional cap of {self.max_total_notional} USDT (new: {notional:.2f}). Skipping.")
                        return False

                    if notional < min_notional:
                        logger.error(f"Adjusted notional {notional:.2f} for {symbol} still below {min_notional}. Skipping entry.")
                        return False

                    margin_ok, required_margin = await self.check_margin(symbol, leverage, quantity, current_price)
                    if not margin_ok:
                        logger.warning(f"Insufficient margin for {symbol}. Required: {required_margin:.2f} USDT.")
                        return False

                    qty_str = "{:.{}f}".format(quantity, qty_precision)

                    # --- IMPROVEMENT 1: RR Check (Skip if direction validation passed) ---
                    # If direction validation passed, skip RR check as signal is high-confidence
                    if direction_validation_passed:
                        logger.info(f"Skipping RR check for {symbol}: Direction validation passed (high-confidence signal)")
                        # Use default TP targets based on ROI config
                        tp1_price_change_pct = self.tp1_roi / 100.0
                        tp2_price_change_pct = self.tp2_roi / 100.0
                        tp1_price = current_price * (1 + (1 if direction == SIDE_BUY else -1) * tp1_price_change_pct)
                        tp2_price = current_price * (1 + (1 if direction == SIDE_BUY else -1) * tp2_price_change_pct)
                        roi_targets = [tp1_price, tp2_price]
                    else:
                        # Direction validation disabled or not passed, perform RR check
                        rr_ok, roi_targets = await self.check_min_risk_reward(
                            symbol, direction, leverage, current_price, self.sl_percentage
                        )
                        if not rr_ok:
                            logger.warning(f"Trade for {symbol} rejected: Does not meet minimum 1:{self.min_risk_reward} Risk:Reward based on current market price and SL configuration.")
                            return False

                    # Place market order
                    try:
                        market_order = await asyncio.wait_for(
                            asyncio.get_running_loop().run_in_executor(self.executor, partial(
                                self.trading_client.futures_create_order,
                                symbol=symbol,
                                side=direction,
                                type='MARKET',
                                quantity=float(qty_str)
                            )),
                            timeout=15
                        )
                        order_id = market_order['orderId']
                        self.placed_order_ids[symbol].add(order_id)
                        try:
                            await self._save_state()
                        except Exception:
                            logger.exception("Failed to save state after placing market order")
                        logger.info(f"Placed market {direction} order for {symbol} with quantity {qty_str} (notional: {notional:.2f} USDT)")
                        
                        # Wait for fill and get actual entry price
                        filled = False
                        avg_entry_price = 0.0
                        total_qty = 0.0
                        for _ in range(10):  # Wait up to 10 seconds for fill
                            if await self.has_symbol_position(symbol):
                                positions = await asyncio.wait_for(
                                    asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_position_information),
                                    timeout=15
                                )
                                position = next((pos for pos in positions if pos['symbol'] == symbol), None)
                                if position:
                                    avg_entry_price = float(position['entryPrice'])
                                    total_qty = abs(float(position['positionAmt']))
                                    filled = True
                                    break
                            await asyncio.sleep(1)

                        if not filled:
                            logger.error(f"Market order for {symbol} not filled within timeout. Canceling.")
                            await self.cancel_all_pending_orders(symbol)
                            return False

                        self.position_info[symbol] = {
                            'total_qty': total_qty,
                            'avg_entry_price': avg_entry_price
                        }
                        try:
                            await self._save_state()
                        except Exception:
                            logger.exception("Failed to save state after market fill")
                        logger.info(f"Market order filled for {symbol}: avg_entry_price={avg_entry_price}, total_qty={total_qty}")
                        
                        # --- Update Trading Journal ---
                        await self.update_journal({
                            'Symbol': symbol,
                            'Direction': 'LONG' if direction == SIDE_BUY else 'SHORT',
                            'Entry Price': avg_entry_price,
                            'Entry Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'Status': 'Open'
                        })
                        
                        # Record this trade for frequency tracking
                        self.record_trade_opened()

                        # The RR check already calculated the required TP prices based on ROI config and leverage
                        # We use these calculated targets: roi_targets = [tp1_price, target_tp2_price]
                        logger.info(f"Using calculated ROI targets for {symbol}: TP1={roi_targets[0]:.{price_precision}f}, TP2={roi_targets[1]:.{price_precision}f}")

                        # BULLETPROOFING: Place TP/SL with verification and emergency close fallback
                        protection_success = await self.place_and_verify_protective_orders(
                            symbol=symbol,
                            direction=direction,
                            opposite_side=opposite_side,
                            avg_entry_price=avg_entry_price,
                            total_qty=total_qty,
                            leverage=leverage,
                            roi_targets=roi_targets,
                            qty_precision=qty_precision,
                            price_precision=price_precision,
                            tick_size=tick_size
                        )
                        
                        if not protection_success:
                            logger.error(f"Failed to place protective orders for {symbol}. Position was emergency closed.")
                            return False

                        self.signal_count_per_symbol[symbol] = self.signal_count_per_symbol.get(symbol, 0) + 1

                        # Reserve the slot for this symbol for the lifetime of the position
                        self.symbols_holding_slots.add(symbol)
                        logger.info(f"Reserved slot for {symbol} after market entry")
                        
                        # Start monitoring with market entry data
                        asyncio.create_task(self.monitor_market_position(
                            symbol=symbol,
                            direction=direction,
                            opposite_side=opposite_side,
                            avg_entry_price=avg_entry_price,
                            total_qty=total_qty,
                            qty_precision=qty_precision,
                            price_precision=price_precision,
                            tick_size=tick_size,
                            leverage=leverage  # Add leverage parameter
                        ))
                        
                        logger.info(f"Processed market signal for {symbol}, total notional: {notional:.2f} USDT. Monitoring started.")
                        return True

                    except BinanceAPIException as e:
                        logger.error(f"Failed to place market order for {symbol}: {e}")
                        return False
                    except asyncio.TimeoutError:
                        logger.error(f"Timeout placing market order for {symbol} after 15 seconds.")
                        return False

                # --- LIMIT ORDER ENTRY (New Signal Format) ---
                elif len(signal.entry_prices) == 1:
                    entry_price = signal.entry_prices[0]
                    logger.info(f"Processing LIMIT ORDER entry for {symbol} at price {entry_price}")
                    
                    # Calculate target notional based on desired margin and leverage
                    target_notional = self.target_margin_per_trade * leverage
                    quantity = target_notional / entry_price
                    notional = quantity * entry_price

                    # Ensure notional meets minimum requirements
                    if notional < min_notional:
                        logger.warning(f"Notional {notional:.2f} for {symbol} below minimum {min_notional}. Adjusting quantity.")
                        quantity = math.ceil((min_notional * 1.01) / entry_price / step_size) * step_size
                        notional = quantity * entry_price

                    quantity = self.round_to_lot_size(quantity, step_size, qty_precision)
                    notional = quantity * entry_price
                    logger.info(f"Limit entry for {symbol}: quantity={quantity}, notional={notional:.2f} USDT, margin={notional/leverage:.2f} USDT")

                    if notional > self.max_total_notional:
                        logger.warning(f"Entry for {symbol} exceeds total notional cap of {self.max_total_notional} USDT (new: {notional:.2f}). Skipping.")
                        return False

                    if notional < min_notional:
                        logger.error(f"Adjusted notional {notional:.2f} for {symbol} still below {min_notional}. Skipping entry.")
                        return False

                    margin_ok, required_margin = await self.check_margin(symbol, leverage, quantity, entry_price)
                    if not margin_ok:
                        logger.warning(f"Insufficient margin for {symbol}. Required: {required_margin:.2f} USDT.")
                        return False

                    qty_str = "{:.{}f}".format(quantity, qty_precision)
                    entry_price_str = "{:.{}f}".format(entry_price, price_precision)

                    # --- RR Check based on entry price (Skip if direction validation passed) ---
                    if direction_validation_passed:
                        logger.info(f"Skipping RR check for {symbol}: Direction validation passed (high-confidence signal)")
                        # Use default TP targets based on ROI config
                        tp1_price_change_pct = self.tp1_roi / 100.0
                        tp2_price_change_pct = self.tp2_roi / 100.0
                        tp1_price = entry_price * (1 + (1 if direction == SIDE_BUY else -1) * tp1_price_change_pct)
                        tp2_price = entry_price * (1 + (1 if direction == SIDE_BUY else -1) * tp2_price_change_pct)
                        roi_targets = [tp1_price, tp2_price]
                    else:
                        # Direction validation disabled or not passed, perform RR check
                        rr_ok, roi_targets = await self.check_min_risk_reward(
                            symbol, direction, leverage, entry_price, self.sl_percentage
                        )
                        if not rr_ok:
                            logger.warning(f"Trade for {symbol} rejected: Does not meet minimum 1:{self.min_risk_reward} Risk:Reward based on entry price and SL configuration.")
                            return False

                    # Place limit order
                    try:
                        limit_order = await asyncio.wait_for(
                            asyncio.get_running_loop().run_in_executor(self.executor, partial(
                                self.trading_client.futures_create_order,
                                symbol=symbol,
                                side=direction,
                                type=ORDER_TYPE_LIMIT,
                                timeInForce=TIME_IN_FORCE_GTC,
                                quantity=float(qty_str),
                                price=entry_price_str
                            )),
                            timeout=15
                        )
                        order_id = limit_order['orderId']
                        self.placed_order_ids[symbol].add(order_id)
                        try:
                            await self._save_state()
                        except Exception:
                            logger.exception("Failed to save state after placing limit order")
                        logger.info(f"Placed limit {direction} order for {symbol} at {entry_price_str} with quantity {qty_str} (notional: {notional:.2f} USDT)")
                        
                        # PHASE 2: Start async monitoring task (non-blocking)
                        asyncio.create_task(self.monitor_limit_order_fill(
                            symbol=symbol,
                            order_id=order_id,
                            direction=direction,
                            opposite_side=opposite_side,
                            entry_price=entry_price,
                            quantity=quantity,
                            leverage=leverage,
                            roi_targets=roi_targets,
                            qty_precision=qty_precision,
                            price_precision=price_precision,
                            tick_size=tick_size,
                            signal_timestamp=signal.timestamp
                        ))
                        
                        logger.info(f"Limit order monitoring started for {symbol}. Processing continues in background.")
                        return True

                    except BinanceAPIException as e:
                        logger.error(f"Failed to place limit order for {symbol}: {e}")
                        return False
                    except asyncio.TimeoutError:
                        logger.error(f"Timeout placing limit order for {symbol} after 15 seconds.")
                        return False

                # --- No valid entry method ---
                else:
                    logger.warning(f"Signal for {symbol} has invalid entry_prices configuration: {signal.entry_prices}. Skipping.")
                    return False

        finally:
            # If we acquired a slot but did not actually reserve it for a symbol (i.e., no position/orders placed), release it
            if acquired_slot and symbol not in getattr(self, 'symbols_holding_slots', set()):
                try:
                    self.open_positions_semaphore.release()
                    logger.debug(f"Released provisional slot permit for {symbol}")
                except ValueError:
                    logger.exception(f"Error releasing provisional semaphore for {symbol}")

    # The original monitor_orders function for limit orders is now deprecated/omitted 
    # as the new signal format uses market entry, and its logic is partially integrated
    # into the monitor_market_position for simplicity.
    
    # ------------------ Persistence Helpers ------------------
    def _state_snapshot(self) -> dict:
        return {
            'position_info': self.position_info,
            'tp_orders': self.tp_orders,
            'sl_orders': self.sl_orders,
            'placed_order_ids': {k: list(v) for k, v in self.placed_order_ids.items()},
            'symbols_holding_slots': list(getattr(self, 'symbols_holding_slots', [])),
            'leverage_set': self.leverage_set,
            'blocked': self.blocked,
            'consecutive_losses_count': self.consecutive_losses_count,
            'trading_paused_until': self.trading_paused_until,
            'last_trade_closed_time': self.last_trade_closed_time # --- IMPROVEMENT 6: Persist cooldown time ---
        }

    async def _save_state(self):
        """Asynchronously write the current state to disk."""
        try:
            state = self._state_snapshot()
            loop = asyncio.get_running_loop()
            def _write():
                with open(self.state_file, 'w', encoding='utf-8') as f:
                    json.dump(state, f, indent=2)
            await loop.run_in_executor(self.executor, _write)
        except Exception:
            logger.exception("Failed to save state to disk")

    def _load_state(self):
        """Load persisted state synchronously at startup."""
        if not os.path.exists(self.state_file):
            return
        try:
            with open(self.state_file, 'r', encoding='utf-8') as f:
                state = json.load(f)
            self.position_info = state.get('position_info', {}) or {}
            self.tp_orders = state.get('tp_orders', {}) or {}
            self.sl_orders = state.get('sl_orders', {}) or {}
            placed = state.get('placed_order_ids', {}) or {}
            self.placed_order_ids = {k: set(v) for k, v in placed.items()}
            self.symbols_holding_slots = set(state.get('symbols_holding_slots', []) or [])
            self.leverage_set = state.get('leverage_set', {}) or {}
            # Load per-symbol blocked info if present
            self.blocked = state.get('blocked', {}) or {}
            # Load consecutive losses count and pause state
            self.consecutive_losses_count = state.get('consecutive_losses_count', 0)
            self.trading_paused_until = state.get('trading_paused_until')
            # --- IMPROVEMENT 6: Load cooldown time ---
            self.last_trade_closed_time = state.get('last_trade_closed_time')
            logger.info(f"Loaded persisted state: {len(self.position_info)} positions, {len(self.tp_orders)} TP entries, {len(self.sl_orders)} SL entries, consecutive_losses_count={self.consecutive_losses_count}")
        except Exception:
            logger.exception("Failed to load persisted state file")

    async def _autosave_loop(self):
        try:
            while True:
                await asyncio.sleep(self.autosave_interval)
                try:
                    await self._save_state()
                    logger.debug("Autosaved bot state to disk")
                except Exception:
                    logger.exception("Autosave failed")
        except asyncio.CancelledError:
            logger.debug("Autosave loop cancelled")

    async def _start_autosave(self):
        if self.autosave_task is None or self.autosave_task.done():
            self.autosave_task = asyncio.create_task(self._autosave_loop())
            logger.info(f"Started autosave task (interval={self.autosave_interval}s)")

    async def _stop_autosave(self):
        if self.autosave_task and not self.autosave_task.done():
            self.autosave_task.cancel()
            try:
                await self.autosave_task
            except asyncio.CancelledError:
                pass
            logger.info("Stopped autosave task")

    async def _reconcile_state_with_exchange(self):
        """Fetch live positions and open orders from the exchange and update in-memory state accordingly.
        This helps the bot pick up existing positions/orders after a restart.
        """
        try:
            # Reset runtime state and rebuild from exchange to avoid keeping stale entries
            self.position_info = {}
            self.tp_orders = {}
            self.sl_orders = {}
            self.placed_order_ids = {}
            self.symbols_holding_slots = set()

            # Fetch positions with error handling
            try:
                positions = await asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_position_information)
            except BinanceAPIException as e:
                action = await self._handle_api_exception(e, "reconciling state (fetching positions)")
                if action == 'abort':
                    logger.error("Cannot reconcile positions due to API permission error. Skipping position reconciliation.")
                    positions = []  # Continue with empty positions
                else:
                    logger.warning("Failed to fetch positions during reconciliation. Continuing with empty state.")
                    positions = []
            except Exception as e:
                logger.error(f"Unexpected error fetching positions during reconciliation: {e}")
                positions = []
            
            for pos in positions:
                try:
                    if float(pos.get('positionAmt', 0)) != 0:
                        symbol = pos['symbol']
                        self.position_info[symbol] = {
                            'total_qty': abs(float(pos['positionAmt'])),
                            'avg_entry_price': float(pos.get('entryPrice', 0.0))
                        }
                        # Reserve slot for this symbol so bot respects max slots
                        try:
                            self.symbols_holding_slots.add(symbol)
                        except Exception:
                            logger.exception(f"Error adding symbol to symbols_holding_slots: {symbol}")
                except Exception:
                    logger.exception(f"Error processing position {pos}")

            # Fetch open orders with error handling
            try:
                open_orders = await asyncio.get_running_loop().run_in_executor(self.executor, partial(self.monitoring_client.futures_get_open_orders))
            except BinanceAPIException as e:
                action = await self._handle_api_exception(e, "reconciling state (fetching open orders)")
                if action == 'abort':
                    logger.error("Cannot reconcile open orders due to API permission error. Skipping order reconciliation.")
                    open_orders = []  # Continue with empty orders
                else:
                    logger.warning("Failed to fetch open orders during reconciliation. Continuing with empty state.")
                    open_orders = []
            except Exception as e:
                logger.error(f"Unexpected error fetching open orders during reconciliation: {e}")
                open_orders = []
            
            # Build a set of symbols which currently have positions (from above)
            symbols_with_positions = set(self.position_info.keys())
            for order in open_orders:
                try:
                    symbol = order['symbol']
                    order_id = order['orderId']
                    otype = order.get('type', '').upper()
                    reduce_only = bool(order.get('reduceOnly'))

                    # If this is a protective reduce-only order but there is no live
                    # position for the symbol, assume it is orphaned and cancel it
                    if reduce_only and symbol not in symbols_with_positions:
                        try:
                            logger.info(f"Found orphan protective order {order_id} for {symbol} (no open position). Canceling on reconcile.")
                            await asyncio.get_running_loop().run_in_executor(
                                self.executor,
                                partial(self.trading_client.futures_cancel_order, symbol=symbol, orderId=order_id)
                            )
                            logger.info(f"Canceled orphan protective order {order_id} for {symbol}")
                        except Exception:
                            logger.exception(f"Failed to cancel orphan protective order {order_id} for {symbol}")
                        # Do not add this order into memory maps
                        continue

                    if reduce_only:
                        # Protective order: SL or TP
                        if symbol not in self.tp_orders:
                            self.tp_orders[symbol] = {}
                        # Normalize stop/market types as SL, others as TP
                        if 'STOP' in otype or 'MARKET' in otype or 'STOP_MARKET' in otype:
                            # Keep only the latest SL id
                            self.sl_orders[symbol] = order_id
                        else:
                            # Assign to TP1/TP2 slots in order of discovery
                            if 'TP1' not in self.tp_orders[symbol]:
                                self.tp_orders[symbol]['TP1'] = order_id
                            elif 'TP2' not in self.tp_orders[symbol]:
                                self.tp_orders[symbol]['TP2'] = order_id
                    else:
                        # Entry order
                        self.placed_order_ids.setdefault(symbol, set()).add(order_id)
                except Exception:
                    logger.exception(f"Error processing open order {order}")

            # CRITICAL FIX: Fetch algo orders separately (post-Dec 2025 migration)
            # futures_get_open_orders() no longer returns algo orders (SL/TP)
            # We must use futures_get_open_algo_orders() to get conditional orders
            try:
                logger.info("Fetching open algo orders (SL/TP) for reconciliation...")
                try:
                    open_algo_orders = await asyncio.get_running_loop().run_in_executor(
                        self.executor, 
                        partial(self.monitoring_client.futures_get_open_algo_orders)
                    )
                    logger.info(f"Found {len(open_algo_orders)} open algo orders")
                except BinanceAPIException as e:
                    action = await self._handle_api_exception(e, "reconciling state (fetching algo orders)")
                    if action == 'abort':
                        logger.error("Cannot reconcile algo orders due to API permission error. Skipping algo order reconciliation.")
                        await self.alerts.send_alert(
                            "WARNING",
                            "Algo Order Reconciliation Failed",
                            f"Cannot fetch algo orders during reconciliation.\n"
                            f"This may indicate missing API permissions for algo trading.\n"
                            f"TP/SL orders may not be tracked correctly."
                        )
                    open_algo_orders = []  # Continue without algo orders
                except Exception as e:
                    logger.error(f"Unexpected error fetching algo orders during reconciliation: {e}")
                    open_algo_orders = []
                
                for algo_order in open_algo_orders:
                    try:
                        symbol = algo_order['symbol']
                        algo_id = algo_order.get('algoId')
                        algo_type = algo_order.get('algoType', '').upper()
                        order_type = algo_order.get('type', '').upper()  # STOP_MARKET or TAKE_PROFIT_MARKET
                        
                        if not algo_id:
                            logger.warning(f"Algo order missing algoId: {algo_order}")
                            continue
                        
                        # Check if this algo order is orphaned (no position)
                        if symbol not in symbols_with_positions:
                            try:
                                logger.info(f"Found orphan algo order {algo_id} for {symbol} (no open position). Canceling on reconcile.")
                                await asyncio.get_running_loop().run_in_executor(
                                    self.executor,
                                    partial(self.trading_client.futures_cancel_algo_order, symbol=symbol, algoId=algo_id)
                                )
                                logger.info(f"Canceled orphan algo order {algo_id} for {symbol}")
                            except Exception:
                                logger.exception(f"Failed to cancel orphan algo order {algo_id} for {symbol}")
                            continue
                        
                        # Classify as SL or TP based on order type
                        if 'STOP' in order_type or order_type == 'STOP_MARKET':
                            # This is a Stop Loss order
                            self.sl_orders[symbol] = algo_id
                            logger.info(f"Reconciled SL algo order for {symbol}: algoId={algo_id}")
                        elif 'TAKE_PROFIT' in order_type or order_type == 'TAKE_PROFIT_MARKET':
                            # This is a Take Profit order
                            if symbol not in self.tp_orders:
                                self.tp_orders[symbol] = {}
                            
                            # Assign to TP1/TP2 slots based on availability
                            if 'TP1' not in self.tp_orders[symbol]:
                                self.tp_orders[symbol]['TP1'] = algo_id
                                logger.info(f"Reconciled TP1 algo order for {symbol}: algoId={algo_id}")
                            elif 'TP2' not in self.tp_orders[symbol]:
                                self.tp_orders[symbol]['TP2'] = algo_id
                                logger.info(f"Reconciled TP2 algo order for {symbol}: algoId={algo_id}")
                            else:
                                logger.warning(f"Found extra TP algo order for {symbol}: algoId={algo_id}. Ignoring.")
                        else:
                            logger.warning(f"Unknown algo order type '{order_type}' for {symbol}: algoId={algo_id}")
                            
                    except Exception:
                        logger.exception(f"Error processing algo order {algo_order}")
                        
            except Exception as e:
                logger.error(f"Failed to fetch or process algo orders during reconciliation: {e}")
                logger.warning("Continuing without algo order reconciliation - protective orders may be missing from state!")


            # Persist reconciled state
            await self._save_state()
            logger.info("Reconciled bot state with exchange and saved to disk")

            # Start monitoring tasks for any live positions we discovered so the bot
            # continues monitoring TP/SL after restart.
            if self.position_info:
                try:
                    exchange_info = await asyncio.get_running_loop().run_in_executor(self.executor, self.monitoring_client.futures_exchange_info)
                    for symbol, info in list(self.position_info.items()):
                        # Do not start duplicate monitor tasks
                        if symbol in self.monitor_tasks and not self.monitor_tasks[symbol].done():
                            continue

                        # Determine direction and opposite_side from positionAmt via futures_position_information lookup
                        try:
                            pos_list = positions
                            pos = next((p for p in pos_list if p['symbol'] == symbol), None)
                            if not pos:
                                logger.warning(f"Could not find live position details for {symbol} to start monitor")
                                continue
                            position_amt = float(pos.get('positionAmt', 0))
                            direction = SIDE_BUY if position_amt > 0 else SIDE_SELL
                            opposite_side = SIDE_SELL if direction == SIDE_BUY else SIDE_BUY

                            symbol_info = next((s for s in exchange_info['symbols'] if s['symbol'] == symbol), None)
                            if not symbol_info:
                                logger.warning(f"Symbol info not found for {symbol}. Skipping monitor start.")
                                continue

                            qty_precision = symbol_info['quantityPrecision']
                            price_precision = symbol_info['pricePrecision']
                            tick_size = float(next(f['tickSize'] for f in symbol_info['filters'] if f['filterType'] == 'PRICE_FILTER'))

                            asyncio.create_task(self.monitor_market_position(
                                symbol=symbol,
                                direction=direction,
                                opposite_side=opposite_side,
                                avg_entry_price=info.get('avg_entry_price', 0.0),
                                total_qty=info.get('total_qty', 0.0),
                                qty_precision=qty_precision,
                                price_precision=price_precision,
                                tick_size=tick_size,
                                leverage=int(pos.get('leverage', self.default_leverage))  # Get leverage from position or use default
                            ))
                            logger.info(f"Started monitoring for existing position: {symbol}")
                        except Exception:
                            logger.exception(f"Failed to start monitor for {symbol}")
                except Exception:
                    logger.exception("Failed to start monitoring tasks for reconciled positions")
        except Exception as e:
            logger.exception(f"Failed to reconcile state with exchange: {e}")

async def main():
    # --- License Check ---
    try:
        logger.info("Validating license...")
        license_info = validate_license()
        # License sets MAXIMUM limits, .env can be lower
        env_max_positions = int(os.getenv('MAX_OPEN_POSITIONS', 1))
        max_positions_license = license_info.get('metadata', {}).get('max_open_positions', 2)
        actual_max_positions = min(env_max_positions, max_positions_license)
        os.environ['MAX_OPEN_POSITIONS'] = str(actual_max_positions)
        logger.info(f"MAX_OPEN_POSITIONS: {actual_max_positions} (env: {env_max_positions}, license limit: {max_positions_license})")
        
        env_max_leverage = int(os.getenv('MAX_LEVERAGE', 20))
        max_leverage_license = license_info.get('metadata', {}).get('max_leverage', 20)
        actual_max_leverage = min(env_max_leverage, max_leverage_license)
        os.environ['MAX_LEVERAGE'] = str(actual_max_leverage)
        logger.info(f"MAX_LEVERAGE: {actual_max_leverage} (env: {env_max_leverage}, license limit: {max_leverage_license})")

    except LicenseError as e:
        logger.error(f"LICENSE VALIDATION FAILED: {e}")
        logger.error("Bot cannot start without a valid license. Exiting.")
        # input("Press Enter to exit...") # Optional: Keep console open
        return # Exit the main function
    except Exception as e:
        logger.error(f"An unexpected error occurred during license check: {e}", exc_info=True)
        logger.error("Bot cannot start due to an error during license check. Exiting.")
        return
    # --- End License Check ---

    # Proceed with bot initialization only if license is valid
    bot = TelegramSignalBot(
        api_id=int(os.getenv('API_ID')),
        api_hash=os.getenv('API_HASH'),
        phone=os.getenv('PHONE_NUMBER'),
        channel_id=os.getenv('CHANNEL_ID'),
        trading_api_key=os.getenv('TRADING_API_KEY'),
        trading_secret_key=os.getenv('TRADING_SECRET_KEY'),
        monitoring_api_key=os.getenv('MONITORING_API_KEY'),
        monitoring_secret_key=os.getenv('MONITORING_SECRET_KEY'),
        trading_mode=os.getenv('TRADING_MODE')
    )
    try:
        await bot.connect()
        await bot.listen_to_channel()
    except KeyboardInterrupt:
        logger.info("Script interrupted by user. Shutting down without canceling pending orders.")
    except Exception as e:
        logger.error(f"Bot failed: {e}")
    finally:
        await bot.cleanup()

if __name__ == "__main__":
    asyncio.run(main())
